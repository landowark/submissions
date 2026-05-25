from openpyxl import Workbook, load_workbook
import pytest
from test_manager_basic import managers
from pathlib import Path
from backend.db.models import Procedure
from backend.validators.pydant import PydProcedureType
from tests.resources.custom_resources import DatabaseTestCase


@pytest.fixture(scope="function")
def db():
    tc = DatabaseTestCase()
    tc.setUp()
    yield tc
    try:
        tc.tearDown()
    except Exception:
        pass


@pytest.fixture(scope="function")
def proceduremanager(db):
    procedure = Procedure.query(limit=1)
    return managers.DefaultProcedureManager(None, input_object=procedure, proceduretype="Test ProcedureType")


@pytest.fixture(scope="function")
def construct_from_worksheet(db):
    p = Path(__file__).parents[2] / "resources" / "226C4100.xlsx"
    p = p.absolute()
    wb = load_workbook(p)
    assert isinstance(wb, Workbook)
    return managers.DefaultProcedureManager(None, input_object=wb['Test ProcedureTy Quality'], proceduretype="Test ProcedureType")


def test_construction_from_sql(proceduremanager):
    assert isinstance(proceduremanager.proceduretype, PydProcedureType)
    assert proceduremanager.proceduretype.name == "Test ProcedureType"
    proceduremanager = managers.DefaultProcedureManager(None, input_object=Procedure.query(limit=1))
    assert isinstance(proceduremanager.proceduretype, PydProcedureType)
    assert proceduremanager.proceduretype.name == "Test ProcedureType"
    # assert isinstance(procedure.pyd, PydProcedure)


def test_construction_from_worksheet(construct_from_worksheet):
    
    assert isinstance(construct_from_worksheet.proceduretype, PydProcedureType)
    assert construct_from_worksheet.proceduretype.name == "Test ProcedureType"
    

def test_write(proceduremanager):
    workbook = proceduremanager.write(Workbook())
    assert isinstance(workbook, Workbook)
    assert "Test ProcedureType Quality" in workbook.sheetnames
    ws = workbook['Test ProcedureType Quality']
    assert ws.cell(1,1).value == "Procedure Type"
    assert ws.cell(13, 3).value == "Lot"
    assert ws.cell(17, 1).value == "Test EquipmentRole"


# def test_construction_from_pyd(clientsubmission):
#     clientsubmission = clientsubmission.to_pydantic()
#     assert clientsubmission.submissiontype.get("value") == "Default SubmissionType"
#     clientmanager = managers.DefaultClientSubmissionManager(None, input_object=clientsubmission)
#     assert isinstance(clientmanager.submissiontype, SubmissionType)
#     assert clientmanager.submissiontype.name == "Default SubmissionType"
#     assert isinstance(clientmanager.pyd, PydClientSubmission)


# def test_write(clientsubmission):
#     clientmanager = managers.DefaultClientSubmissionManager(None, input_object=clientsubmission)
#     wb = Workbook()
#     workbook = clientmanager.write(wb)
#     assert isinstance(workbook, Workbook)
#     assert "Client Info" in workbook.sheetnames
#     ws = workbook['Client Info']
#     assert ws.cell(1,1).value == "Submitter Info"
#     assert ws.cell(13, 3).value == "Row"

