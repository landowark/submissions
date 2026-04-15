from openpyxl import Workbook
import pytest
from test_manager_basic import managers
from pathlib import Path
from backend.db.models import SubmissionType, ClientSubmission
from backend.validators.pydant import PydClientSubmission
from tests.resources.custom_resources import DatabaseTestCase


@pytest.fixture(scope="function", autouse=True)
def db():
    tc = DatabaseTestCase()
    tc.setUp()
    yield tc
    try:
        tc.tearDown()
    except Exception:
        pass


@pytest.fixture(scope="function")
def clientsubmission(db):
    return ClientSubmission.query(limit=1)

@pytest.fixture(scope="function")
def construct_from_excel():
    io_ = Path(r"tests\resources\226C4100.xlsx").absolute()
    return managers.DefaultClientSubmissionManager(None, input_object=io_)

@pytest.fixture(scope="function")
def construct_from_sql(clientsubmission):
    return managers.DefaultClientSubmissionManager(None, input_object=clientsubmission)


def test_construction_from_excel(construct_from_excel):
    clientmanager = construct_from_excel
    assert isinstance(clientmanager.submissiontype, SubmissionType)
    assert clientmanager.submissiontype.name == "Default SubmissionType"
    assert isinstance(clientmanager.pyd, PydClientSubmission)


def test_construction_from_sql(construct_from_sql):
    # assert clientsubmission.submissiontype.name == "Default SubmissionType"
    # clientmanager = managers.DefaultClientSubmissionManager(None, input_object=clientsubmission)
    assert isinstance(construct_from_sql.submissiontype, SubmissionType)
    assert construct_from_sql.submissiontype.name == "Default SubmissionType"
    assert isinstance(construct_from_sql.pyd, PydClientSubmission)


def test_construction_from_pyd(clientsubmission):
    clientsubmission = clientsubmission.to_pydantic()
    assert clientsubmission.submissiontype.get("value") == "Default SubmissionType"
    clientmanager = managers.DefaultClientSubmissionManager(None, input_object=clientsubmission)
    assert isinstance(clientmanager.submissiontype, SubmissionType)
    assert clientmanager.submissiontype.name == "Default SubmissionType"
    assert isinstance(clientmanager.pyd, PydClientSubmission)


def test_write(construct_from_sql):
    # clientmanager = managers.DefaultClientSubmissionManager(None, input_object=clientsubmission)
    # wb = Workbook()
    workbook = construct_from_sql.write(Workbook())
    assert isinstance(workbook, Workbook)
    assert "Client Info" in workbook.sheetnames
    ws = workbook['Client Info']
    assert ws.cell(1,1).value == "Submitter Info"
    # assert ws.cell(13, 3).value == "Row"


def test_find_procedures(construct_from_excel):
    clientmanager = construct_from_excel
    assert isinstance(clientmanager.input_object, Workbook)
    procedures = clientmanager.found_procedures
    assert "Test ProcedureTy Quality" in procedures