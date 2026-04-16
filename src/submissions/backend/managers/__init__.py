"""
Module for manager defaults.
"""
from copy import deepcopy
import logging
from pprint import pformat
from pathlib import Path
from frontend.widgets.functions import select_open_file
from tools import get_application_from_parent
from backend.validators import pydant
from backend.db.models import BaseClass
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import csv

logger = logging.getLogger(f"submissions.{__name__}")

class DefaultManager(object):

    """
    The job of the manager class is to convert all inputs into a Pydantic object for portability.
    This object will be stored as self.pyd
    """

    def __new__(cls, *args, **kwargs):
        """
        Is called before __init__. Ensures filepath is present.
        """
        try:
            input_object = kwargs.get('input_object') or args[1]
        except IndexError:
            input_object = None
        if isinstance(input_object, str):
            input_object = Path(input_object)
        if isinstance(input_object, Path):
            try:
                assert input_object.exists()
            except AssertionError:
                raise FileNotFoundError(f"File {input_object} does not exist.")
        instance = super().__new__(cls)
        instance.input_object = input_object
        return instance

    def __init__(self, parent, input_object: Path | str | pydant.PydBaseClass | BaseClass | Workbook | Worksheet | None = None, **kwargs):
        
        self.parent = parent
        self.input_object = input_object
        self.sheets = kwargs.get("sheets", None)
        if self.sheets is None:
            try:
                self.sheets = self.__class__.sheets
            except AttributeError:
                self.sheets = {}
        self.set_pyd()
        
    def set_pyd(self):
        if isinstance(self.input_object, str):
            self.input_object = Path(self.input_object)
        if isinstance(self.input_object, Path):
            self.input_object = self.input_object.absolute()
            filepath = deepcopy(self.input_object)
            if self.input_object.suffix == ".csv":
                self.input_object = self.csv2xlsx(self.input_object)
                if isinstance(self.input_object, tuple):
                    self.input_object = self.input_object[1]
            elif self.input_object.suffix == ".xlsx":
                self.input_object = load_workbook(self.input_object, data_only=True)
            else:
                raise TypeError(f"Unknown file type: {self.input_object.suffix}")
            self.input_object.file = filepath
        # NOTE: If input_object is a str or path, use parser to construct object
        match self.input_object:
            case Workbook() | Worksheet():
                self.pyd = self.parse()
            case x if issubclass(self.input_object.__class__, pydant.PydBaseClass):
                self.pyd = self.input_object
            case x if issubclass(self.input_object.__class__, BaseClass):
                self.pyd = self.input_object.to_pydantic()
            case _:
                logger.warning(f"Unmatched input object: {type(self.input_object)}. Looking for file.")
                if self.parent is not None:
                    # TODO: Allow for multiple filters. For now, just look for xlsx.
                    self.input_object = select_open_file(file_extension="xlsx", obj=get_application_from_parent(self.parent))
                    self.set_pyd()
                else:
                    raise ValueError(f"No parent, cannot get user input.")

    def parse(self):
        raise NotImplementedError("Parse only implemented in subclasses.")

    @property
    def _pyd_object(self):
        try:
            return getattr(pydant, f"Pyd{self.__class__.__name__.replace('Manager', '').replace('Default', '')}")
        except AttributeError as e:
            logger.error(
                f"Couldn't get pyd object: Pyd{self.__class__.__name__.replace('Manager', '').replace('Default', '')}, using {self.__class__.pyd_name}")
            try:
                return getattr(pydant, self.__class__.pyd_name)
            except AttributeError:
                logger.error(f"Couldn't get pyd object using pyd_name. Returning None")
                return None
        
    @classmethod
    def csv2xlsx(cls, filepath):
        wb = Workbook()
        ws = wb.active
        with open(filepath, "r") as f:
            reader = csv.reader(f, delimiter=",")
            for row in reader:
                ws.append(row)
        return wb, ws

    def to_pydantic(self):
        return self.pyd


    def get_worksheet(self, sheet: Worksheet | str | int = 0):
        match sheet:
            case Worksheet():
                return sheet
            case str():
                return self.input_object[sheet]
            case int():
                return self.input_object.worksheets[sheet - 1]
            case _:
                raise TypeError(f"Invalid type for worksheet retrieval: {type(sheet)}")

    def ratchet_start_row(self):
        """
        Changes start_row of sample_parser to end_row of info parser. Used when chaining parsers together.
        """
        output = []
        for sheet in self.info_parser.sheets:
            item = {k: v for k, v in sheet.items() if k != "end_row"}
            item['start_row'] = sheet.get("end_row", 1)
            output.append(item)
        return output


from .clientsubmissions import DefaultClientSubmissionManager
from .procedures import DefaultProcedureManager
from .results import DefaultResultsManager
from .runs import DefaultRunManager
