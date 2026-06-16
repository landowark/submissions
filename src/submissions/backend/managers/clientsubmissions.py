"""
Module for manager of ClientSubmission object
"""
from __future__ import annotations
from pprint import pformat
import logging, sys
from typing import TYPE_CHECKING, Generator
from pathlib import Path
from openpyxl.workbook import Workbook
from backend.validators import ClientSubmissionNamer
from backend.managers import DefaultManager
from backend.excel.parsers import clientsubmission_parser
from backend.excel.writers import clientsubmission_writer 
if TYPE_CHECKING:
    from backend.db.models import SubmissionType, ClientSubmission
    from backend.validators.pydant import PydClientSubmission, PydProcedure, PydRun

logger = logging.getLogger(f"submissions.{__name__}")


class DefaultClientSubmissionManager(DefaultManager):
   
    sheets = {
        "info":[dict(sheet="Client Info", start_row=1)],
        "sample":[dict(sheet="Client Info", start_row=1)]
        }

    def __init__(self, parent, submissiontype: SubmissionType | str | None = None,
                 input_object: Path | str | ClientSubmission | PydClientSubmission | None = None, **kwargs):
        from backend.db.models import SubmissionType, ClientSubmission
        from backend.validators.pydant import PydClientSubmission
        # NOTE: So the submissiontype schtick is mostly for a future incident in which I have to scrape 
        # specialized excel sheets.
        match input_object:
            case str() | Path() | Workbook():
                self.namer = ClientSubmissionNamer(filepath=input_object)
            case _ if isinstance(input_object, ClientSubmission):
                self.namer = input_object
            case _ if isinstance(input_object, PydClientSubmission):
                self.namer = input_object
            case _:
                logger.warning(f"Skipping submission type")
        if submissiontype is None:
            try:
                submissiontype = self.namer.submissiontype
            except Exception as e:
                raise TypeError(f"Unknown type for submissiontype of {type(submissiontype)}")
        match submissiontype:
            case str():
                submissiontype = SubmissionType.query(name=submissiontype)
            case dict():
                q = submissiontype.get("name") or submissiontype.get("value")
                submissiontype = SubmissionType.query(name=q)
            case SubmissionType():
                pass
            case _:
                # NOTE: if unmatched, try to get from input_object
                pass
        self.submissiontype = submissiontype
        super().__init__(parent=parent, input_object=input_object, **kwargs)
        if isinstance(self.input_object, Workbook):
            for procedure in self.scraped_procedures:
                run = next((item for item in self.clientsubmission.run if item.rsl_plate_number==procedure.run), None)
                if run is None:
                    run = self.construct_run(procedure=procedure)
                run.add_samples(procedure.sample)
                run.procedure.append(procedure)
                if run not in self.clientsubmission.run:
                    self.clientsubmission.run.append(run)

    @property
    def _pyd_object(self):
        from backend.validators.pydant import PydClientSubmission
        return PydClientSubmission

    def construct_run(self, procedure: PydProcedure) -> PydRun:
        from backend.validators.pydant import PydRun
        return PydRun(
            rsl_plate_number = procedure.run,
            clientsubmission = self.clientsubmission,
            started_date = procedure.started_date,
            completed_date = procedure.completed_date,
        )

    @property
    def scraped_procedures(self) -> Generator[PydProcedure, None, None]:
        from backend.db.models import ProcedureType
        from backend.managers.procedures import DefaultProcedureManager
        for procedure in self.found_procedures:
            proceduretype = procedure.strip(" Quality")
            proceduretype = ProcedureType.query(name=proceduretype)
            try:
                worksheet = self.input_object[procedure]
            except KeyError:
                continue
            manager = DefaultProcedureManager(parent=self.parent, input_object=worksheet, proceduretype=proceduretype)
            yield manager.to_pydantic()
            
    @property
    def found_procedures(self) -> Generator[str, None, None]:
        # At this point strings should be parsed into path
        from backend.db.models import ProcedureType
        if not isinstance(self.input_object, Workbook):
            yield from ()
        else:
            ptypes = [item.name for item in ProcedureType.query()]
            for sheet in self.input_object.sheetnames:
                # Check if the base name or specific extraction name matches ptypes
                for pt in ptypes:
                    if pt.startswith(sheet.removesuffix(' Quality')):
                        yield sheet
                
    def parse(self):
        from backend.validators.pydant import PydSample
        try:
            info_parser = getattr(clientsubmission_parser, f"{self.submissiontype.name}InfoParser")
        except AttributeError:
            info_parser = clientsubmission_parser.ClientSubmissionInfoParser
        info = {}
        for sheet in self.sheets['info']:
            ws = self.get_worksheet(sheet.get("sheet", 1))
            start_row = sheet.get("start_row", 1)
            self.info_parser = info_parser(worksheet=ws, start_row=start_row)
            info.update(self.info_parser.parsed_info)
            for s in self.sheets['sample']:
                if s['sheet'] == sheet['sheet']:
                    s['start_row'] = self.info_parser.end_row + 1
        self.clientsubmission = self._pyd_object(**info)
        # NOTE: Alter sheets List[dict] so that the start_row sent to sample parser is the end row of the info parser
        try:
            sample_parser = getattr(clientsubmission_parser, f"{self.submissiontype.name}InfoParser")
        except AttributeError:
            sample_parser = clientsubmission_parser.ClientSubmissionSampleParser
        samples = []
        for sheet in self.sheets['sample']:
            ws = self.get_worksheet(sheet.get("sheet", 1))
            start_row = sheet.get("start_row", 1)
            self.sample_parser = sample_parser(worksheet=ws, start_row=start_row, submitter_id=self.clientsubmission.submitter_plate_id.value)
            for sample in self.sample_parser.parsed_info:
                samples.append(sample)
        for sample in samples:
            try:
                self.clientsubmission.sample.append(PydSample(**sample))
            except Exception as e:
                logger.error(f"Couldn't add sample {sample} due to {e}")
                continue
        return self.clientsubmission
    
    def write(self, workbook: Workbook) -> Workbook:
        self.info_writer = clientsubmission_writer.ClientSubmissionInfoWriter(pydant_obj=self.pyd)
        assert isinstance(self.info_writer, clientsubmission_writer.ClientSubmissionInfoWriter)
        workbook = self.info_writer.write_to_workbook(workbook)
        self.sample_writer = clientsubmission_writer.ClientSubmissionSampleWriter(pydant_obj=self.pyd)
        workbook = self.sample_writer.write_to_workbook(workbook, start_row=self.info_writer.worksheet.max_row + 1)
        return workbook

__all__ = ["DefaultClientSubmissionManager"]