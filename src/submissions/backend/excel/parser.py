"""
contains parser objects for pulling values from client generated run sheets.
"""
import logging
from copy import copy
from getpass import getuser
from pprint import pformat
from typing import List
from openpyxl import load_workbook, Workbook
from pathlib import Path
from backend.db.models import *
from backend.validators import PydSubmission, RSLNamer
from collections import OrderedDict
from tools import check_not_nan, is_missing, check_key_or_attr

logger = logging.getLogger(f"submissions.{__name__}")


class SheetParser(object):
    """
    object to pull and contain data from excel file
    """

    def __init__(self, filepath: Path | None = None):
        """
        Args:
            filepath (Path | None, optional): file path to excel sheet. Defaults to None.
        """
        logger.info(f"\n\nParsing {filepath.__str__()}\n\n")
        match filepath:
            case Path():
                self.filepath = filepath
            case str():
                self.filepath = Path(filepath)
            case _:
                logger.error(f"No filepath given.")
                raise ValueError("No filepath given.")
        try:
            self.xl = load_workbook(filepath, data_only=True)
        except ValueError as e:
            logger.error(f"Incorrect value: {e}")
            raise FileNotFoundError(f"Couldn't parse file {self.filepath}")
        self.sub = OrderedDict()
        # NOTE: make decision about type of sample we have
        self.sub['submission_type'] = dict(value=RSLNamer.retrieve_submission_type(filename=self.filepath),
                                           missing=True)
        self.submission_type = SubmissionType.query(name=self.sub['submission_type'])
        self.sub_object = BasicRun.find_polymorphic_subclass(polymorphic_identity=self.submission_type)
        # NOTE: grab the info map from the run type in database
        self.parse_info()
        self.import_kit_validation_check()
        self.parse_reagents()
        self.parse_samples()
        self.parse_equipment()
        self.parse_tips()

    def parse_info(self):
        """
        Pulls basic information from the excel sheet
        """
        parser = InfoParser(xl=self.xl, submission_type=self.submission_type, sub_object=self.sub_object)
        self.info_map = parser.info_map
        # NOTE: in order to accommodate generic run types we have to check for the type in the excel sheet and rerun accordingly
        try:
            check = parser.parsed_info['submission_type']['value'] not in [None, "None", "", " "]
        except KeyError as e:
            logger.error(f"Couldn't check run type due to KeyError: {e}")
            return
        logger.info(
            f"Checking for updated run type: {self.submission_type.name} against new: {parser.parsed_info['submission_type']['value']}")
        if self.submission_type.name != parser.parsed_info['submission_type']['value']:
            if check:
                # NOTE: If initial run type doesn't match parsed run type, defer to parsed run type.
                self.submission_type = SubmissionType.query(name=parser.parsed_info['submission_type']['value'])
                logger.info(f"Updated self.submission_type to {self.submission_type}. Rerunning parse.")
                self.parse_info()
            else:
                self.submission_type = RSLNamer.retrieve_submission_type(filename=self.filepath)
                self.parse_info()
        for k, v in parser.parsed_info.items():
            self.sub.__setitem__(k, v)

    def parse_reagents(self, extraction_kit: str | None = None):
        """
        Calls reagent parser class to pull info from the excel sheet

        Args:
            extraction_kit (str | None, optional): Relevant extraction kit for reagent map. Defaults to None.
        """
        if extraction_kit is None:
            extraction_kit = self.sub['extraction_kit']
        parser = ReagentParser(xl=self.xl, submission_type=self.submission_type,
                               extraction_kit=extraction_kit)
        self.sub['reagents'] = parser.parsed_reagents

    def parse_samples(self):
        """
        Calls sample parser to pull info from the excel sheet
        """
        parser = SampleParser(xl=self.xl, submission_type=self.submission_type)
        self.sub['samples'] = parser.parsed_samples

    def parse_equipment(self):
        """
        Calls equipment parser to pull info from the excel sheet
        """
        parser = EquipmentParser(xl=self.xl, submission_type=self.submission_type)
        self.sub['equipment'] = parser.parsed_equipment

    def parse_tips(self):
        """
        Calls tips parser to pull info from the excel sheet
        """
        parser = TipParser(xl=self.xl, submission_type=self.submission_type)
        self.sub['tips'] = parser.parsed_tips

    def import_kit_validation_check(self):
        """
        Enforce that the parser has an extraction kit
        """
        if 'extraction_kit' not in self.sub.keys() or not check_not_nan(self.sub['extraction_kit']['value']):
            from frontend.widgets.pop_ups import ObjectSelector
            dlg = ObjectSelector(title="Kit Needed", message="At minimum a kit is needed. Please select one.",
                                 obj_type=KitType)
            if dlg.exec():
                self.sub['extraction_kit'] = dict(value=dlg.parse_form(), missing=True)
            else:
                raise ValueError("Extraction kit needed.")
        else:
            if isinstance(self.sub['extraction_kit'], str):
                self.sub['extraction_kit'] = dict(value=self.sub['extraction_kit'], missing=True)

    def to_pydantic(self) -> PydSubmission:
        """
        Generates a pydantic model of scraped data for validation

        Returns:
            PydSubmission: output pydantic model
        """
        return PydSubmission(filepath=self.filepath, run_custom=True, **self.sub)


class InfoParser(object):
    """
    Object to parse generic info from excel sheet.
    """

    def __init__(self, xl: Workbook, submission_type: str | SubmissionType, sub_object: BasicRun | None = None):
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (str | SubmissionType): Type of run expected (Wastewater, Bacterial Culture, etc.)
            sub_object (BasicRun | None, optional): Submission object holding methods. Defaults to None.
        """
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        if sub_object is None:
            sub_object = BasicRun.find_polymorphic_subclass(polymorphic_identity=submission_type.name)
        self.submission_type_obj = submission_type
        self.submission_type = dict(value=self.submission_type_obj.name, missing=True)
        self.sub_object = sub_object
        self.xl = xl

    @property
    def info_map(self) -> dict:
        """
        Gets location of basic info from the submission_type object in the database.

        Returns:
            dict: Location map of all info for this run type
        """
        # NOTE: Get the parse_info method from the run type specified
        return self.sub_object.construct_info_map(submission_type=self.submission_type_obj, mode="read")

    @property
    def parsed_info(self) -> dict:
        """
        Pulls basic info from the excel sheet.

        Returns:
            dict: key:value of basic info
        """
        dicto = {}
        # NOTE: This loop parses generic info
        for sheet in self.xl.sheetnames:
            ws = self.xl[sheet]
            relevant = []
            for k, v in self.info_map.items():
                # NOTE: If the value is hardcoded put it in the dictionary directly. Ex. Artic kit
                if k == "custom":
                    continue
                if isinstance(v, str):
                    dicto[k] = dict(value=v, missing=False)
                    continue
                for location in v:
                    try:
                        check = location['sheet'] == sheet
                    except TypeError:
                        logger.warning(f"Location is likely a string, skipping")
                        dicto[k] = dict(value=location, missing=False)
                        check = False
                    if check:
                        new = location
                        new['name'] = k
                        relevant.append(new)
            # NOTE: make sure relevant is not an empty list.
            if not relevant:
                continue
            for item in relevant:
                # NOTE: Get cell contents at this location
                value = ws.cell(row=item['row'], column=item['column']).value
                match item['name']:
                    case "submission_type":
                        value, missing = is_missing(value)
                        value = value.title()
                    case "submitted_date":
                        value, missing = is_missing(value)
                    # NOTE: is field a JSON? Includes: Extraction info, PCR info, comment, custom
                    case thing if thing in self.sub_object.jsons:
                        value, missing = is_missing(value)
                        if missing: continue
                        value = dict(name=f"Parser_{sheet}", text=value, time=datetime.now())
                        try:
                            dicto[item['name']]['value'] += value
                            continue
                        except KeyError:
                            logger.error(f"New value for {item['name']}")
                    case _:
                        value, missing = is_missing(value)
                if item['name'] not in dicto.keys():
                    try:
                        dicto[item['name']] = dict(value=value, missing=missing)
                    except (KeyError, IndexError):
                        continue
        # NOTE: Return after running the parser components held in run object.
        return self.sub_object.custom_info_parser(input_dict=dicto, xl=self.xl, custom_fields=self.info_map['custom'])


class ReagentParser(object):
    """
    Object to pull reagents from excel sheet.
    """

    def __init__(self, xl: Workbook, submission_type: str | SubmissionType, extraction_kit: str,
                 run_object: BasicRun | None = None):
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (str|SubmissionType): Type of run expected (Wastewater, Bacterial Culture, etc.)
            extraction_kit (str): Extraction kit used.
            run_object (BasicRun | None, optional): Submission object holding methods. Defaults to None.
        """
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        self.submission_type_obj = submission_type
        if not run_object:
            run_object = submission_type.submission_class
        self.run_object = run_object
        if isinstance(extraction_kit, dict):
            extraction_kit = extraction_kit['value']
        self.kit_object = KitType.query(name=extraction_kit)
        self.xl = xl

    @property
    def kit_map(self) -> dict:
        """
        Gets location of kit reagents from database

        Args:
            submission_type (str): Name of run type.

        Returns:
            dict: locations of reagent info for the kit.
        """
        associations, self.kit_object = self.kit_object.construct_xl_map_for_use(submission_type=self.submission_type_obj)
        reagent_map = {k: v for k, v in associations.items() if k != 'info'}
        try:
            del reagent_map['info']
        except KeyError:
            pass
        return reagent_map

    @property
    def parsed_reagents(self) -> Generator[dict, None, None]:
        """
        Extracts reagent information from the Excel form.

        Returns:
            List[PydReagent]: List of parsed reagents.
        """
        for sheet in self.xl.sheetnames:
            ws = self.xl[sheet]
            relevant = {k.strip(): v for k, v in self.kit_map.items() if sheet in self.kit_map[k]['sheet']}
            if not relevant:
                continue
            for item in relevant:
                try:
                    reagent = relevant[item]
                    name = ws.cell(row=reagent['name']['row'], column=reagent['name']['column']).value
                    lot = ws.cell(row=reagent['lot']['row'], column=reagent['lot']['column']).value
                    expiry = ws.cell(row=reagent['expiry']['row'], column=reagent['expiry']['column']).value
                    if 'comment' in relevant[item].keys():
                        comment = ws.cell(row=reagent['comment']['row'], column=reagent['comment']['column']).value
                    else:
                        comment = ""
                except (KeyError, IndexError):
                    yield dict(role=item.strip(), lot=None, expiry=None, name=None, comment="", missing=True)
                # NOTE: If the cell is blank tell the PydReagent
                if check_not_nan(lot):
                    missing = False
                else:
                    missing = True
                lot = str(lot)
                try:
                    check = name.lower() != "not applicable"
                except AttributeError:
                    logger.warning(f"name is not a string.")
                    check = True
                if check:
                    yield dict(role=item.strip(), lot=lot, expiry=expiry, name=name, comment=comment,
                               missing=missing)


class SampleParser(object):
    """
    Object to pull data for samples in excel sheet and construct individual sample objects
    """

    def __init__(self, xl: Workbook, submission_type: SubmissionType, sample_map: dict | None = None,
                 sub_object: BasicRun | None = None) -> None:
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (SubmissionType): Type of run expected (Wastewater, Bacterial Culture, etc.)
            sample_map (dict | None, optional): Locations in database where samples are found. Defaults to None.
            sub_object (BasicRun | None, optional): Submission object holding methods. Defaults to None.
        """
        self.samples = []
        self.xl = xl
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        self.submission_type = submission_type.name
        self.submission_type_obj = submission_type
        if sub_object is None:
            logger.warning(
                f"Sample parser attempting to fetch run class with polymorphic identity: {self.submission_type}")
            sub_object = BasicRun.find_polymorphic_subclass(polymorphic_identity=self.submission_type)
        self.sub_object = sub_object
        self.sample_type = self.sub_object.get_default_info("sample_type", submission_type=submission_type)
        self.samp_object = BasicSample.find_polymorphic_subclass(polymorphic_identity=self.sample_type)

    @property
    def sample_map(self) -> dict:
        """
        Gets info locations in excel book for run type.

        Args:
            submission_type (str): run type

        Returns:
            dict: Info locations.
        """

        return self.sub_object.construct_sample_map(submission_type=self.submission_type_obj)

    @property
    def plate_map_samples(self) -> List[dict]:
        """
        Parse sample location/name from plate map

        Returns:
            List[dict]: List of sample ids and locations.
        """
        invalids = [0, "0", "EMPTY"]
        smap = self.sample_map['plate_map']
        ws = self.xl[smap['sheet']]
        plate_map_samples = []
        for ii, row in enumerate(range(smap['start_row'], smap['end_row'] + 1), start=1):
            for jj, column in enumerate(range(smap['start_column'], smap['end_column'] + 1), start=1):
                id = str(ws.cell(row=row, column=column).value)
                if check_not_nan(id):
                    if id not in invalids:
                        sample_dict = dict(id=id, row=ii, column=jj)
                        sample_dict['sample_type'] = self.sample_type
                        plate_map_samples.append(sample_dict)
                    else:
                        pass
                else:
                    pass
        return plate_map_samples

    @property
    def lookup_samples(self) -> List[dict]:
        """
        Parse misc info from lookup table.

        Returns:
            List[dict]: List of basic sample info.
        """

        lmap = self.sample_map['lookup_table']
        ws = self.xl[lmap['sheet']]
        lookup_samples = []
        for ii, row in enumerate(range(lmap['start_row'], lmap['end_row'] + 1), start=1):
            row_dict = {k: ws.cell(row=row, column=v).value for k, v in lmap['sample_columns'].items()}
            try:
                row_dict[lmap['merge_on_id']] = str(row_dict[lmap['merge_on_id']])
            except KeyError:
                pass
            row_dict['sample_type'] = self.sample_type
            row_dict['submission_rank'] = ii
            try:
                check = check_not_nan(row_dict[lmap['merge_on_id']])
            except KeyError:
                check = False
            if check:
                lookup_samples.append(self.samp_object.parse_sample(row_dict))
        return lookup_samples

    @property
    def parsed_samples(self) -> Generator[dict, None, None]:
        """
        Merges sample info from lookup table and plate map.

        Returns:
            List[dict]: Reconciled samples
        """
        if not self.plate_map_samples or not self.lookup_samples:
            logger.warning(f"No separate samples")
            samples = self.lookup_samples or self.plate_map_samples
            for new in samples:
                if not check_key_or_attr(key='submitter_id', interest=new, check_none=True):
                    new['submitter_id'] = new['id']
                new = self.sub_object.parse_samples(new)
                try:
                    del new['id']
                except KeyError:
                    pass
                yield new
        else:
            merge_on_id = self.sample_map['lookup_table']['merge_on_id']
            logger.info(f"Merging sample info using {merge_on_id}")
            plate_map_samples = sorted(copy(self.plate_map_samples), key=itemgetter('id'))
            lookup_samples = sorted(copy(self.lookup_samples), key=itemgetter(merge_on_id))
            for ii, psample in enumerate(plate_map_samples):
                # NOTE: See if we can do this the easy way and just use the same list index.
                try:
                    check = psample['id'] == lookup_samples[ii][merge_on_id]
                except (KeyError, IndexError):
                    check = False
                if check:
                    new = lookup_samples[ii] | psample
                    lookup_samples[ii] = {}
                else:
                    logger.warning(f"Match for {psample['id']} not direct, running search.")
                    searchables = [(jj, sample) for jj, sample in enumerate(lookup_samples)
                                   if merge_on_id in sample.keys()]
                    jj, new = next(((jj, lsample | psample) for jj, lsample in searchables
                                    if lsample[merge_on_id] == psample['id']), (-1, psample))
                    if jj >= 0:
                        lookup_samples[jj] = {}
                if not check_key_or_attr(key='submitter_id', interest=new, check_none=True):
                    new['submitter_id'] = psample['id']
                new = self.sub_object.parse_samples(new)
                try:
                    del new['id']
                except KeyError:
                    pass
                yield new


class EquipmentParser(object):
    """
    Object to pull data for equipment in excel sheet
    """

    def __init__(self, xl: Workbook, submission_type: str | SubmissionType) -> None:
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (str | SubmissionType): Type of run expected (Wastewater, Bacterial Culture, etc.)
        """
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        self.submission_type = submission_type
        self.xl = xl

    @property
    def equipment_map(self) -> dict:
        """
        Gets the map of equipment locations in the run type's spreadsheet

        Returns:
            List[dict]: List of locations
        """
        return {k: v for k, v in self.submission_type.construct_field_map("equipment")}

    def get_asset_number(self, input: str) -> str:
        """
        Pulls asset number from string.

        Args:
            input (str): String to be scraped

        Returns:
            str: asset number
        """
        regex = Equipment.manufacturer_regex
        try:
            return regex.search(input).group().strip("-")
        except AttributeError as e:
            logger.error(f"Error getting asset number for {input}: {e}")
            return input

    @property
    def parsed_equipment(self) -> Generator[dict, None, None]:
        """
        Scrapes equipment from xl sheet

        Returns:
            List[dict]: list of equipment
        """
        for sheet in self.xl.sheetnames:
            ws = self.xl[sheet]
            try:
                relevant = {k: v for k, v in self.equipment_map.items() if v['sheet'] == sheet}
            except (TypeError, KeyError) as e:
                logger.error(f"Error creating relevant equipment list: {e}")
                continue
            previous_asset = ""
            for k, v in relevant.items():
                asset = ws.cell(v['name']['row'], v['name']['column']).value
                if not check_not_nan(asset):
                    asset = previous_asset
                else:
                    previous_asset = asset
                asset = self.get_asset_number(input=asset)
                eq = Equipment.query(asset_number=asset)
                if eq is None:
                    eq = Equipment.query(name=asset)
                process = ws.cell(row=v['process']['row'], column=v['process']['column']).value
                try:
                    yield dict(name=eq.name, processes=[process], role=k, asset_number=eq.asset_number,
                               nickname=eq.nickname)
                except AttributeError:
                    logger.error(f"Unable to add {eq} to list.")
                    continue


class TipParser(object):
    """
    Object to pull data for tips in excel sheet
    """

    def __init__(self, xl: Workbook, submission_type: str | SubmissionType) -> None:
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (str | SubmissionType): Type of run expected (Wastewater, Bacterial Culture, etc.)
        """
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        self.submission_type = submission_type
        self.xl = xl

    @property
    def tip_map(self) -> dict:
        """
        Gets the map of equipment locations in the run type's spreadsheet

        Returns:
            List[dict]: List of locations
        """
        return {k: v for k, v in self.submission_type.construct_field_map("tip")}

    @property
    def parsed_tips(self) -> Generator[dict, None, None]:
        """
        Scrapes equipment from xl sheet

        Returns:
            List[dict]: list of equipment
        """
        for sheet in self.xl.sheetnames:
            ws = self.xl[sheet]
            try:
                relevant = {k: v for k, v in self.tip_map.items() if v['sheet'] == sheet}
            except (TypeError, KeyError) as e:
                logger.error(f"Error creating relevant equipment list: {e}")
                continue
            previous_asset = ""
            for k, v in relevant.items():
                asset = ws.cell(v['name']['row'], v['name']['column']).value
                if "lot" in v.keys():
                    lot = ws.cell(v['lot']['row'], v['lot']['column']).value
                else:
                    lot = None
                if not check_not_nan(asset):
                    asset = previous_asset
                else:
                    previous_asset = asset
                eq = Tips.query(lot=lot, name=asset, limit=1)
                try:
                    yield dict(name=eq.name, role=k, lot=lot)
                except AttributeError:
                    logger.error(f"Unable to add {eq} to PydTips list.")


class PCRParser(object):
    """Object to pull data from Design and Analysis PCR export file."""

    def __init__(self, filepath: Path | None = None, submission: BasicRun | None = None) -> None:
        """
        Args:
            filepath (Path | None, optional): file to parse. Defaults to None.
            submission (BasicRun | None, optional): Submission parsed data to be added to.
        """
        if filepath is None:
            logger.error('No filepath given.')
            self.xl = None
        else:
            try:
                self.xl = load_workbook(filepath)
            except ValueError as e:
                logger.error(f'Incorrect value: {e}')
                self.xl = None
            except PermissionError:
                logger.error(f"Couldn't get permissions for {filepath.__str__()}. Operation might have been cancelled.")
                return None
        if submission is None:
            self.submission_obj = Wastewater
            rsl_plate_num = None
        else:
            self.submission_obj = submission
            rsl_plate_num = self.submission_obj.rsl_plate_num
        self.samples = self.submission_obj.parse_pcr(xl=self.xl, rsl_plate_num=rsl_plate_num)
        self.controls = self.submission_obj.parse_pcr_controls(xl=self.xl, rsl_plate_num=rsl_plate_num)

    @property
    def pcr_info(self) -> dict:
        """
        Parse general info rows for all types of PCR results
        """
        info_map = self.submission_obj.get_submission_type().sample_map['pcr_general_info']
        sheet = self.xl[info_map['sheet']]
        iter_rows = sheet.iter_rows(min_row=info_map['start_row'], max_row=info_map['end_row'])
        pcr = {}
        for row in iter_rows:
            try:
                key = row[0].value.lower().replace(' ', '_')
            except AttributeError as e:
                logger.error(f"No key: {row[0].value} due to {e}")
                continue
            value = row[1].value or ""
            pcr[key] = value
        pcr['imported_by'] = getuser()
        return pcr


class ConcentrationParser(object):

    def __init__(self, filepath: Path | None = None, run: BasicRun | None = None) -> None:
        if filepath is None:
            logger.error('No filepath given.')
            self.xl = None
        else:
            try:
                self.xl = load_workbook(filepath)
            except ValueError as e:
                logger.error(f'Incorrect value: {e}')
                self.xl = None
            except PermissionError:
                logger.error(f"Couldn't get permissions for {filepath.__str__()}. Operation might have been cancelled.")
                return None
        if run is None:
            self.submission_obj = BasicRun()
            rsl_plate_num = None
        else:
            self.submission_obj = run
            rsl_plate_num = self.submission_obj.rsl_plate_num
        self.samples = self.submission_obj.parse_concentration(xl=self.xl, rsl_plate_num=rsl_plate_num)

# NOTE: Generified parsers below

class InfoParserV2(object):
    """
    Object for retrieving submitter info from sample list sheet
    """

    default_range = dict(
        start_row=2,
        end_row=18,
        start_column=7,
        end_column=8,
        sheet="Sample List"
    )

