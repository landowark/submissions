"""
Default Parser archetypes.
"""
from __future__ import annotations
from logging import getLogger
logger = getLogger(f"submissions.{__name__}")
from re import sub as rsub
from typing import Generator
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame


class DefaultParser(object):

    range_dict = dict(start_row = 1)

    def __repr__(self):
        return f"{self.__class__.__name__}<{self.worksheet.title}>"

    def __init__(self, worksheet: Worksheet, start_row: int = 1, end_row: int | None = None, *args, **kwargs):
        """

        Args:
            filepath (Path|str): Must be given as a kwarg. eg. filepath=X
            procedure ():
            range_dict ():
            *args ():
            **kwargs ():
        """
        logger.info(f"\n\nHello from {self.__class__.__name__}\n\n")
        self.worksheet = worksheet
        self.start_row = self.delineate_start_row(worksheet=worksheet, start_row=start_row)
        if end_row is None:
            self.end_row = self.delineate_end_row(worksheet=worksheet, start_row=self.start_row)
        else:
            self.end_row = self.delineate_end_row(worksheet=worksheet, start_row=end_row)
        assert self.start_row <= self.end_row
        
    @classmethod
    def delineate_start_row(cls, worksheet: Worksheet, start_row: int = 1) -> int:
        """
        Determines the start row by finding the first non-empty row.

        Returns:
            int: Start row number
        """
        for iii, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
            if not all([item.value is None for item in row]):
                return iii
        return worksheet.min_row

    @classmethod
    def delineate_end_row(cls, worksheet: Worksheet, start_row: int = 1) -> int:
        """
        Determines the end row by finding the first empty row.

        Returns:
            int: End row number
        """
        for iii, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
            if all([item.value is None for item in row]):
                return iii
        return worksheet.max_row + 1
    
    @staticmethod
    def fix_key(key: str) -> str | None:
        key = rsub(r"\(.*\)", "", key)
        key = rsub(r"\s+", "_", key.lower().replace(":", "").strip())
        if key.count("_") > 3:
            logger.warning(f"There are more than 3 spaces in {key}, skipping")
            return None
        match key:
            case "comments":
                return "comment"
            case _:
                return key


class DefaultKEYVALUEParser(DefaultParser):

    @property
    def parsed_info(self) -> Generator[tuple, None, None]:
        """
        Generates key, value tuples for rows in an excel sheet.

        Returns:
            Generator[tuple, None, None]: (key, value) tuple.
        """
        rows = range(self.start_row, self.end_row)
        for row in rows:
            check_row = [item for item in self.worksheet.rows][row-1]
            if any([isinstance(cell, MergedCell) for cell in check_row]):
                continue
            key = self.worksheet.cell(row, 1).value
            if key:
                key = self.fix_key(key)
                # NOTE: If there are more than 3 spaces in the key, continue
                if not key:
                    continue
                value = self.worksheet.cell(row, 2).value
                missing = False if value else True
                value = dict(value=value, missing=missing)
                yield key, value

    def to_pydantic(self):
        return self._pyd_object({k:v for k,v in self.parsed_info})


class DefaultTABLEParser(DefaultParser):

    @property
    def parsed_info(self) -> Generator[dict, None, None]:
        """
        Generates dictionaries of data from Excel rows.

        Returns:
            Generator[dict, None, None]: {column_header: row column value}
        """
        rows = list(self.worksheet.iter_rows(
            min_row=self.start_row,
            max_row=self.end_row - 1,
            values_only=True
        ))
        if not rows:
            return
        df = DataFrame(rows[1:], columns=rows[0])
        df = df.dropna(axis=1, how='all')
        for row in df.iterrows():
            output = {}
            for key, value in row[1].to_dict().items():
                if isinstance(key, str):
                    key = self.fix_key(key)
                if key:    
                    output[key] = value
            yield output

    def to_pydantic(self, **kwargs):
        return [self._pyd_object(**output) for output in self.parsed_info]


from .procedure_parsers import *
from .results_parsers import *
from .clientsubmission_parser import *

__all__ = ["DefaultKEYVALUEParser", "DefaultTABLEParser", "ProcedureInfoParser", "ProcedureSampleParser", "ProcedureReagentParser", "ProcedureEquipmentParser",
           "DefaultResultsInfoParser", "DefaultResultsSampleParser", "DiomniPCRInfoParser", "DiomniPCRSampleParser", "QubitInfoParser", "QubitSampleParser",
           "ClientSubmissionInfoParser", "ClientSubmissionSampleParser"]