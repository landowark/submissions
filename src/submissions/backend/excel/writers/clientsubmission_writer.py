"""
Module for ClientSubmission writing
"""
from __future__ import annotations
from openpyxl.cell import MergedCell
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from . import DefaultKEYVALUEWriter, DefaultTABLEWriter


class ClientSubmissionInfoWriter(DefaultKEYVALUEWriter):

    def __init__(self, pydant_obj, *args, **kwargs):
        super().__init__(pydant_obj=pydant_obj, *args, **kwargs)
        # Add comment back due to exclusion
        self.fill_dictionary['comment'] = pydant_obj.comment

    def prewrite(self, worksheet: Worksheet, start_row: int) -> Worksheet:
        worksheet.cell(row=start_row, column=1, value="Submitter Info")
        worksheet.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
        worksheet.cell(row=start_row, column=1).font = Font(bold=True, color="FFFFFF")
        worksheet.cell(row=start_row, column=1).fill = PatternFill(start_color='376589', end_color='376589', fill_type="solid")
        worksheet.cell(row=start_row, column=2).fill = PatternFill(start_color='376589', end_color='376589', fill_type="solid")
        return worksheet


class ClientSubmissionSampleWriter(DefaultTABLEWriter):

    def __init__(self, pydant_obj, *args, **kwargs):
        self.submissiontype = pydant_obj.submissiontype
        super().__init__(pydant_obj=pydant_obj, *args, **kwargs)
        self.pydant_obj = self.pad_submission_samples_to_length()
        self.excluded = self.pydant_obj[0].class_config.excluded
        self.key_value_order = self.pydant_obj[0].class_config.key_value_order

    def write_to_workbook(self, workbook: Workbook, sheet: str | None = None,
                          start_row: int | None = None, *args, **kwargs) -> Workbook:
        
        workbook = super().write_to_workbook(workbook=workbook, sheet=sheet, start_row=start_row, *args, **kwargs)
        return workbook

    def postwrite(self, worksheet: Worksheet, **kwargs) -> Worksheet:
        worksheet = super().postwrite(worksheet, **kwargs)
        for row in worksheet.iter_rows(min_row=self.start_row, max_row=self.end_row):
            for cell in row:
                if cell.value in [0, "0", "None"]:
                    if isinstance(cell, MergedCell):
                        continue
                    else:
                        cell.value = ""
                cell.alignment = Alignment(horizontal="center")
        return worksheet
    
    def pad_submission_samples_to_length(self):
        from backend.validators.pydant import PydClientSubmissionSampleAssociation
        output_samples = []
        rng = self.pydant_obj.max_sample_rank + 1
        for iii in range(1, rng):
            iterator = self.pydant_obj.sql_instance.clientsubmissionsampleassociation
            try:
                sample = next(item.to_pydantic() for item in iterator if item.submission_rank == iii)
            except StopIteration:
                sample = PydClientSubmissionSampleAssociation(sample="", clientsubmission=self.pydant_obj.name, submission_rank=iii)
            output_samples.append(sample)
        return sorted(output_samples, key=lambda x: x.submission_rank)


__all__ = ["ClientSubmissionInfoWriter", "ClientSubmissionSampleWriter"]