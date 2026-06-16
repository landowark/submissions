"""
Writers for PCR results from Design and Analysis Software
"""
from __future__ import annotations
import logging
from pprint import pformat
from typing import Generator
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from pandas import DataFrame
from . import DefaultResultsInfoWriter, DefaultResultsSampleWriter

logger = logging.getLogger(f"submissions.{__name__}")


class DiomniPCRInfoWriter(DefaultResultsInfoWriter):

    pass

    

class DiomniPCRSampleWriter(DefaultResultsSampleWriter):

    def write_to_workbook(self, workbook: Workbook, sheet: str | None = None, start_row: int = 1, *args, **kwargs) -> Workbook:
        # super().write_to_workbook(workbook, sheet, start_row, *args, **kwargs)
        font = Font(bold=True, color="ffffffff", size=16)
        fill = PatternFill(start_color='376589', end_color='376589', fill_type="solid")
        align = Alignment(horizontal="center")
        start_row += 1
        self.start_row = start_row + 1
        try:
            self.worksheet = workbook[self.sheet]
        except KeyError:
            self.worksheet = workbook.create_sheet(title=self.sheet)
        for df in self.create_results_dataframes():
            rows = dataframe_to_rows(df, index=False)
            cell = self.worksheet.cell(row=start_row, column=1, value=df.caption)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            for row_data in rows:
                start_row += 1
                for col_idx, value in enumerate(row_data, start=1):
                    self.worksheet.cell(row=start_row, column=col_idx, value=value)
                # Move to the next row immediately after writing the columns
        self.postwrite(worksheet=self.worksheet)
        return workbook


    def create_results_dataframes(self) -> Generator[DataFrame, None, None]:
        """Flattens PydResults objects into a unified DataFrame.

        Each target (e.g., N1, N2) becomes a distinct row.
        """
        
        sheets = {key for obj in self.pydant_obj for key in obj.result.keys()}
        logger.debug(sheets)
        for sheet in sheets:
            all_rows = []
            for obj in self.pydant_obj:
                # Extract base metadata from the object
                metadata = {
                    "resultstype": obj.resultstype,
                    "procedure": obj.procedure,
                    "sample": obj.sample,
                }
                # Navigate to the target dictionary
                target_call = obj.result.get(sheet, {})
                if isinstance(target_call, dict):
                    for target_name, target_data in target_call.items():
                        # Merge metadata, the target key, and target metrics
                        row_data = {"Sample ID": metadata['sample'], "Target": target_name, **target_data}
                        all_rows.append(row_data)
            # Convert the list of dictionaries into a DataFrame
            df = DataFrame(all_rows)
            if df.empty:
                continue
            df.caption = sheet
            logger.debug(df)
            yield df

__all__ = ["DiomniPCRInfoWriter", "DiomniPCRSampleWriter"]