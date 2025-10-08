"""
Writers for PCR results from Design and Analysis Software
"""
from __future__ import annotations
import logging
from pprint import pformat
from typing import Generator, TYPE_CHECKING
from openpyxl import Workbook
from openpyxl.styles import Alignment
# from backend.excel.writers import DefaultKEYVALUEWriter, DefaultTABLEWriter
from . import DefaultResultsInfoWriter, DefaultResultsSampleWriter
from tools import flatten_list
if TYPE_CHECKING:
    from backend.db.models import ProcedureType

logger = logging.getLogger(f"submissions.{__name__}")

class PCRInfoWriter(DefaultResultsInfoWriter):

    start_row = 1

    def __init__(self, pydant_obj, proceduretype: "ProcedureType" | None = None, *args, **kwargs):
        super().__init__(pydant_obj=pydant_obj, proceduretype=proceduretype, *args, **kwargs)
        self.fill_dictionary = self.pydant_obj.improved_dict()['result']

    def write_to_workbook(self, workbook: Workbook, sheet: str | None = None,
                          start_row: int | None = None, *args, **kwargs) -> Workbook:
        workbook = super().write_to_workbook(workbook=workbook, sheet=f"{self.proceduretype.name} Results")
        return workbook


class PCRSampleWriter(DefaultResultsSampleWriter):

    def write_to_workbook(self, workbook: Workbook) -> Workbook:
        worksheet = workbook[f"{self.proceduretype.name} Results"]
        header_row = self.proceduretype.allowed_result_methods['PCR']['sample']['header_row']
        proto_columns = [(1, "sample"), (2, "target")]
        columns = []
        for iii, header in enumerate(self.column_headers, start=3):
            worksheet.cell(row=header_row, column=iii, value=header.replace("_", " ").title())
            columns.append((iii, header))
        columns = sorted(columns, key=lambda x: x[0])
        columns = proto_columns + columns
        all_results = flatten_list([[item for item in self.rearrange_results(result)] for result in self.pydant_obj])
        if len(all_results) > 0 :
            worksheet.cell(row=header_row, column=1, value="Sample")
            worksheet.cell(row=header_row, column=2, value="Target")
        for iii, item in enumerate(all_results, start=1):
            row = header_row + iii
            for k, v in item.items():
                column = next((col[0] for col in columns if col[1]==k), None)
                cell = worksheet.cell(row=row, column=column)
                cell.value = v
                cell.alignment = Alignment(horizontal='left')
        return workbook

    @classmethod
    def rearrange_results(cls, result) -> Generator[dict, None, None]:
        for target, values in result.result.items():
            if not isinstance(values, dict):
                continue
            values['target'] = target
            try:
                values['sample'] = result.sample_id
            except AttributeError as e:
                logger.error(f"No sample_id found for {pformat(result.__dict__)}")
                raise e
            yield values

    @property
    def column_headers(self):
        output = []
        for item in self.pydant_obj:
            dicto: dict = item.result
            for value in dicto.values():
                if not isinstance(value, dict):
                    continue
                for key in value.keys():
                    output.append(key)
        return sorted(list(set(output)))
