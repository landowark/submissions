"""
Module for default excel writers
"""
from __future__ import annotations
from logging import getLogger
logger = getLogger(f"submissions.{__name__}")
from numpy import nan as npnan
from typing import  TYPE_CHECKING
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows
from tools import flatten_list, sort_dict_by_list, handle_keys, handle_results
if TYPE_CHECKING:
    from backend.db.models import ProcedureType


class DefaultWriter(object):

    def __repr__(self):
        try:
            return f"{self.__class__.__name__}<{self.filepath.stem}>"
        except AttributeError:
            return f"{self.__class__.__name__}<Unknown Filepath>"

    def __init__(self, pydant_obj, *args, **kwargs):
        self.pydant_obj = pydant_obj
        self.write_sheet = pydant_obj.class_config.write_sheet
        
    def write_to_workbook(self, workbook: Workbook, sheet: str | None = None,
                          start_row: int | None = None, *args, **kwargs):
        if not start_row:
            try:
                start_row = self.__class__.start_row
            except AttributeError as e:
                logger.exception(f"Couldn't get start row due to {e}")
                start_row = 1
        if not sheet:
            sheet = self.write_sheet
        self.sheet = sheet
        sheetnames = workbook.sheetnames
        if isinstance(sheetnames, property):
            try:
                sheetnames = sheetnames.fget(workbook)
            except Exception as e:
                logger.exception(f"Couldn't resolve workbook sheetnames property due to {e}")
                sheetnames = []
        if self.sheet not in sheetnames:
            try:
                self.worksheet = workbook["Sheet"]
                self.worksheet.title = self.sheet
            except KeyError:
                self.worksheet = workbook.create_sheet(self.sheet)
        else:
            self.worksheet = workbook[self.sheet]
        self.worksheet = self.prewrite(self.worksheet, start_row=start_row)
        self.start_row = self.delineate_start_row(start_row=start_row)
        # NOTE: Declared in child classes
        self.end_row = self.delineate_end_row(start_row=start_row)
        return workbook

    def delineate_start_row(self, start_row: int = 1) -> int:
        """
        Gets the first black row.
        Args:
            start_row (int): row to start looking at.

        Returns:
            int
        """
        for iii, row in enumerate(self.worksheet.iter_rows(min_row=start_row), start=start_row):
            if all([item.value is None for item in row]):
                return iii
        if self.worksheet.max_row == 1:
            return self.worksheet.max_row + 1
        else:
            return self.worksheet.max_row + 2

    def prewrite(self, worksheet: Worksheet, start_row: int) -> Worksheet:
        return worksheet

    def columns_best_fit(self, worksheet: Worksheet) -> None:
        """
        Make all columns best fit
        """
        for col in worksheet.columns:
            setlen = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if len(str(cell.value)) > setlen:
                    setlen = len(str(cell.value))
            set_col_width = setlen + 5
            # Note: Setting the column width
            worksheet.column_dimensions[column].width = set_col_width
        return worksheet


class DefaultKEYVALUEWriter(DefaultWriter):
    
    
    def __init__(self, pydant_obj, proceduretype: ProcedureType | None = None, *args, **kwargs):
        super().__init__(pydant_obj=pydant_obj, proceduretype=proceduretype, *args, **kwargs)
        self.excluded = pydant_obj.class_config.excluded
        self.key_value_order = pydant_obj.class_config.key_value_order
        fill_dictionary = {k: v for k, v in self.pydant_obj.improved_dict.items() if k not in self.excluded}
        self.fill_dictionary = sort_dict_by_list(fill_dictionary, self.key_value_order)

    def delineate_end_row(self, start_row: int = 1):
        return len(self.fill_dictionary) + start_row

    @classmethod
    def check_location(cls, locations: list, sheet: str):
        return any([item['sheet'] == sheet for item in locations])

    def write_to_workbook(self, workbook: Workbook, sheet: str | None = None,
                          start_row: int = 1, *args, **kwargs) -> Workbook:
        workbook = super().write_to_workbook(workbook=workbook, sheet=sheet, start_row=start_row)
        for ii, (k, v) in enumerate(self.fill_dictionary.items(), start=self.start_row):
            value = handle_results(v)
            if value is None:
                continue
            self.worksheet.cell(column=1, row=ii, value=handle_keys(k))
            self.worksheet.cell(column=2, row=ii, value=handle_results(value))
        self.worksheet = self.postwrite(self.worksheet)
        
        return workbook

    def postwrite(self, worksheet: Worksheet) -> Worksheet:
        worksheet = self.columns_best_fit(worksheet=worksheet)
        return worksheet


class DefaultTABLEWriter(DefaultWriter):

    def __init__(self, pydant_obj, actual_objs_type: str | None = None, *args, **kwargs):
        super().__init__(pydant_obj=pydant_obj, *args, **kwargs)
        self.pydant_obj = getattr(self.pydant_obj, actual_objs_type) if actual_objs_type else self.pydant_obj
        try:
            self.excluded = self.pydant_obj[0].class_config.excluded
        except (IndexError, AttributeError, TypeError) as e:
            logger.exception(f"Error occurred while initializing TABLE writer: {e}")
            self.excluded = []
        try:
            self.key_value_order = self.pydant_obj[0].class_config.key_value_order
        except (IndexError, AttributeError, TypeError) as e:
            logger.exception(f"Error occurred while initializing TABLE writer: {e}")
            self.key_value_order = []

    def get_row_count(self, start_row: int = 1):
        list_df = DataFrame([item for item in self.worksheet.values][start_row - 1:])
        row_count = list_df.shape[0]
        return row_count

    def delineate_end_row(self, start_row: int = 1) -> int:
        end_row = start_row + len(self.pydant_obj) + 2
        return end_row

    def write_to_workbook(self, workbook: Workbook, sheet: str | None = None,
                          start_row: int | None = None, *args, **kwargs) -> Workbook:
        workbook = super().write_to_workbook(workbook=workbook, sheet=sheet, start_row=start_row, *args, **kwargs)
        
        records = [getattr(item, 'improved_dict', {}) for item in self.pydant_obj]
        df = DataFrame(records)[self.sorted_header_row]
        df.replace("", npnan, inplace=True)

        # Serialize list-valued columns so relationship fields like tipslot are preserved
        for column in df.columns:
            if df[column].apply(lambda value: isinstance(value, list)).any():
                df[column] = df[column].apply(
                    lambda value: "\n".join(str(item) for item in value) if isinstance(value, list) else value
                )

        # Identify columns where ALL values are zero
        is_all_zero = (df == 0).all()

        # Drop columns where the data is empty
        df = df.loc[:, ~is_all_zero]
        df.dropna(axis=1, how='all', inplace=True)
        df.fillna("", inplace=True)
        df.drop(columns=[col for col in df.columns if col in self.excluded], inplace=True, errors='ignore')
        # Rename column Headers.
        df = df.rename(columns=handle_keys)
        rows = dataframe_to_rows(df, index=False, header=True)
        for r_idx, row in enumerate(rows, start_row + 1 ):
            for c_idx, value in enumerate(row, 1):
                self.worksheet.cell(row=r_idx, column=c_idx, value=handle_results(value))
        self.worksheet = self.postwrite(self.worksheet)
        return workbook

    @property
    def sorted_header_row(self) -> list:
        output = []
        header_list = list(set(flatten_list([item.fields for item in self.pydant_obj])))
        class_config = self.pydant_obj[0].class_config
        for item in class_config.key_value_order:
            if item in [header for header in header_list if header not in class_config.excluded]:
                output.append(header_list.pop(header_list.index(item)))
        return output + sorted([item for item in header_list if item not in class_config.excluded])

    def postwrite(self, worksheet: Worksheet) -> Worksheet:
        worksheet = self.columns_best_fit(worksheet=worksheet)
        worksheet = self.colour_start_row(worksheet=worksheet)
        return worksheet
    
    def colour_start_row(self, worksheet: Worksheet) -> Worksheet:
        font = Font(bold=True, color="ffffffff")
        fill = PatternFill(start_color='376589', end_color='376589', fill_type="solid")
        align = Alignment(horizontal="center")
        for cell in worksheet[self.start_row]:
            cell.font = font
            cell.fill = fill
            cell.alignment = align
        return worksheet


from .procedure_writers import *
from .results_writers import *
from .clientsubmission_writer import *

__all__ = ["DefaultKEYVALUEWriter", "DefaultTABLEWriter", "ProcedureInfoWriter", "ProcedureReagentWriter", "ProcedureEquipmentWriter", 
           "ProcedureSampleWriter", "DefaultResultsInfoWriter", "DefaultResultsSampleWriter", "DiomniPCRInfoWriter", "DiomniPCRSampleWriter", 
           "QubitInfoWriter", "QubitSampleWriter", "ClientSubmissionInfoWriter", "ClientSubmissionSampleWriter"]
