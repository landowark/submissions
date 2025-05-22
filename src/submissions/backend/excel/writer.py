"""
contains writer objects for pushing values to procedure sheet templates.
"""
import logging
from copy import copy
from datetime import datetime
from operator import itemgetter
from pprint import pformat
from typing import List, Generator, Tuple
from openpyxl import load_workbook, Workbook
from backend.db.models import SubmissionType, KitType, Run
from backend.validators.pydant import PydSubmission
from io import BytesIO
from collections import OrderedDict

logger = logging.getLogger(f"submissions.{__name__}")


class SheetWriter(object):
    """
    object to manage data placement into excel file
    """

    def __init__(self, submission: PydSubmission):
        """
        Args:
            submission (PydSubmission): Object containing procedure information.
        """
        self.sub = OrderedDict(submission.improved_dict())
        # NOTE: Set values from pydantic object.
        for k, v in self.sub.items():
            match k:
                case 'filepath':
                    self.__setattr__(k, v)
                case 'proceduretype':
                    self.sub[k] = v['value']
                    self.submission_type = SubmissionType.query(name=v['value'])
                    self.run_object = BasicRun.find_polymorphic_subclass(
                        polymorphic_identity=self.submission_type)
                case _:
                    if isinstance(v, dict):
                        self.sub[k] = v['value']
                    else:
                        self.sub[k] = v
        template = self.submission_type.template_file
        if not template:
            logger.error(f"No template file found, falling back to Bacterial Culture")
            template = SubmissionType.basic_template
        workbook = load_workbook(BytesIO(template))
        self.xl = workbook
        self.write_info()
        self.write_reagents()
        self.write_samples()
        self.write_equipment()
        self.write_tips()

    def write_info(self):
        """
        Calls info writer
        """
        disallowed = ['filepath', 'reagents', 'sample', 'equipment', 'control']
        info_dict = {k: v for k, v in self.sub.items() if k not in disallowed}
        writer = InfoWriter(xl=self.xl, submission_type=self.submission_type, info_dict=info_dict)
        self.xl = writer.write_info()

    def write_reagents(self):
        """
        Calls reagent writer
        """
        reagent_list = self.sub['reagents']
        writer = ReagentWriter(xl=self.xl, submission_type=self.submission_type,
                               extraction_kit=self.sub['kittype'], reagent_list=reagent_list)
        self.xl = writer.write_reagents()

    def write_samples(self):
        """
        Calls sample writer
        """
        sample_list = self.sub['sample']
        writer = SampleWriter(xl=self.xl, submission_type=self.submission_type, sample_list=sample_list)
        self.xl = writer.write_samples()

    def write_equipment(self):
        """
        Calls equipment writer
        """
        equipment_list = self.sub['equipment']
        writer = EquipmentWriter(xl=self.xl, submission_type=self.submission_type, equipment_list=equipment_list)
        self.xl = writer.write_equipment()

    def write_tips(self):
        """
        Calls tip writer
        """
        tips_list = self.sub['tips']
        writer = TipWriter(xl=self.xl, submission_type=self.submission_type, tips_list=tips_list)
        self.xl = writer.write_tips()


class InfoWriter(object):
    """
    object to write general procedure info into excel file
    """

    def __init__(self, xl: Workbook, submission_type: SubmissionType | str, info_dict: dict,
                 sub_object: Run | None = None):
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (SubmissionType | str): Type of procedure expected (Wastewater, Bacterial Culture, etc.)
            info_dict (dict): Dictionary of information to write.
            sub_object (BasicRun | None, optional): Submission object containing methods. Defaults to None.
        """
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        if sub_object is None:
            sub_object = Run.find_polymorphic_subclass(polymorphic_identity=submission_type.name)
        self.submission_type = submission_type
        self.sub_object = sub_object
        self.xl = xl
        self.info_map = submission_type.construct_info_map(mode='write')
        self.info = self.reconcile_map(info_dict, self.info_map)

    def reconcile_map(self, info_dict: dict, info_map: dict) -> Generator[(Tuple[str, dict]), None, None]:
        """
        Merge info with its locations

        Args:
            info_dict (dict): dictionary of info items
            info_map (dict): dictionary of info locations

        Returns:
            dict: merged dictionary
        """
        for k, v in info_dict.items():
            if v is None:
                continue
            if k == "custom":
                continue
            dicto = {}
            try:
                dicto['locations'] = info_map[k]
            except KeyError:
                pass
            dicto['value'] = v
            if len(dicto) > 0:
                yield k, dicto

    def write_info(self) -> Workbook:
        """
        Performs write operations

        Returns:
            Workbook: workbook with info written.
        """
        final_info = {}
        for k, v in self.info:
            match k:
                case "custom":
                    continue
                case "comment":
                    # NOTE: merge all comments to fit in single cell.
                    if isinstance(v['value'], list):
                        json_join = [item['text'] for item in v['value'] if 'text' in item.keys()]
                        v['value'] = "\n".join(json_join)
                case thing if thing in self.sub_object.timestamps:
                    v['value'] = v['value'].date()
                case _:
                    pass
            final_info[k] = v
            try:
                locations = v['locations']
            except KeyError:
                logger.error(f"No locations for {k}, skipping")
                continue
            for loc in locations:
                sheet = self.xl[loc['sheet']]
                try:
                    # logger.debug(f"Writing {v['value']} to row {loc['row']} and column {loc['column']}")
                    sheet.cell(row=loc['row'], column=loc['column'], value=v['value'])
                except AttributeError as e:
                    logger.error(f"Can't write {k} to that cell due to AttributeError: {e}")
                except ValueError as e:
                    logger.error(f"Can't write {v} to that cell due to ValueError: {e}")
                    sheet.cell(row=loc['row'], column=loc['column'], value=v['value'].name)
        return self.sub_object.custom_info_writer(self.xl, info=final_info, custom_fields=self.info_map['custom'])


class ReagentWriter(object):
    """
    object to write reagent data into excel file
    """

    def __init__(self, xl: Workbook, submission_type: SubmissionType | str, extraction_kit: KitType | str,
                 reagent_list: list):
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (SubmissionType | str): Type of procedure expected (Wastewater, Bacterial Culture, etc.)
            extraction_kit (KitType | str): Extraction kittype used.
            reagent_list (list): List of reagent dicts to be written to excel.
        """
        self.xl = xl
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        self.submission_type_obj = submission_type
        if isinstance(extraction_kit, str):
            extraction_kit = KitType.query(name=extraction_kit)
        self.kit_object = extraction_kit
        associations, self.kit_object = self.kit_object.construct_xl_map_for_use(
            proceduretype=self.submission_type_obj)
        reagent_map = {k: v for k, v in associations.items()}
        self.reagents = self.reconcile_map(reagent_list=reagent_list, reagent_map=reagent_map)

    def reconcile_map(self, reagent_list: List[dict], reagent_map: dict) -> Generator[dict, None, None]:
        """
        Merge reagents with their locations

        Args:
            reagent_list (List[dict]): List of reagent dictionaries
            reagent_map (dict): Reagent locations

        Returns:
            List[dict]: merged dictionary
        """
        filled_roles = [item['reagentrole'] for item in reagent_list]
        for map_obj in reagent_map.keys():
            if map_obj not in filled_roles:
                reagent_list.append(dict(name="Not Applicable", role=map_obj, lot="Not Applicable", expiry="Not Applicable"))
        for reagent in reagent_list:
            try:
                mp_info = reagent_map[reagent['reagentrole']]
            except KeyError:
                continue
            placeholder = copy(reagent)
            for k, v in reagent.items():
                try:
                    dicto = dict(value=v, row=mp_info[k]['row'], column=mp_info[k]['column'])
                except KeyError as e:
                    logger.error(f"KeyError: {e}")
                    dicto = v
                placeholder[k] = dicto
                placeholder['sheet'] = mp_info['sheet']
            yield placeholder

    def write_reagents(self) -> Workbook:
        """
        Performs write operations

        Returns:
            Workbook: Workbook with reagents written
        """
        for reagent in self.reagents:
            sheet = self.xl[reagent['sheet']]
            for v in reagent.values():
                if not isinstance(v, dict):
                    continue
                match v['value']:
                    case datetime():
                        v['value'] = v['value'].date()
                    case _:
                        pass
                sheet.cell(row=v['row'], column=v['column'], value=v['value'])
        return self.xl


class SampleWriter(object):
    """
    object to write sample data into excel file
    """

    def __init__(self, xl: Workbook, submission_type: SubmissionType | str, sample_list: list):
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (SubmissionType | str): Type of procedure expected (Wastewater, Bacterial Culture, etc.)
            sample_list (list): List of sample dictionaries to be written to excel file.
        """
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        self.submission_type = submission_type
        self.xl = xl
        self.sample_map = submission_type.sample_map['lookup_table']
        # NOTE: exclude any sample without a procedure rank.
        samples = [item for item in self.reconcile_map(sample_list) if item['submission_rank'] > 0]
        self.samples = sorted(samples, key=itemgetter('submission_rank'))
        self.blank_lookup_table()

    def reconcile_map(self, sample_list: list) -> Generator[dict, None, None]:
        """
        Merge sample info with locations

        Args:
            sample_list (list): List of sample information

        Returns:
            List[dict]: List of merged dictionaries
        """
        multiples = ['row', 'column', 'assoc_id', 'submission_rank']
        for sample in sample_list:
            sample = self.submission_type.submission_class.custom_sample_writer(sample)
            for assoc in zip(sample['row'], sample['column'], sample['submission_rank']):
                new = dict(row=assoc[0], column=assoc[1], submission_rank=assoc[2])
                for k, v in sample.items():
                    if k in multiples:
                        continue
                    new[k] = v
                yield new

    def blank_lookup_table(self):
        """
        Blanks out columns in the lookup table to ensure help values are removed before writing.
        """
        sheet = self.xl[self.sample_map['sheet']]
        for row in range(self.sample_map['start_row'], self.sample_map['end_row'] + 1):
            for column in self.sample_map['sample_columns'].values():
                if sheet.cell(row, column).data_type != 'f':
                    sheet.cell(row=row, column=column, value="")

    def write_samples(self) -> Workbook:
        """
        Performs writing operations.

        Returns:
            Workbook: Workbook with sample written
        """
        sheet = self.xl[self.sample_map['sheet']]
        columns = self.sample_map['sample_columns']
        for sample in self.samples:
            row = self.sample_map['start_row'] + (sample['submission_rank'] - 1)
            for k, v in sample.items():
                if isinstance(v, dict):
                    try:
                        v = v['value']
                    except KeyError:
                        logger.error(f"Cant convert {v} to single string.")
                try:
                    column = columns[k]
                except KeyError:
                    continue
                sheet.cell(row=row, column=column, value=v)
        return self.xl


class EquipmentWriter(object):
    """
    object to write equipment data into excel file
    """

    def __init__(self, xl: Workbook, submission_type: SubmissionType | str, equipment_list: list):
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (SubmissionType | str): Type of procedure expected (Wastewater, Bacterial Culture, etc.)
            equipment_list (list): List of equipment dictionaries to write to excel file.
        """
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        self.submission_type = submission_type
        self.xl = xl
        equipment_map = {k: v for k, v in self.submission_type.construct_field_map("equipment")}
        self.equipment = self.reconcile_map(equipment_list=equipment_list, equipment_map=equipment_map)

    def reconcile_map(self, equipment_list: list, equipment_map: dict) -> Generator[dict, None, None]:
        """
        Merges equipment with location data

        Args:
            equipment_list (list): List of equipment
            equipment_map (dict): Dictionary of equipment locations

        Returns:
            List[dict]: List of merged dictionaries
        """
        if equipment_list is None:
            return
        for ii, equipment in enumerate(equipment_list, start=1):
            try:
                mp_info = equipment_map[equipment['reagentrole']]
            except KeyError:
                logger.error(f"No {equipment['reagentrole']} in {pformat(equipment_map)}")
                mp_info = None
            placeholder = copy(equipment)
            if not mp_info:
                for jj, (k, v) in enumerate(equipment.items(), start=1):
                    dicto = dict(value=v, row=ii, column=jj)
                    placeholder[k] = dicto
            else:
                for jj, (k, v) in enumerate(equipment.items(), start=1):
                    try:
                        dicto = dict(value=v, row=mp_info[k]['row'], column=mp_info[k]['column'])
                    except KeyError as e:
                        continue
                    placeholder[k] = dicto
                if "asset_number" not in mp_info.keys():
                    placeholder['name']['value'] = f"{equipment['name']} - {equipment['asset_number']}"
            try:
                placeholder['sheet'] = mp_info['sheet']
            except KeyError:
                placeholder['sheet'] = "Equipment"
            yield placeholder

    def write_equipment(self) -> Workbook:
        """
        Performs write operations

        Returns:
            Workbook: Workbook with equipment written
        """
        for equipment in self.equipment:
            if not equipment['sheet'] in self.xl.sheetnames:
                self.xl.create_sheet("Equipment")
            sheet = self.xl[equipment['sheet']]
            for k, v in equipment.items():
                if not isinstance(v, dict):
                    continue
                if isinstance(v['value'], list):
                    v['value'] = v['value'][0]
                try:
                    sheet.cell(row=v['row'], column=v['column'], value=v['value'])
                except AttributeError as e:
                    logger.error(f"Couldn't write to {equipment['sheet']}, row: {v['row']}, column: {v['column']}")
                    logger.error(e)
        return self.xl


class TipWriter(object):
    """
    object to write tips data into excel file
    """

    def __init__(self, xl: Workbook, submission_type: SubmissionType | str, tips_list: list):
        """
        Args:
            xl (Workbook): Openpyxl workbook from submitted excel file.
            submission_type (SubmissionType | str): Type of procedure expected (Wastewater, Bacterial Culture, etc.)
            tips_list (list): List of tip dictionaries to write to the excel file.
        """
        if isinstance(submission_type, str):
            submission_type = SubmissionType.query(name=submission_type)
        self.submission_type = submission_type
        self.xl = xl
        tips_map = {k: v for k, v in self.submission_type.construct_field_map("tip")}
        self.tips = self.reconcile_map(tips_list=tips_list, tips_map=tips_map)

    def reconcile_map(self, tips_list: List[dict], tips_map: dict) -> Generator[dict, None, None]:
        """
        Merges tips with location data

        Args:
            tips_list (List[dict]): List of tips
            tips_map (dict): Tips locations

        Returns:
            List[dict]: List of merged dictionaries
        """
        if tips_list is None:
            return
        for ii, tips in enumerate(tips_list, start=1):
            mp_info = tips_map[tips.role]
            placeholder = {}
            if mp_info == {}:
                for jj, (k, v) in enumerate(tips.__dict__.items(), start=1):
                    dicto = dict(value=v, row=ii, column=jj)
                    placeholder[k] = dicto
            else:
                for jj, (k, v) in enumerate(tips.__dict__.items(), start=1):
                    try:
                        dicto = dict(value=v, row=mp_info[k]['row'], column=mp_info[k]['column'])
                    except KeyError as e:
                        continue
                    placeholder[k] = dicto
            try:
                placeholder['sheet'] = mp_info['sheet']
            except KeyError:
                placeholder['sheet'] = "Tips"
            yield placeholder

    def write_tips(self) -> Workbook:
        """
        Performs write operations

        Returns:
            Workbook: Workbook with tips written
        """
        for tips in self.tips:
            if not tips['sheet'] in self.xl.sheetnames:
                self.xl.create_sheet("Tips")
            sheet = self.xl[tips['sheet']]
            for k, v in tips.items():
                if not isinstance(v, dict):
                    continue
                if isinstance(v['value'], list):
                    v['value'] = v['value'][0]
                try:
                    sheet.cell(row=v['row'], column=v['column'], value=v['value'])
                except AttributeError as e:
                    logger.error(f"Couldn't write to {tips['sheet']}, row: {v['row']}, column: {v['column']}")
                    logger.error(e)
        return self.xl
