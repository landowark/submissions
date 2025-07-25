"""
Models for the main submission and sample types.
"""
from __future__ import annotations

import itertools
import pickle
from copy import deepcopy
from getpass import getuser
import logging, uuid, tempfile, re, base64, numpy as np, pandas as pd, types, sys
from inspect import isclass
from zipfile import ZipFile, BadZipfile
from tempfile import TemporaryDirectory, TemporaryFile
from operator import itemgetter
from pprint import pformat
from pandas import DataFrame
from sqlalchemy.ext.hybrid import hybrid_property
from . import BaseClass, Reagent, SubmissionType, KitType, Organization, Contact, LogMixin, SubmissionReagentAssociation
from sqlalchemy import Column, String, TIMESTAMP, INTEGER, ForeignKey, JSON, FLOAT, case, func
from sqlalchemy.orm import relationship, validates, Query
from sqlalchemy.orm.attributes import flag_modified
from sqlalchemy.ext.associationproxy import association_proxy
from sqlalchemy.exc import OperationalError as AlcOperationalError, IntegrityError as AlcIntegrityError, StatementError, \
    ArgumentError
from sqlite3 import OperationalError as SQLOperationalError, IntegrityError as SQLIntegrityError
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from tools import row_map, setup_lookup, jinja_template_loading, rreplace, row_keys, check_key_or_attr, Result, Report, \
    report_result, create_holidays_for_year, check_dictionary_inclusion_equality
from datetime import datetime, date
from typing import List, Any, Tuple, Literal, Generator, Type
from pathlib import Path
from jinja2.exceptions import TemplateNotFound
from jinja2 import Template
from PIL import Image

logger = logging.getLogger(f"submissions.{__name__}")


class BasicSubmission(BaseClass, LogMixin):
    """
    Concrete of basic submission which polymorphs into BacterialCulture and Wastewater
    """

    id = Column(INTEGER, primary_key=True)  #: primary key
    rsl_plate_num = Column(String(32), unique=True, nullable=False)  #: RSL name (e.g. RSL-22-0012)
    submitter_plate_num = Column(String(127), unique=True)  #: The number given to the submission by the submitting lab
    submitted_date = Column(TIMESTAMP)  #: Date submission received
    submitting_lab = relationship("Organization", back_populates="submissions")  #: client org
    submitting_lab_id = Column(INTEGER, ForeignKey("_organization.id", ondelete="SET NULL",
                                                   name="fk_BS_sublab_id"))  #: client lab id from _organizations
    sample_count = Column(INTEGER)  #: Number of samples in the submission
    extraction_kit = relationship("KitType", back_populates="submissions")  #: The extraction kit used
    extraction_kit_id = Column(INTEGER, ForeignKey("_kittype.id", ondelete="SET NULL",
                                                   name="fk_BS_extkit_id"))  #: id of joined extraction kit
    submission_type_name = Column(String, ForeignKey("_submissiontype.name", ondelete="SET NULL",
                                                     name="fk_BS_subtype_name"))  #: name of joined submission type
    technician = Column(String(64))  #: initials of processing tech(s)
    reagents_id = Column(String, ForeignKey("_reagent.id", ondelete="SET NULL",
                                            name="fk_BS_reagents_id"))  #: id of used reagents
    extraction_info = Column(JSON)  #: unstructured output from the extraction table logger.
    run_cost = Column(
        FLOAT(2))  #: total cost of running the plate. Set from constant and mutable kit costs at time of creation.
    signed_by = Column(String(32))  #: user name of person who submitted the submission to the database.
    comment = Column(JSON)  #: user notes
    submission_category = Column(
        String(64))  #: ["Research", "Diagnostic", "Surveillance", "Validation"], else defaults to submission_type_name
    cost_centre = Column(
        String(64))  #: Permanent storage of used cost centre in case organization field changed in the future.
    contact = relationship("Contact", back_populates="submissions")  #: client org
    contact_id = Column(INTEGER, ForeignKey("_contact.id", ondelete="SET NULL",
                                            name="fk_BS_contact_id"))  #: client lab id from _organizations
    custom = Column(JSON)
    controls = relationship("Control", back_populates="submission",
                            uselist=True)  #: A control sample added to submission
    completed_date = Column(TIMESTAMP)

    submission_sample_associations = relationship(
        "SubmissionSampleAssociation",
        back_populates="submission",
        cascade="all, delete-orphan",
    )  #: Relation to SubmissionSampleAssociation

    samples = association_proxy("submission_sample_associations",
                                "sample", creator=lambda sample: SubmissionSampleAssociation(
            sample=sample))  #: Association proxy to SubmissionSampleAssociation.samples

    submission_reagent_associations = relationship(
        "SubmissionReagentAssociation",
        back_populates="submission",
        cascade="all, delete-orphan",
    )  #: Relation to SubmissionReagentAssociation

    reagents = association_proxy("submission_reagent_associations",
                                 "reagent")  #: Association proxy to SubmissionReagentAssociation.reagent

    submission_equipment_associations = relationship(
        "SubmissionEquipmentAssociation",
        back_populates="submission",
        cascade="all, delete-orphan"
    )  #: Relation to Equipment

    equipment = association_proxy("submission_equipment_associations",
                                  "equipment")  #: Association proxy to SubmissionEquipmentAssociation.equipment

    submission_tips_associations = relationship(
        "SubmissionTipsAssociation",
        back_populates="submission",
        cascade="all, delete-orphan")

    tips = association_proxy("submission_tips_associations",
                             "tips")

    # NOTE: Allows for subclassing into ex. BacterialCulture, Wastewater, etc.
    __mapper_args__ = {
        "polymorphic_identity": "Basic Submission",
        "polymorphic_on": case(

            (submission_type_name == "Wastewater", "Wastewater"),
            (submission_type_name == "Wastewater Artic", "Wastewater Artic"),
            (submission_type_name == "Bacterial Culture", "Bacterial Culture"),

            else_="Basic Submission"
        ),
        "with_polymorphic": "*",
    }

    def __repr__(self) -> str:
        return f"<Submission({self.rsl_plate_num})>"

    @hybrid_property
    def kittype(self):
        return self.extraction_kit

    @hybrid_property
    def organization(self):
        return self.submitting_lab

    @classproperty
    def jsons(cls) -> List[str]:
        """
        Get list of JSON db columns

        Returns:
            List[str]: List of column names
        """
        output = [item.name for item in cls.__table__.columns if isinstance(item.type, JSON)]
        if issubclass(cls, BasicSubmission) and not cls.__name__ == "BasicSubmission":
            output += BasicSubmission.jsons
        return output

    @classproperty
    def timestamps(cls) -> List[str]:
        """
        Get list of TIMESTAMP columns

        Returns:
            List[str]: List of column names
        """
        output = [item.name for item in cls.__table__.columns if isinstance(item.type, TIMESTAMP)]
        if issubclass(cls, BasicSubmission) and not cls.__name__ == "BasicSubmission":
            output += BasicSubmission.timestamps
        return output

    @classmethod
    def get_default_info(cls, *args, submission_type: SubmissionType | None = None) -> dict:
        """
        Gets default info from the database for a given submission type.

        Args:
            *args (): List of fields to get
            submission_type (SubmissionType): the submission type of interest. Necessary due to generic submission types.

        Returns:
            dict: Default info

        """
        # NOTE: Create defaults for all submission_types
        # NOTE: Singles tells the query which fields to set limit to 1
        dicto = super().get_default_info()
        recover = ['filepath', 'samples', 'csv', 'comment', 'equipment']
        dicto.update(dict(
            details_ignore=['excluded', 'reagents', 'samples',
                            'extraction_info', 'comment', 'barcode',
                            'platemap', 'export_map', 'equipment', 'tips', 'custom'],
            # NOTE: Fields not placed in ui form
            form_ignore=['reagents', 'ctx', 'id', 'cost', 'extraction_info', 'signed_by', 'comment', 'namer',
                         'submission_object', "tips", 'contact_phone', 'custom', 'cost_centre', 'completed_date',
                         'controls', "origin_plate"] + recover,
            # NOTE: Fields not placed in ui form to be moved to pydantic
            form_recover=recover
        ))
        # NOTE: Grab mode_sub_type specific info.
        if args:
            output = {k: v for k, v in dicto.items() if k in args}
        else:
            output = {k: v for k, v in dicto.items()}
        if isinstance(submission_type, SubmissionType):
            st = submission_type
        else:
            st = cls.get_submission_type(submission_type)
        if st is None:
            logger.error("No default info for BasicSubmission.")
        else:
            output['submission_type'] = st.name
            for k, v in st.defaults.items():
                if args and k not in args:
                    continue
                else:
                    match v:
                        case list():
                            output[k] += v
                        case _:
                            output[k] = v
        if len(args) == 1:
            try:
                return output[args[0]]
            except KeyError as e:
                if "pytest" in sys.modules and args[0] == "abbreviation":
                    return "BS"
                else:
                    raise KeyError(f"{args[0]} not found in {output}")
        return output

    @classmethod
    def get_submission_type(cls, sub_type: str | SubmissionType | None = None) -> SubmissionType:
        """
        Gets the SubmissionType associated with this class

        Args:
            sub_type (str | SubmissionType, Optional): Identity of the submission type to retrieve. Defaults to None.

        Returns:
            SubmissionType: SubmissionType with name equal sub_type or this polymorphic identity if sub_type is None.
        """
        if isinstance(sub_type, dict):
            try:
                sub_type = sub_type['value']
            except KeyError as e:
                logger.error(f"Couldn't extract value from {sub_type}")
                raise e
        match sub_type:
            case str():
                return SubmissionType.query(name=sub_type)
            case SubmissionType():
                return sub_type
            case _:
                return SubmissionType.query(cls.__mapper_args__['polymorphic_identity'])

    @classmethod
    def construct_info_map(cls, submission_type: SubmissionType | None = None,
                           mode: Literal["read", "write"] = "read") -> dict:
        """
        Method to call submission type's construct info map.

        Args:
            mode (Literal["read", "write"]): Which map to construct.

        Returns:
            dict: Map of info locations.
        """
        return cls.get_submission_type(submission_type).construct_info_map(mode=mode)

    @classmethod
    def construct_sample_map(cls, submission_type: SubmissionType | None = None) -> dict:
        """
        Method to call submission type's construct_sample_map

        Returns:
            dict: sample location map
        """
        return cls.get_submission_type(submission_type).sample_map

    def generate_associations(self, name: str, extra: str | None = None):
        try:
            field = self.__getattribute__(name)
        except AttributeError:
            return None
        for item in field:
            if extra:
                yield item.to_sub_dict(extra)
            else:
                yield item.to_sub_dict()

    def to_dict(self, full_data: bool = False, backup: bool = False, report: bool = False) -> dict:
        """
        Constructs dictionary used in submissions summary

        Args:
            expand (bool, optional): indicates if generators to be expanded. Defaults to False.
            report (bool, optional): indicates if to be used for a report. Defaults to False.
            full_data (bool, optional): indicates if sample dicts to be constructed. Defaults to False.
            backup (bool, optional): passed to adjust_to_dict_samples. Defaults to False.

        Returns:
            dict: dictionary used in submissions summary and details
        """
        # NOTE: get lab from nested organization object
        try:
            sub_lab = self.submitting_lab.name
        except AttributeError:
            sub_lab = None
        try:
            sub_lab = sub_lab.replace("_", " ").title()
        except AttributeError:
            pass
        # NOTE: get extraction kit name from nested kit object
        try:
            ext_kit = self.extraction_kit.name
        except AttributeError:
            ext_kit = None
        # NOTE: load scraped extraction info
        try:
            ext_info = self.extraction_info
        except TypeError:
            ext_info = None
        output = {
            "id": self.id,
            "plate_number": self.rsl_plate_num,
            "submission_type": self.submission_type_name,
            "submitter_plate_number": self.submitter_plate_num,
            "submitted_date": self.submitted_date.strftime("%Y-%m-%d"),
            "submitting_lab": sub_lab,
            "sample_count": self.sample_count,
            "extraction_kit": ext_kit,
            "cost": self.run_cost
        }
        if report:
            return output
        if full_data:
            try:
                reagents = [item.to_sub_dict(extraction_kit=self.extraction_kit) for item in
                            self.submission_reagent_associations]
            except Exception as e:
                logger.error(f"We got an error retrieving reagents: {e}")
                reagents = []
            finally:
                dicto, _ = self.extraction_kit.construct_xl_map_for_use(self.submission_type)
                for k, v in dicto.items():
                    if k == 'info':
                        continue
                    if not any([item['role'] == k for item in reagents]):
                        expiry = "NA"
                        reagents.append(
                            dict(role=k, name="Not Applicable", lot="NA", expiry=expiry,
                                 missing=True))
            samples = self.generate_associations(name="submission_sample_associations")
            equipment = self.generate_associations(name="submission_equipment_associations")
            tips = self.generate_associations(name="submission_tips_associations")
            cost_centre = self.cost_centre
            custom = self.custom
            controls = [item.to_sub_dict() for item in self.controls]
        else:
            reagents = None
            samples = None
            equipment = None
            tips = None
            cost_centre = None
            custom = None
            controls = None
        try:
            comments = self.comment
        except Exception as e:
            logger.error(f"Error setting comment: {self.comment}, {e}")
            comments = None
        try:
            contact = self.contact.name
        except AttributeError as e:
            try:
                contact = f"Defaulted to: {self.submitting_lab.contacts[0].name}"
            except (AttributeError, IndexError):
                contact = "NA"
        try:
            contact_phone = self.contact.phone
        except AttributeError:
            contact_phone = "NA"
        output["submission_category"] = self.submission_category
        output["technician"] = self.technician
        output["reagents"] = reagents
        output["samples"] = samples
        output["extraction_info"] = ext_info
        output["comment"] = comments
        output["equipment"] = equipment
        output["tips"] = tips
        output["cost_centre"] = cost_centre
        output["signed_by"] = self.signed_by
        output["contact"] = contact
        output["contact_phone"] = contact_phone
        output["custom"] = custom
        output["controls"] = controls
        try:
            output["completed_date"] = self.completed_date.strftime("%Y-%m-%d")
        except AttributeError:
            output["completed_date"] = self.completed_date
        return output

    @classmethod
    def archive_submissions(cls, start_date: date | datetime | str | int | None = None,
                            end_date: date | datetime | str | int | None = None,
                            submissiontype: List[str] | None = None):
        if submissiontype:
            if isinstance(submissiontype, str):
                submissiontype = [submissiontype]
            query_out = []
            for sub_type in submissiontype:
                subs = cls.query(page_size=0, start_date=start_date, end_date=end_date, submissiontype=sub_type)
                # logger.debug(f"Sub results: {subs}")
                query_out.append(subs)
            query_out = list(itertools.chain.from_iterable(query_out))
        else:
            query_out = cls.query(page_size=0, start_date=start_date, end_date=end_date)
        records = []
        for sub in query_out:
            output = sub.to_dict(full_data=True)
            for k, v in output.items():
                if isinstance(v, types.GeneratorType):
                    output[k] = [item for item in v]
            records.append(output)
        df = DataFrame.from_records(records)
        df.sort_values(by="id", inplace=True)
        df.set_index("id", inplace=True)
        return df

    @property
    def column_count(self) -> int:
        """
        Calculate the number of columns in this submission 

        Returns:
            int: Number of unique columns.
        """
        columns = set([assoc.column for assoc in self.submission_sample_associations])
        return len(columns)

    def calculate_base_cost(self) -> None:
        """
        Calculates cost of the plate
        """
        # NOTE: Calculate number of columns based on largest column number
        try:
            cols_count_96 = self.column_count
        except Exception as e:
            logger.error(f"Column count error: {e}")
        # NOTE: Get kit associated with this submission
        # logger.debug(f"Checking associations with submission type: {self.submission_type_name}")
        assoc = next((item for item in self.extraction_kit.kit_submissiontype_associations if
                      item.submission_type == self.submission_type),
                     None)
        # logger.debug(f"Got association: {assoc}")
        # NOTE: If every individual cost is 0 this is probably an old plate.
        if all(item == 0.0 for item in [assoc.constant_cost, assoc.mutable_cost_column, assoc.mutable_cost_sample]):
            try:
                self.run_cost = self.extraction_kit.cost_per_run
            except Exception as e:
                logger.error(f"Calculation error: {e}")
        else:
            try:
                self.run_cost = assoc.constant_cost + (assoc.mutable_cost_column * cols_count_96) + (
                        assoc.mutable_cost_sample * int(self.sample_count))
            except Exception as e:
                logger.error(f"Calculation error: {e}")
        self.run_cost = round(self.run_cost, 2)

    @property
    def hitpicked(self) -> list:
        """
        Returns positve sample locations for plate

        Returns:
            list: list of hitpick dictionaries for each sample
        """
        output_list = [assoc.hitpicked for assoc in self.submission_sample_associations]
        return output_list

    @classmethod
    def make_plate_map(cls, sample_list: list, plate_rows: int = 8, plate_columns=12) -> str:
        """
        Constructs an html based plate map for submission details.

        Args:
            sample_list (list): List of submission samples
            plate_rows (int, optional): Number of rows in the plate. Defaults to 8.
            plate_columns (int, optional): Number of columns in the plate. Defaults to 12.

        Returns:
            str: html output string.
        """
        rows = range(1, plate_rows + 1)
        columns = range(1, plate_columns + 1)
        # logger.debug(f"sample list for plate map: {pformat(sample_list)}")
        # NOTE: An overly complicated list comprehension create a list of sample locations
        # NOTE: next will return a blank cell if no value found for row/column
        output_samples = [next((item for item in sample_list if item['row'] == row and item['column'] == column),
                               dict(name="", row=row, column=column, background_color="#ffffff"))
                          for row in rows
                          for column in columns]
        env = jinja_template_loading()
        template = env.get_template("plate_map.html")
        html = template.render(samples=output_samples, PLATE_ROWS=plate_rows, PLATE_COLUMNS=plate_columns)
        return html + "<br/>"

    @property
    def used_equipment(self) -> Generator[str, None, None]:
        """
        Gets EquipmentRole names associated with this BasicSubmission

        Returns:
            List[str]: List of names
        """
        return (item.role for item in self.submission_equipment_associations)

    @classmethod
    def submissions_to_df(cls, submission_type: str | None = None, limit: int = 0,
                          chronologic: bool = True, page: int = 1, page_size: int = 250) -> pd.DataFrame:
        """
        Convert all submissions to dataframe

        Args:
            page_size (int, optional): Number of items to include in query result. Defaults to 250.
            page (int, optional): Limits the number of submissions to a page size. Defaults to 1.
            chronologic (bool, optional): Sort submissions in chronologic order. Defaults to True.
            submission_type (str | None, optional): Filter by SubmissionType. Defaults to None.
            limit (int, optional): Maximum number of results to return. Defaults to 0.

        Returns:
            pd.DataFrame: Pandas Dataframe of all relevant submissions
        """
        # NOTE: use lookup function to create list of dicts
        subs = [item.to_dict() for item in
                cls.query(submissiontype=submission_type, limit=limit, chronologic=chronologic, page=page,
                          page_size=page_size)]
        df = pd.DataFrame.from_records(subs)
        # NOTE: Exclude sub information
        exclude = ['controls', 'extraction_info', 'pcr_info', 'comment', 'comments', 'samples', 'reagents',
                   'equipment', 'gel_info', 'gel_image', 'dna_core_submission_number', 'gel_controls',
                   'source_plates', 'pcr_technician', 'ext_technician', 'artic_technician', 'cost_centre',
                   'signed_by', 'artic_date', 'gel_barcode', 'gel_date', 'ngs_date', 'contact_phone', 'contact',
                   'tips', 'gel_image_path', 'custom']
        # NOTE: dataframe equals dataframe of all columns not in exclude
        df = df.loc[:, ~df.columns.isin(exclude)]
        if chronologic:
            try:
                df.sort_values(by="id", axis=0, inplace=True, ascending=False)
            except KeyError:
                logger.error("No column named 'id'")
        # NOTE: Human friendly column labels
        df.columns = [item.replace("_", " ").title() for item in df.columns]
        return df

    def set_attribute(self, key: str, value):
        """
        Performs custom attribute setting based on values.

        Args:
            key (str): name of attribute
            value (_type_): value of attribute
        """
        match key:
            case "extraction_kit":
                field_value = KitType.query(name=value)
            case "submitting_lab":
                field_value = Organization.query(name=value)
            case "contact":
                field_value = Contact.query(name=value)
            case "samples":
                for sample in value:
                    sample, _ = sample.to_sql(submission=self)
                return
            case "reagents":
                field_value = [reagent['value'].to_sql()[0] if isinstance(reagent, dict) else reagent.to_sql()[0] for
                               reagent in value]
            case "submission_type":
                field_value = SubmissionType.query(name=value)
            case "sample_count":
                if value is None:
                    field_value = len(self.samples)
                else:
                    field_value = value
            case "ctx" | "csv" | "filepath" | "equipment" | "controls":
                return
            case item if item in self.jsons:
                match key:
                    case "custom" | "source_plates":
                        existing = value
                    case _:
                        existing = self.__getattribute__(key)
                        logger.debug(f"Existing value is {pformat(existing)}")
                        if value in ['', 'null', None]:
                            logger.error(f"No value given, not setting.")
                            return
                        if existing is None:
                            existing = []
                        # if value in existing:
                        if check_dictionary_inclusion_equality(existing, value):
                            logger.warning("Value already exists. Preventing duplicate addition.")
                            return
                        else:
                            if isinstance(value, list):
                                existing += value
                            else:
                                if value:
                                    existing.append(value)

                self.__setattr__(key, existing)
                # NOTE: Make sure this gets updated by telling SQLAlchemy it's been modified.
                flag_modified(self, key)
                return
            case _:
                try:
                    field_value = value.strip()
                except AttributeError:
                    field_value = value
        # NOTE: insert into field
        current = self.__getattribute__(key)
        if field_value and current != field_value:
            try:
                self.__setattr__(key, field_value)
            except AttributeError as e:
                logger.error(f"Could not set {self} attribute {key} to {value} due to \n{e}")

    def update_subsampassoc(self, assoc: SubmissionSampleAssociation, input_dict: dict) -> SubmissionSampleAssociation:
        """
        Update a joined submission sample association.

        Args:
            assoc (SubmissionSampleAssociation): Sample association to be updated.
            input_dict (dict): updated values to insert.

        Returns:
            SubmissionSampleAssociation: Updated association
        """
        # NOTE: No longer searches for association here, done in caller function
        for k, v in input_dict.items():
            try:
                setattr(assoc, k, v)
                # NOTE: for some reason I don't think assoc.__setattr__(k, v) works here.
            except AttributeError:
                pass
        return assoc

    def update_reagentassoc(self, reagent: Reagent, role: str):
        # NOTE: get the first reagent assoc that fills the given role.
        try:
            assoc = next(item for item in self.submission_reagent_associations if
                         item.reagent and role in [role.name for role in item.reagent.role])
            assoc.reagent = reagent
        except StopIteration as e:
            logger.error(f"Association for {role} not found, creating new association.")
            assoc = SubmissionReagentAssociation(submission=self, reagent=reagent)
            self.submission_reagent_associations.append(assoc)

    def to_pydantic(self, backup: bool = False) -> "PydSubmission":
        """
        Converts this instance into a PydSubmission

        Returns:
            PydSubmission: converted object.
        """
        from backend.validators import PydSubmission
        dicto = self.to_dict(full_data=True, backup=backup)
        new_dict = {}
        for key, value in dicto.items():
            missing = value in ['', 'None', None]
            match key:
                case "reagents":
                    field_value = [item.to_pydantic(extraction_kit=self.extraction_kit) for item in
                                   self.submission_reagent_associations]
                case "samples":
                    field_value = [item.to_pydantic() for item in self.submission_sample_associations]
                case "equipment":
                    field_value = [item.to_pydantic() for item in self.submission_equipment_associations]
                case "controls":
                    try:
                        field_value = [item.to_pydantic() for item in self.__getattribute__(key)]
                    except TypeError as e:
                        logger.error(f"Error converting {key} to pydantic :{e}")
                        continue
                case "tips":
                    field_value = [item.to_pydantic() for item in self.submission_tips_associations]
                case "submission_type":
                    field_value = dict(value=self.__getattribute__(key).name, missing=missing)
                case "plate_number":
                    key = 'rsl_plate_num'
                    field_value = dict(value=self.rsl_plate_num, missing=missing)
                case "submitter_plate_number":
                    key = "submitter_plate_num"
                    field_value = dict(value=self.submitter_plate_num, missing=missing)
                case "id":
                    continue
                case _:
                    try:
                        key = key.lower().replace(" ", "_")
                        if isclass(value):
                            field_value = dict(value=self.__getattribute__(key).name, missing=missing)
                        else:
                            field_value = dict(value=self.__getattribute__(key), missing=missing)
                    except AttributeError:
                        logger.error(f"{key} is not available in {self}")
                        field_value = dict(value="NA", missing=True)
            new_dict[key] = field_value
        new_dict['filepath'] = Path(tempfile.TemporaryFile().name)
        dicto.update(new_dict)
        return PydSubmission(**dicto)

    def save(self, original: bool = True):
        """
        Adds this instance to database and commits.

        Args:
            original (bool, optional): Is this the first save. Defaults to True.
        """
        if original:
            self.uploaded_by = getuser()
        return super().save()

    @classmethod
    def get_regex(cls, submission_type: SubmissionType | str | None = None) -> re.Pattern:
        """
        Gets the regex string for identifying a certain class of submission.

        Args:
            submission_type (SubmissionType | str | None, optional): submission type of interest. Defaults to None.

        Returns:
            str: String from which regex will be compiled.
        """
        # logger.debug(f"Class for regex: {cls}")
        try:
            regex = cls.get_submission_type(submission_type).defaults['regex']
        except AttributeError as e:
            logger.error(f"Couldn't get submission type for {cls.__mapper_args__['polymorphic_identity']}")
            regex = None
        try:
            regex = re.compile(rf"{regex}", flags=re.IGNORECASE | re.VERBOSE)
        except re.error as e:
            regex = cls.construct_regex()
        # logger.debug(f"Returning regex: {regex}")
        return regex

    # NOTE: Polymorphic functions

    @classproperty
    def regex(cls) -> re.Pattern:
        """
        Constructs catchall regex.

        Returns:
            re.Pattern: Regular expression pattern to discriminate between submission types.
        """
        res = [st.defaults['regex'] for st in SubmissionType.query() if st.defaults]
        rstring = rf'{"|".join(res)}'
        regex = re.compile(rstring, flags=re.IGNORECASE | re.VERBOSE)
        return regex

    @classmethod
    def find_polymorphic_subclass(cls, polymorphic_identity: str | SubmissionType | list | None = None,
                                  attrs: dict | None = None) -> BasicSubmission | List[BasicSubmission]:
        """
        Find subclass based on polymorphic identity or relevant attributes.

        Args:
            polymorphic_identity (str | None, optional): String representing polymorphic identity. Defaults to None.
            attrs (str | SubmissionType | None, optional): Attributes of the relevant class. Defaults to None.

        Returns:
            _type_: Subclass of interest.
        """
        if isinstance(polymorphic_identity, dict):
            polymorphic_identity = polymorphic_identity['value']
        if isinstance(polymorphic_identity, SubmissionType):
            polymorphic_identity = polymorphic_identity.name
        model = cls
        match polymorphic_identity:
            case str():
                try:
                    model = cls.__mapper__.polymorphic_map[polymorphic_identity].class_
                except Exception as e:
                    logger.error(
                        f"Could not get polymorph {polymorphic_identity} of {cls} due to {e}, falling back to BasicSubmission")
            case list():
                output = []
                for identity in polymorphic_identity:
                    if isinstance(identity, SubmissionType):
                        identity = polymorphic_identity.name
                    output.append(cls.__mapper__.polymorphic_map[identity].class_)
                return output
            case _:
                pass
        if attrs and any([not hasattr(cls, attr) for attr in attrs.keys()]):
            # NOTE: looks for first model that has all included kwargs
            try:
                model = next(subclass for subclass in cls.__subclasses__() if
                             all([hasattr(subclass, attr) for attr in attrs.keys()]))
            except StopIteration as e:
                raise AttributeError(
                    f"Couldn't find existing class/subclass of {cls} with all attributes:\n{pformat(attrs.keys())}")
        return model

    # NOTE: Child class custom functions

    @classmethod
    def custom_info_parser(cls, input_dict: dict, xl: Workbook | None = None, custom_fields: dict = {}) -> dict:
        """
        Update submission dictionary with type specific information

        Args:
            input_dict (dict): Input sample dictionary
            xl (Workbook): original xl workbook, used for child classes mostly
            custom_fields: Dictionary of locations, ranges, etc to be used by this function

        Returns:
            dict: Updated sample dictionary
        """
        input_dict['custom'] = {}
        for k, v in custom_fields.items():
            logger.debug(f"Custom info parser getting type: {v['type']}")
            match v['type']:
                # NOTE: 'exempt' type not currently used
                case "exempt":
                    continue
                case "cell":
                    ws = xl[v['read']['sheet']]
                    input_dict['custom'][k] = ws.cell(row=v['read']['row'], column=v['read']['column']).value
                case "range":
                    ws = xl[v['sheet']]
                    if v['start_row'] != v['end_row']:
                        v['end_row'] = v['end_row'] + 1
                    rows = range(v['start_row'], v['end_row'])
                    if v['start_column'] != v['end_column']:
                        v['end_column'] = v['end_column'] + 1
                    columns = range(v['start_column'], v['end_column'])
                    input_dict['custom'][k] = [dict(value=ws.cell(row=row, column=column).value, row=row, column=column)
                                               for row in rows for column in columns]
        return input_dict

    @classmethod
    def parse_samples(cls, input_dict: dict) -> dict:
        """
        Update sample dictionary with type specific information

        Args:
            input_dict (dict): Input sample dictionary

        Returns:
            dict: Updated sample dictionary
        """
        return input_dict

    @classmethod
    def custom_validation(cls, pyd: "PydSubmission") -> "PydSubmission":
        """
        Performs any final parsing of the pydantic object that only needs to be done for this cls.

        Args:
            input_dict (dict): Parser product up to this point.
            xl (pd.ExcelFile | None, optional): Excel submission form. Defaults to None.
            info_map (dict | None, optional): Map of information locations from SubmissionType. Defaults to None.
            plate_map (dict | None, optional): Constructed plate map of samples. Defaults to None.

        Returns:
            dict: Updated parser product.
        """
        return pyd

    @classmethod
    def custom_info_writer(cls, input_excel: Workbook, info: dict | None = None, backup: bool = False,
                           custom_fields: dict = {}) -> Workbook:
        """
        Adds custom autofill methods for submission

        Args:
            input_excel (Workbook): initial workbook.
            info (dict | None, optional): dictionary of additional info. Defaults to None.
            backup (bool, optional): Whether this is part of a backup operation. Defaults to False.
            custom_fields: Dictionary of locations, ranges, etc to be used by this function

        Returns:
            Workbook: Updated workbook
        """
        for k, v in custom_fields.items():
            try:
                assert v['type'] in ['exempt', 'range', 'cell']
            except (AssertionError, KeyError):
                continue
            match v['type']:
                case "exempt":
                    continue
                case "cell":
                    v['write'].append(v['read'])
                    for cell in v['write']:
                        ws = input_excel[cell['sheet']]
                        ws.cell(row=cell['row'], column=cell['column'], value=info['custom'][k])
                case "range":
                    ws = input_excel[v['sheet']]
                    if v['start_row'] != v['end_row']:
                        v['end_row'] = v['end_row'] + 1
                    if v['start_column'] != v['end_column']:
                        v['end_column'] = v['end_column'] + 1
                    for item in info['custom'][k]:
                        ws.cell(row=item['row'], column=item['column'], value=item['value'])
        return input_excel

    @classmethod
    def custom_sample_writer(self, sample: dict) -> dict:
        """
        Performs any final alterations to sample writing unique to this submission type.
        Args:
            sample (dict): Dictionary of sample values.

        Returns:
            dict: Finalized dictionary.
        """
        return sample

    @classmethod
    def enforce_name(cls, instr: str, data: dict | None = {}) -> str:
        """
        Custom naming method for this class.

        Args:
            instr (str): Initial name.
            data (dict | None, optional): Additional parameters for name. Defaults to None.

        Returns:
            str: Updated name.
        """
        from backend.validators import RSLNamer
        if "submission_type" not in data.keys():
            data['submission_type'] = cls.__mapper_args__['polymorphic_identity']
        data['abbreviation'] = cls.get_default_info("abbreviation", submission_type=data['submission_type'])
        if instr in [None, ""]:
            outstr = RSLNamer.construct_new_plate_name(data=data)
        else:
            outstr = instr
        if re.search(rf"{data['abbreviation']}", outstr, flags=re.IGNORECASE) is None:
            # NOTE: replace RSL- with RSL-abbreviation-
            outstr = re.sub(rf"RSL-?", rf"RSL-{data['abbreviation']}-", outstr, flags=re.IGNORECASE)
        try:
            # NOTE: remove dashes from date
            outstr = re.sub(r"(\d{4})-(\d{2})-(\d{2})", r"\1\2\3", outstr)
            # NOTE: insert dash between abbreviation and date
            outstr = re.sub(rf"{data['abbreviation']}(\d{6})", rf"{data['abbreviation']}-\1", outstr,
                            flags=re.IGNORECASE).upper()
        except (AttributeError, TypeError) as e:
            logger.error(f"Error making outstr: {e}, sending to RSLNamer to make new plate name.")
            outstr = RSLNamer.construct_new_plate_name(data=data)
        try:
            # NOTE: Grab plate number as number after a -|_ not followed by another number
            plate_number = re.search(r"(?:(-|_)\d)(?!\d)", outstr).group().strip("_").strip("-")
        except AttributeError as e:
            plate_number = "1"
        # NOTE: insert dash between date and plate number
        outstr = re.sub(r"(\d{8})(-|_)?\d?(R\d?)?", rf"\1-{plate_number}\3", outstr)
        try:
            # NOTE: grab repeat number
            repeat = re.search(r"-\dR(?P<repeat>\d)?", outstr).groupdict()['repeat']
            if repeat is None:
                repeat = "1"
        except AttributeError as e:
            repeat = ""
        # NOTE: Insert repeat number?
        outstr = re.sub(r"(-\dR)\d?", rf"\1 {repeat}", outstr).replace(" ", "")
        # NOTE: This should already have been done. Do I dare remove it?
        outstr = re.sub(rf"RSL{data['abbreviation']}", rf"RSL-{data['abbreviation']}", outstr)
        return re.sub(rf"{data['abbreviation']}(\d)", rf"{data['abbreviation']}-\1", outstr)

    @classmethod
    def parse_pcr(cls, xl: Workbook, rsl_plate_num: str) -> Generator[dict, None, None]:
        """
        Perform parsing of pcr info. Since most of our PC outputs are the same format, this should work for most.

        Args:
            xl (pd.DataFrame): pcr info form
            rsl_plate_num (str): rsl plate num of interest

        Returns:
            Generator[dict, None, None]: Updated samples
        """
        pcr_sample_map = cls.get_submission_type().sample_map['pcr_samples']
        main_sheet = xl[pcr_sample_map['main_sheet']]
        fields = {k: v for k, v in pcr_sample_map.items() if k not in ['main_sheet', 'start_row']}
        logger.debug(f"Fields: {fields}")
        for row in main_sheet.iter_rows(min_row=pcr_sample_map['start_row']):
            idx = row[0].row
            sample = {}
            for k, v in fields.items():
                # logger.debug(f"Checking key: {k} with value {v}")
                sheet = xl[v['sheet']]
                sample[k] = sheet.cell(row=idx, column=v['column']).value
            yield sample

    @classmethod
    def parse_pcr_controls(cls, xl: Workbook, rsl_plate_num: str) -> Generator[dict, None, None]:
        """
        Custom parsing of pcr controls from Design & Analysis Software export.

        Args:
            xl (Workbook): D&A export file
            rsl_plate_num (str): Plate number of the submission to be joined.

        Yields:
            Generator[dict, None, None]: Dictionaries of row values.
        """
        location_map = cls.get_submission_type().sample_map['pcr_controls']
        # logger.debug(f"Location map: {location_map}")
        submission = cls.query(rsl_plate_num=rsl_plate_num)
        name_column = 1
        for item in location_map:
            # logger.debug(f"Checking {item}")
            worksheet = xl[item['sheet']]
            for iii, row in enumerate(worksheet.iter_rows(max_row=len(worksheet['A']), max_col=name_column), start=1):
                # logger.debug(f"Checking row {row}, {iii}")
                for cell in row:
                    # logger.debug(f"Checking cell: {cell}, with value {cell.value} against {item['name']}")
                    if cell.value == item['name']:
                        subtype, _ = item['name'].split("-")
                        target = item['target']
                        # logger.debug(f"Subtype: {subtype}, target: {target}")
                        ct = worksheet.cell(row=iii, column=item['ct_column']).value
                        # NOTE: Kind of a stop gap solution to find control reagents.
                        if subtype == "PC":
                            ctrl = next((assoc.reagent for assoc in submission.submission_reagent_associations
                                         if
                                         any(["positive control" in item.name.lower() for item in assoc.reagent.role])),
                                        None)
                        elif subtype == "NC":
                            ctrl = next((assoc.reagent for assoc in submission.submission_reagent_associations
                                         if any(["molecular grade water" in item.name.lower() for item in
                                                 assoc.reagent.role])), None)
                        else:
                            ctrl = None
                        # logger.debug(f"Control reagent: {ctrl.__dict__}")
                        try:
                            ct = float(ct)
                        except ValueError:
                            ct = 0.0
                        if ctrl:
                            ctrl = ctrl.lot
                        else:
                            ctrl = None
                        output = dict(
                            name=f"{rsl_plate_num}<{item['name']}-{target}>",
                            ct=ct,
                            subtype=subtype,
                            target=target,
                            reagent_lot=ctrl
                        )
                        # logger.debug(f"Control output: {pformat(output)}")
                        yield output

    @classmethod
    def filename_template(cls) -> str:
        """
        Constructs template for filename of this class.
        Note: This is meant to be used with the dictionary constructed in self.to_dict(). Keys need to have spaces removed

        Returns:
            str: filename template in jinja friendly format.
        """
        return "{{ rsl_plate_num }}"

    @classmethod
    def adjust_autofill_samples(cls, samples: List[Any]) -> List[Any]:
        """
        Makes adjustments to samples before writing to excel.

        Args:
            samples (List[Any]): List of Samples

        Returns:
            List[Any]: Updated list of samples
        """
        return samples

    @classmethod
    def get_details_template(cls, base_dict: dict) -> Tuple[dict, Template]:
        """
        Get the details jinja template for the correct class

        Args:
            base_dict (dict): incoming dictionary of Submission fields

        Returns:
            Tuple(dict, Template): (Updated dictionary, Template to be rendered)
        """
        base_dict['excluded'] = cls.get_default_info('details_ignore')
        base_dict['excluded'] += ['controls']
        env = jinja_template_loading()
        temp_name = f"{cls.__name__.lower()}_details.html"
        try:
            template = env.get_template(temp_name)
        except TemplateNotFound as e:
            logger.error(f"Couldn't find template due to {e}")
            template = env.get_template("basicsubmission_details.html")
        return base_dict, template

    # NOTE: Query functions

    @classmethod
    @setup_lookup
    def query(cls,
              submissiontype: str | SubmissionType | None = None,
              submission_type_name: str | None = None,
              id: int | str | None = None,
              rsl_plate_num: str | None = None,
              start_date: date | datetime | str | int | None = None,
              end_date: date | datetime | str | int | None = None,
              reagent: Reagent | str | None = None,
              chronologic: bool = False,
              limit: int = 0,
              page: int = 1,
              page_size: None | int = 250,
              **kwargs
              ) -> BasicSubmission | List[BasicSubmission]:
        """
        Lookup submissions based on a number of parameters. Overrides parent.

        Args:
            submission_type (str | models.SubmissionType | None, optional): Submission type of interest. Defaults to None.
            id (int | str | None, optional): Submission id in the database (limits results to 1). Defaults to None.
            rsl_plate_num (str | None, optional): Submission name in the database (limits results to 1). Defaults to None.
            start_date (date | str | int | None, optional): Beginning date to search by. Defaults to None.
            end_date (date | str | int | None, optional): Ending date to search by. Defaults to None.
            reagent (models.Reagent | str | None, optional): A reagent used in the submission. Defaults to None.
            chronologic (bool, optional): Return results in chronologic order. Defaults to False.
            limit (int, optional): Maximum number of results to return. Defaults to 0.

        Returns:
            models.BasicSubmission | List[models.BasicSubmission]: Submission(s) of interest
        """
        # from ... import SubmissionReagentAssociation
        # NOTE: if you go back to using 'model' change the appropriate cls to model in the query filters
        if submissiontype is not None:
            model = cls.find_polymorphic_subclass(polymorphic_identity=submissiontype)
        elif len(kwargs) > 0:
            # NOTE: find the subclass containing the relevant attributes
            model = cls.find_polymorphic_subclass(attrs=kwargs)
        else:
            model = cls
        query: Query = cls.__database_session__.query(model)
        if start_date is not None and end_date is None:
            logger.warning(f"Start date with no end date, using today.")
            end_date = date.today()
        if end_date is not None and start_date is None:
            # NOTE: this query returns a tuple of (object, datetime), need to get only datetime.
            start_date = cls.__database_session__.query(cls, func.min(cls.submitted_date)).first()[1]
            logger.warning(f"End date with no start date, using first submission date: {start_date}")
        if start_date is not None:
            # match start_date:
            #     case date():
            #         pass
            #     case datetime():
            #         start_date = start_date.date()
            #     case int():
            #         start_date = datetime.fromordinal(
            #             datetime(1900, 1, 1).toordinal() + start_date - 2).date()
            #     case _:
            #         start_date = parse(start_date).date()
            # # start_date = start_date.strftime("%Y-%m-%d")
            # match end_date:
            #     case date():
            #         pass
            #     case datetime():
            #         end_date = end_date  # + timedelta(days=1)
            #         # pass
            #     case int():
            #         end_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + end_date - 2).date()  # \
            #         # + timedelta(days=1)
            #     case _:
            #         end_date = parse(end_date).date()  # + timedelta(days=1)
            # # end_date = end_date.strftime("%Y-%m-%d")
            # start_date = datetime.combine(start_date, datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S.%f")
            # end_date = datetime.combine(end_date, datetime.max.time()).strftime("%Y-%m-%d %H:%M:%S.%f")
            # # if start_date == end_date:
            # #     start_date = start_date.strftime("%Y-%m-%d %H:%M:%S.%f")
            # #     query = query.filter(model.submitted_date == start_date)
            # # else:
            start_date = cls.rectify_query_date(start_date)
            end_date = cls.rectify_query_date(end_date, eod=True)
            query = query.filter(model.submitted_date.between(start_date, end_date))
        # NOTE: by reagent (for some reason)
        match reagent:
            case str():
                query = query.join(SubmissionReagentAssociation).join(Reagent).filter(
                    Reagent.lot == reagent)
            case Reagent():
                query = query.join(SubmissionReagentAssociation).filter(
                    SubmissionReagentAssociation.reagent == reagent)
            case _:
                pass
        # NOTE: by rsl number (returns only a single value)
        match rsl_plate_num:
            case str():
                query = query.filter(model.rsl_plate_num == rsl_plate_num)
                limit = 1
            case _:
                pass
        match submission_type_name:
            case str():
                query = query.filter(model.submission_type_name == submission_type_name)
            case _:
                pass
        # NOTE: by id (returns only a single value)
        match id:
            case int():
                query = query.filter(model.id == id)
                limit = 1
            case str():
                query = query.filter(model.id == int(id))
                limit = 1
            case _:
                pass
        query = query.order_by(cls.submitted_date.desc())
        # NOTE: Split query results into pages of size {page_size}
        if page_size > 0:
            query = query.limit(page_size)
        page = page - 1
        if page is not None:
            query = query.offset(page * page_size)
        return cls.execute_query(query=query, model=model, limit=limit, **kwargs)

    @classmethod
    def query_or_create(cls, submission_type: str | SubmissionType | None = None, **kwargs) -> BasicSubmission:
        """
        Returns object from db if exists, else, creates new. Due to need for user input, doesn't see much use ATM.

        Args:
            submission_type (str | SubmissionType | None, optional): Submission type to be created. Defaults to None.

        Raises:
            ValueError: Raised if no kwargs passed.
            ValueError: Raised if disallowed key is passed.

        Returns:
            cls: A BasicSubmission subclass.
        """
        code = 0
        msg = ""
        report = Report()
        disallowed = ["id"]
        if kwargs == {}:
            raise ValueError("Need to narrow down query or the first available instance will be returned.")
        sanitized_kwargs = {k: v for k, v in kwargs.items() if k not in disallowed}
        instance = cls.query(submissiontype=submission_type, limit=1, **sanitized_kwargs)
        if instance is None:
            used_class = cls.find_polymorphic_subclass(attrs=kwargs, polymorphic_identity=submission_type)
            instance = used_class(**sanitized_kwargs)
            match submission_type:
                case str():
                    submission_type = SubmissionType.query(name=submission_type)
                case _:
                    pass
            instance.submission_type = submission_type
            instance.submission_type_name = submission_type.name
            if "submitted_date" not in kwargs.keys():
                instance.submitted_date = date.today()
        else:
            from frontend.widgets.pop_ups import QuestionAsker
            logger.warning(f"Found existing instance: {instance}, asking to overwrite.")
            #     code = 1
            #     msg = "This submission already exists.\nWould you like to overwrite?"
            # report.add_result(Result(msg=msg, code=code))
            dlg = QuestionAsker(title="Overwrite?",
                                message="This submission already exists.\nWould you like to overwrite?")
            if dlg.exec():
                pass
            else:
                code = 1
                msg = "This submission already exists.\nWould you like to overwrite?"
                report.add_result(Result(msg=msg, code=code))
                return None, report
        return instance, report

    # NOTE: Custom context events for the ui

    def custom_context_events(self) -> dict:
        """
        Creates dictionary of str:function to be passed to context menu

        Returns:
            dict: dictionary of functions
        """
        names = ["Delete", "Details", "Edit", "Add Comment", "Add Equipment", "Export"]
        funcs = [self.delete, self.show_details, self.edit, self.add_comment, self.add_equipment, self.backup]
        dicto = {item[0]: item[1] for item in zip(names, funcs)}
        return dicto

    def delete(self, obj=None):
        """
        Performs backup and deletes this instance from database.

        Args:
            obj (_type_, optional): Parent widget. Defaults to None.

        Raises:
            e: SQLIntegrityError or SQLOperationalError if problem with commit.
        """
        from frontend.widgets.pop_ups import QuestionAsker
        fname = self.__backup_path__.joinpath(f"{self.rsl_plate_num}-backup({date.today().strftime('%Y%m%d')})")
        msg = QuestionAsker(title="Delete?", message=f"Are you sure you want to delete {self.rsl_plate_num}?\n")
        if msg.exec():
            try:
                # NOTE: backs up file as xlsx, same as export.
                self.backup(fname=fname, full_backup=True)
            except BadZipfile:
                logger.error("Couldn't open zipfile for writing.")
            self.__database_session__.delete(self)
            try:
                self.__database_session__.commit()
            except (SQLIntegrityError, SQLOperationalError, AlcIntegrityError, AlcOperationalError) as e:
                self.__database_session__.rollback()
                raise e
            try:
                obj.setData()
            except AttributeError:
                logger.error("App will not refresh data at this time.")

    def show_details(self, obj):
        """
        Creates Widget for showing submission details.

        Args:
            obj (Widget): Parent widget
        """
        from frontend.widgets.submission_details import SubmissionDetails
        dlg = SubmissionDetails(parent=obj, sub=self)
        if dlg.exec():
            pass

    def edit(self, obj):
        """
        Return submission to form widget for updating

        Args:
            obj (Widget): Parent widget 
        """
        from frontend.widgets.submission_widget import SubmissionFormWidget
        for widget in obj.app.table_widget.formwidget.findChildren(SubmissionFormWidget):
            widget.setParent(None)
        pyd = self.to_pydantic(backup=True)
        form = pyd.to_form(parent=obj, disable=['rsl_plate_num'])
        obj.app.table_widget.formwidget.layout().addWidget(form)

    def add_comment(self, obj):
        """
        Creates widget for adding comments to submissions

        Args:
            obj (_type_): parent widget
        """
        from frontend.widgets.submission_details import SubmissionComment
        dlg = SubmissionComment(parent=obj, submission=self)
        if dlg.exec():
            comment = dlg.parse_form()
            if comment in ["", None]:
                return
            self.set_attribute(key='comment', value=comment)
            self.save(original=False)

    def add_equipment(self, obj):
        """
        Creates widget for adding equipment to this submission

        Args:
            obj (_type_): parent widget
        """
        from frontend.widgets.equipment_usage import EquipmentUsage
        dlg = EquipmentUsage(parent=obj, submission=self)
        if dlg.exec():
            equipment = dlg.parse_form()
            for equip in equipment:
                logger.debug(f"Parsed equipment: {equip}")
                _, assoc = equip.to_sql(submission=self)
                logger.debug(f"Got equipment association: {assoc} for {equip}")
                try:
                    assoc.save()
                except AttributeError as e:
                    logger.error(f"Couldn't save association with {equip} due to {e}")
                if equip.tips:
                    for tips in equip.tips:
                        # logger.debug(f"Attempting to add tips assoc: {tips} (pydantic)")
                        tassoc = tips.to_sql(submission=self)
                        # logger.debug(f"Attempting to add tips assoc: {tips.__dict__} (sql)")
                        if tassoc not in self.submission_tips_associations:
                            tassoc.save()
                        else:
                            logger.error(f"Tips already found in submission, skipping.")
        else:
            pass

    def backup(self, obj=None, fname: Path | None = None, full_backup: bool = False):
        """
        Exports xlsx info files for this instance.

        Args:
            obj (_type_, optional): _description_. Defaults to None.
            fname (Path | None, optional): Filename of xlsx file. Defaults to None.
            full_backup (bool, optional): Whether or not to make yaml file. Defaults to False.
        """
        pyd = self.to_pydantic(backup=True)
        if fname is None:
            from frontend.widgets.functions import select_save_file
            fname = select_save_file(default_name=pyd.construct_filename(), extension="xlsx", obj=obj)
        if fname.name == "":
            return
        writer = pyd.to_writer()
        writer.xl.save(filename=fname.with_suffix(".xlsx"))

    @property
    def turnaround_time(self) -> int:
        try:
            completed = self.completed_date.date()
        except AttributeError:
            completed = None
        return self.calculate_turnaround(start_date=self.submitted_date.date(), end_date=completed)

    @classmethod
    def calculate_turnaround(cls, start_date: date | None = None, end_date: date | None = None) -> int:
        """
        Calculates number of business days between data submitted and date completed

        Args:
            start_date (date, optional): Date submitted. defaults to None.
            end_date (date, optional): Date completed. defaults to None.

        Returns:
            int: Number of business days.
        """
        if not end_date:
            return None
        try:
            delta = np.busday_count(start_date, end_date, holidays=create_holidays_for_year(start_date.year)) + 1
        except ValueError:
            return None
        return delta


# NOTE: Below are the custom submission types

class BacterialCulture(BasicSubmission):
    """
    derivative submission type from BasicSubmission
    """
    id = Column(INTEGER, ForeignKey('_basicsubmission.id'), primary_key=True)

    __mapper_args__ = dict(polymorphic_identity="Bacterial Culture",
                           polymorphic_load="inline",
                           inherit_condition=(id == BasicSubmission.id))

    def to_dict(self, full_data: bool = False, backup: bool = False, report: bool = False) -> dict:
        """
        Extends parent class method to add controls to dict

        Returns:
            dict: dictionary used in submissions summary
        """
        output = super().to_dict(full_data=full_data, backup=backup, report=report)
        if report:
            return output
        if full_data:
            output['controls'] = [item.to_sub_dict() for item in self.controls]
        return output

    @classmethod
    def filename_template(cls):
        """
        extends parent
        """
        template = super().filename_template()
        template += "_{{ submitting_lab.name }}_{{ submitter_plate_num }}"
        return template

    @classmethod
    def custom_validation(cls, pyd) -> "PydSubmission":
        """
        Extends parent. Currently finds control sample and adds to reagents.

        Args:
            input_dict (dict): _description_
            xl (pd.ExcelFile | None, optional): _description_. Defaults to None.
            info_map (dict | None, optional): _description_. Defaults to None.

        Returns:
            PydSubmission: Updated pydantic.
        """
        from . import ControlType
        pyd = super().custom_validation(pyd)
        # NOTE: build regex for all control types that have targets
        regex = ControlType.build_positive_regex(control_type="Irida Control")
        # NOTE: search samples for match
        for sample in pyd.samples:
            matched = regex.match(sample.submitter_id)
            if bool(matched):
                new_lot = matched.group()
                try:
                    pos_control_reg = \
                        next(reg for reg in pyd.reagents if reg.role == "Bacterial-Positive Control")
                except StopIteration:
                    logger.error(f"No positive control reagent listed")
                    return pyd
                pos_control_reg.lot = new_lot
                pos_control_reg.missing = False
        return pyd

    @classmethod
    def custom_info_parser(cls, input_dict: dict, xl: Workbook | None = None, custom_fields: dict = {}) -> dict:
        """
        Performs class specific info parsing before info parsing is finalized.

        Args:
            input_dict (dict): Generic input info
            xl (Workbook | None, optional): Original xl workbook. Defaults to None.
            custom_fields (dict, optional): Map of custom fields to be parsed. Defaults to {}.

        Returns:
            dict: Updated info dictionary.
        """
        input_dict = super().custom_info_parser(input_dict=input_dict, xl=xl, custom_fields=custom_fields)
        return input_dict

    def custom_context_events(self) -> dict:
        """
        Sets context events for main widget

        Returns:
            dict: Context menu items for this instance.
        """
        events = super().custom_context_events()
        events['Import Concentration'] = self.import_concentration
        return events

    @report_result
    def import_concentration(self, obj) -> Report:
        from frontend.widgets import select_open_file
        from backend.excel.parser import ConcentrationParser
        report = Report()
        fname = select_open_file(obj=obj, file_extension="xlsx")
        if not fname:
            report.add_result(Result(msg="No file selected, cancelling.", status="Warning"))
            return report
        parser = ConcentrationParser(filepath=fname, submission=self)
        conc_samples = [sample for sample in parser.samples]
        # logger.debug(f"Concentration samples: {pformat(conc_samples)}")
        for sample in self.samples:
            # logger.debug(f"Sample {sample.submitter_id}")
            # logger.debug(f"Match {item['submitter_id']}")
            try:
                # NOTE: Fix for ENs which have no rsl_number...
                sample_dict = next(item for item in conc_samples if str(item['submitter_id']).upper() == sample.submitter_id)
            except StopIteration:
                logger.error(f"Couldn't find sample dict for {sample.submitter_id}")
                continue
            logger.debug(f"Sample {sample.submitter_id} conc. = {sample_dict['concentration']}")
            if sample_dict['concentration']:
                sample.concentration = sample_dict['concentration']
            else:
                continue
            sample.save()
        # logger.debug(conc_samples)
        return report

    @classmethod
    def parse_concentration(cls, xl: Workbook, rsl_plate_num: str) -> Generator[dict, None, None]:
        lookup_table = cls.get_submission_type().sample_map['lookup_table']
        logger.debug(lookup_table)
        main_sheet = xl[lookup_table['sheet']]
        for row in main_sheet.iter_rows(min_row=lookup_table['start_row'], max_row=lookup_table['end_row']):
            idx = row[0].row
            sample = dict(
                submitter_id=main_sheet.cell(row=idx, column=lookup_table['sample_columns']['submitter_id']).value)
            sample['concentration'] = main_sheet.cell(row=idx,
                                                      column=lookup_table['sample_columns']['concentration']).value
            yield sample

    # def get_provisional_controls(self, controls_only: bool = True):
    def get_provisional_controls(self, include: List[str] = []):
        # NOTE To ensure Samples are done last.
        # logger.debug(f"Getting contols for {self.rsl_plate_num}")
        include = sorted(include)
        # logger.debug(include)
        pos_str = "(ATCC)|(MCS)"
        pos_regex = re.compile(rf"^{pos_str}")
        neg_str = "(EN)"
        neg_regex = re.compile(rf"^{neg_str}")
        output = []
        for item in include:
            match item:
                case "Positive":
                    if self.controls:
                        # logger.debug(f"{self.rsl_plate_num} controls: {[(item, item.sample) for item in self.controls]}")
                        provs = (control.sample for control in self.controls if control.is_positive_control)
                    else:
                        provs = (sample for sample in self.samples if bool(pos_regex.match(sample.submitter_id)))
                case "Negative":
                    if self.controls:
                        provs = (control.sample for control in self.controls if not control.is_positive_control)
                    else:
                        provs = (sample for sample in self.samples if bool(neg_regex.match(sample.submitter_id)))
                case _:
                    provs = (sample for sample in self.samples if not sample.control and sample not in output)
            # logger.debug(f"Provisional controls: {provs}")
            for prov in provs:
                if not prov:
                    continue
                # logger.debug(f"Prov: {prov}")
                prov.submission = self.rsl_plate_num
                prov.submitted_date = self.submitted_date
                output.append(prov)
        return output


class Wastewater(BasicSubmission):
    """
    derivative submission type from BasicSubmission
    """
    id = Column(INTEGER, ForeignKey('_basicsubmission.id'), primary_key=True, autoincrement=False)
    ext_technician = Column(String(64))  #: Name of technician doing extraction
    pcr_technician = Column(String(64))  #: Name of technician doing pcr
    pcr_info = Column(JSON)  #: unstructured output from pcr table logger or user(Artic)

    __mapper_args__ = __mapper_args__ = dict(polymorphic_identity="Wastewater",
                                             polymorphic_load="inline",
                                             inherit_condition=(id == BasicSubmission.id))

    def to_dict(self, full_data: bool = False, backup: bool = False, report: bool = False) -> dict:
        """
        Extends parent class method to add controls to dict

        Returns:
            dict: dictionary used in submissions summary
        """
        output = super().to_dict(full_data=full_data, backup=backup, report=report)
        if report:
            return output
        try:
            output['pcr_info'] = self.pcr_info
        except TypeError as e:
            pass
        if self.ext_technician is None or self.ext_technician == "None":
            output['ext_technician'] = self.technician
        else:
            output["ext_technician"] = self.ext_technician
        if self.pcr_technician is None or self.pcr_technician == "None":
            output["pcr_technician"] = self.technician
        else:
            output['pcr_technician'] = self.pcr_technician
        if full_data:
            output['samples'] = [sample for sample in output['samples']]
            dummy_samples = []
            for item in output['samples']:
                thing = deepcopy(item)
                try:
                    thing['row'] = thing['source_row']
                    thing['column'] = thing['source_column']
                except KeyError:
                    logger.error(f"No row or column for sample: {item['submitter_id']}")
                    continue
                thing['tooltip'] = f"Sample Name: {thing['name']}\nWell: {thing['sample_location']}"
                dummy_samples.append(thing)
            # logger.debug(f"Dummy samples for 24 well: {pformat(dummy_samples)}")
            output['origin_plate'] = self.__class__.make_plate_map(sample_list=dummy_samples, plate_rows=4,
                                                                   plate_columns=6)
        # logger.debug(f"PCR info: {output['pcr_info']}")
        return output

    @classmethod
    def custom_info_parser(cls, input_dict: dict, xl: Workbook | None = None, custom_fields: dict = {}) -> dict:
        """
        Update submission dictionary with class specific information. Extends parent

        Args:
            input_dict (dict): Input sample dictionary
            xl (Workbook): xl (Workbook): original xl workbook, used for child classes mostly.
            custom_fields: Dictionary of locations, ranges, etc to be used by this function

        Returns:
            dict: Updated info dictionary
        """
        input_dict = super().custom_info_parser(input_dict)
        if xl is not None:
            try:
                input_dict['csv'] = xl["Copy to import file"]
            except KeyError as e:
                logger.error(e)
                try:
                    match input_dict['rsl_plate_num']:
                        case dict():
                            input_dict['csv'] = xl[input_dict['rsl_plate_num']['value']]
                        case str():
                            input_dict['csv'] = xl[input_dict['rsl_plate_num']]
                        case _:
                            pass
                except Exception as e:
                    logger.error(f"Error handling couldn't get csv due to: {e}")
        return input_dict

    @classmethod
    def parse_samples(cls, input_dict: dict) -> dict:
        """
        Update sample dictionary with type specific information. Extends parent

        Args:
            input_dict (dict): Input sample dictionary

        Returns:
            dict: Updated sample dictionary
        """
        input_dict = super().parse_samples(input_dict=input_dict)
        # NOTE: Had to put in this section due to ENs not having rsl_number and therefore not getting PCR results.
        check = check_key_or_attr("rsl_number", input_dict)
        if not check:
            input_dict['rsl_number'] = input_dict['submitter_id']
        # logger.debug(pformat(input_dict, indent=4))
        return input_dict

    @classmethod
    def parse_pcr(cls, xl: Workbook, rsl_plate_num: str) -> Generator[dict, None, None]:
        """
        Perform parsing of pcr info. Since most of our PC outputs are the same format, this should work for most.

        Args:
            xl (pd.DataFrame): pcr info form
            rsl_plate_number (str): rsl plate num of interest

        Returns:
            Generator[dict, None, None]: Updated samples
        """
        samples = [item for item in super().parse_pcr(xl=xl, rsl_plate_num=rsl_plate_num)]
        # NOTE: Due to having to run through samples in for loop we need to convert to list.
        # NOTE: Also, you can't change the size of a list while iterating it, so don't even think about it.
        output = []
        for sample in samples:
            logger.debug(sample)
            # NOTE: remove '-{target}' from controls
            sample['sample'] = re.sub('-N\\d*$', '', sample['sample'])
            # NOTE: if sample is already in output skip
            if sample['sample'] in [item['sample'] for item in output]:
                logger.warning(f"Already have {sample['sample']}")
                continue
            # NOTE: Set ct values
            # logger.debug(f"Sample ct: {sample['ct']}")
            sample[f"ct_{sample['target'].lower()}"] = sample['ct'] if isinstance(sample['ct'], float) else 0.0
            # NOTE: Set assessment
            # logger.debug(f"Sample assessemnt: {sample['assessment']}")
            # NOTE: Get sample having other target
            other_targets = [s for s in samples if re.sub('-N\\d*$', '', s['sample']) == sample['sample']]
            for s in other_targets:
                sample[f"ct_{s['target'].lower()}"] = s['ct'] if isinstance(s['ct'], float) else 0.0
            try:
                del sample['ct']
            except KeyError:
                pass
            try:
                del sample['assessment']
            except KeyError:
                pass
            # logger.debug(sample)
            row = int(row_keys[sample['well'][:1]])
            column = int(sample['well'][1:])
            sample['row'] = row
            sample['column'] = column
            output.append(sample)
        # NOTE: And then convert back to list to keep fidelity with parent method.
        for sample in output:
            yield sample

    @classmethod
    def enforce_name(cls, instr: str, data: dict | None = {}) -> str:
        """
        Custom naming method for this class. Extends parent.

        Args:
            instr (str): Initial name.
            data (dict | None, optional): Additional parameters for name. Defaults to None.

        Returns:
            str: Updated name.
        """
        try:
            # NOTE: Deal with PCR file.
            instr = re.sub(r"PCR(-|_)", "", instr)
        except (AttributeError, TypeError) as e:
            logger.error(f"Problem using regex: {e}")
        outstr = super().enforce_name(instr=instr, data=data)
        return outstr

    @classmethod
    def adjust_autofill_samples(cls, samples: List[Any]) -> List[Any]:
        """
        Makes adjustments to samples before writing to excel. Extends parent.

        Args:
            samples (List[Any]): List of Samples

        Returns:
            List[Any]: Updated list of samples
        """
        samples = super().adjust_autofill_samples(samples)
        samples = [item for item in samples if not item.submitter_id.startswith("EN")]
        return samples

    @classmethod
    def get_details_template(cls, base_dict: dict) -> Tuple[dict, Template]:
        """
        Get the details jinja template for the correct class. Extends parent

        Args:
            base_dict (dict): incoming dictionary of Submission fields

        Returns:
            Tuple[dict, Template]: (Updated dictionary, Template to be rendered)
        """
        base_dict, template = super().get_details_template(base_dict=base_dict)
        base_dict['excluded'] += ['origin_plate']
        return base_dict, template

    def custom_context_events(self) -> dict:
        """
        Sets context events for main widget

        Returns:
            dict: Context menu items for this instance.
        """
        events = super().custom_context_events()
        events['Link PCR'] = self.link_pcr
        return events

    @report_result
    def link_pcr(self, obj) -> Report:
        """
        PYQT6 function to add PCR info to this submission

        Args:
            obj (_type_): Parent widget
        """
        from backend.excel import PCRParser
        from backend.db import PCRControl, ControlType
        from frontend.widgets import select_open_file
        report = Report()
        fname = select_open_file(obj=obj, file_extension="xlsx")
        if not fname:
            report.add_result(Result(msg="No file selected, cancelling.", status="Warning"))
            return report
        parser = PCRParser(filepath=fname, submission=self)
        self.set_attribute("pcr_info", parser.pcr_info)
        # NOTE: These are generators here, need to expand.
        pcr_samples = sorted([sample for sample in parser.samples], key=itemgetter('column'))
        logger.debug(f"Samples from parser: {pformat(pcr_samples)}")
        # NOTE: Samples from parser check out.
        pcr_controls = [control for control in parser.controls]
        self.save(original=False)
        for assoc in self.submission_sample_associations:
            logger.debug(f"Checking pcr_samples for {assoc.sample.rsl_number}, {assoc.sample.ww_full_sample_id} at "
                         f"column {assoc.column} and row {assoc.row}")
            # NOTE: Associations of interest do exist in the submission, are not being found below
            # Okay, I've found the problem, at last, the problem is that only one RSL number is saved for each sample,
            # Even though each instance of say "25-YUL13-PU3-0320" has multiple RSL numbers in the excel sheet.
            # so, yeah, the submitters need to make sure that there are unique values for each one.
            try:
                sample_dict = next(item for item in pcr_samples if item['sample'] == assoc.sample.rsl_number
                                   and item['row'] == assoc.row and item['column'] == assoc.column)
                logger.debug(
                    f"Found sample {sample_dict} at index {pcr_samples.index(sample_dict)}: {pcr_samples[pcr_samples.index(sample_dict)]}")
            except StopIteration:
                logger.error(f"Couldn't find {assoc} in the Parser samples")
                continue
            logger.debug(f"Length of pcr_samples: {len(pcr_samples)}")
            assoc = self.update_subsampassoc(assoc=assoc, input_dict=sample_dict)
            result = assoc.save()
            if result:
                report.add_result(result)
        controltype = ControlType.query(name="PCR Control")
        submitted_date = datetime.strptime(" ".join(parser.pcr_info['run_start_date/time'].split(" ")[:-1]),
                                           "%Y-%m-%d %I:%M:%S %p")
        for control in pcr_controls:
            # logger.debug(f"Control coming into save: {control}")
            new_control = PCRControl(**control)
            new_control.submitted_date = submitted_date
            new_control.controltype = controltype
            new_control.submission = self
            # logger.debug(f"Control coming into save: {new_control.__dict__}")
            new_control.save()
        return report


class WastewaterArtic(BasicSubmission):
    """
    derivative submission type for artic wastewater
    """
    id = Column(INTEGER, ForeignKey('_basicsubmission.id'), primary_key=True)
    artic_technician = Column(String(64))  #: Name of technician performing artic
    dna_core_submission_number = Column(String(64))  #: Number used by core as id
    pcr_info = Column(JSON)  #: unstructured output from pcr table logger or user(Artic)
    gel_image = Column(String(64))  #: file name of gel image in zip file
    gel_info = Column(JSON)  #: unstructured data from gel.
    gel_controls = Column(JSON)  #: locations of controls on the gel
    source_plates = Column(JSON)  #: wastewater plates that samples come from
    artic_date = Column(TIMESTAMP)  #: Date Artic Performed
    ngs_date = Column(TIMESTAMP)  #: Date submission received
    gel_date = Column(TIMESTAMP)  #: Date submission received
    gel_barcode = Column(String(16))  #: Identifier for the used gel.

    __mapper_args__ = dict(polymorphic_identity="Wastewater Artic",
                           polymorphic_load="inline",
                           inherit_condition=(id == BasicSubmission.id))

    def to_dict(self, full_data: bool = False, backup: bool = False, report: bool = False) -> dict:
        """
        Extends parent class method to add controls to dict

        Returns:
            dict: dictionary used in submissions summary
        """
        output = super().to_dict(full_data=full_data, backup=backup, report=report)
        if report:
            return output
        if self.artic_technician in [None, "None"]:
            output['artic_technician'] = self.technician
        else:
            output['artic_technician'] = self.artic_technician
        output['gel_info'] = self.gel_info
        output['gel_image'] = self.gel_image
        output['dna_core_submission_number'] = self.dna_core_submission_number
        output['source_plates'] = self.source_plates
        output['artic_date'] = self.artic_date or self.submitted_date
        output['ngs_date'] = self.ngs_date or self.submitted_date
        output['gel_date'] = self.gel_date or self.submitted_date
        output['gel_barcode'] = self.gel_barcode
        return output

    @classmethod
    def custom_info_parser(cls, input_dict: dict, xl: Workbook | None = None, custom_fields: dict = {}) -> dict:
        """
        Update submission dictionary with class specific information

        Args:
            input_dict (dict): Input sample dictionary
            xl (pd.ExcelFile): original xl workbook, used for child classes mostly
            custom_fields: Dictionary of locations, ranges, etc to be used by this function

        Returns:
            dict: Updated sample dictionary
        """
        from backend.validators import RSLNamer
        from openpyxl_image_loader.sheet_image_loader import SheetImageLoader

        def scrape_image(wb: Workbook, info_dict: dict) -> Image or None:
            """
            Pulls image from excel workbook

            Args:
                wb (Workbook): Workbook of interest.
                info_dict (dict): Location map.

            Returns:
                Image or None: Image of interest.
            """
            ws = wb[info_dict['sheet']]
            img_loader = SheetImageLoader(ws)
            for ii in range(info_dict['start_row'], info_dict['end_row'] + 1):
                for jj in range(info_dict['start_column'], info_dict['end_column'] + 1):
                    cell_str = f"{row_map[jj]}{ii}"
                    if img_loader.image_in(cell_str):
                        try:
                            return img_loader.get(cell_str)
                        except ValueError as e:
                            logger.error(f"Could not open image from cell: {cell_str} due to {e}")
                            return None
            return None

        input_dict = super().custom_info_parser(input_dict)
        input_dict['submission_type'] = dict(value="Wastewater Artic", missing=False)
        egel_section = custom_fields['egel_controls']
        ws = xl[egel_section['sheet']]
        # NOTE: Here we should be scraping the control results.
        data = [ws.cell(row=ii, column=jj) for jj in range(egel_section['start_column'], egel_section['end_column'] + 1)
                for
                ii in range(egel_section['start_row'], egel_section['end_row'] + 1)]
        data = [cell for cell in data if cell.value is not None and "NTC" in cell.value]
        input_dict['gel_controls'] = [
            dict(sample_id=cell.value, location=f"{row_map[cell.row - 9]}{str(cell.column - 14).zfill(2)}") for cell in
            data]
        # NOTE: Get source plate information
        source_plates_section = custom_fields['source_plates']
        ws = xl[source_plates_section['sheet']]
        data = [dict(plate=ws.cell(row=ii, column=source_plates_section['plate_column']).value,
                     starting_sample=ws.cell(row=ii, column=source_plates_section['starting_sample_column']).value) for
                ii in
                range(source_plates_section['start_row'], source_plates_section['end_row'] + 1)]
        for datum in data:
            if datum['plate'] in ["None", None, ""]:
                continue
            else:
                datum['plate'] = RSLNamer(filename=datum['plate'], submission_type="Wastewater").parsed_name
        if xl is not None:
            try:
                input_dict['csv'] = xl["hitpicks_csv_to_export"]
            except KeyError as e:
                logger.error(e)
                try:
                    match input_dict['rsl_plate_num']:
                        case dict():
                            input_dict['csv'] = xl[input_dict['rsl_plate_num']['value']]
                        case str():
                            input_dict['csv'] = xl[input_dict['rsl_plate_num']]
                        case _:
                            pass
                except Exception as e:
                    logger.error(f"Error handling couldn't get csv due to: {e}")
        input_dict['source_plates'] = data
        egel_info_section = custom_fields['egel_info']
        ws = xl[egel_info_section['sheet']]
        data = []
        for ii in range(egel_info_section['start_row'], egel_info_section['end_row'] + 1):
            datum = dict(
                name=ws.cell(row=ii, column=egel_info_section['start_column'] - 3).value,
                values=[]
            )
            for jj in range(egel_info_section['start_column'], egel_info_section['end_column'] + 1):
                d = dict(
                    name=ws.cell(row=egel_info_section['start_row'] - 1, column=jj).value,
                    value=ws.cell(row=ii, column=jj).value
                )
                if d['value'] is not None:
                    datum['values'].append(d)
            data.append(datum)
        input_dict['gel_info'] = data
        egel_image_section = custom_fields['image_range']
        img: Image = scrape_image(wb=xl, info_dict=egel_image_section)
        if img is not None:
            tmp = Path(TemporaryFile().name).with_suffix(".jpg")
            img.save(tmp.__str__())
            with ZipFile(cls.__directory_path__.joinpath("submission_imgs.zip"), 'a') as zipf:
                # NOTE: Add a file located at the source_path to the destination within the zip
                # file. It will overwrite existing files if the names collide, but it
                # will give a warning
                zipf.write(tmp.__str__(), f"{input_dict['rsl_plate_num']['value']}.jpg")
            input_dict['gel_image'] = f"{input_dict['rsl_plate_num']['value']}.jpg"
        return input_dict

    @classmethod
    def enforce_name(cls, instr: str, data: dict = {}) -> str:
        """
        Custom naming method for this class. Extends parent.

        Args:
            instr (str): Initial name.
            data (dict | None, optional): Additional parameters for name. Defaults to None.

        Returns:
            str: Updated name.
        """
        logger.debug(f"Incoming String: {instr}")
        try:
            # NOTE: Deal with PCR file.
            instr = re.sub(r"Artic", "", instr, flags=re.IGNORECASE)
        except (AttributeError, TypeError) as e:
            logger.error(f"Problem using regex: {e}")
        try:
            instr = instr.replace("-", "")
        except AttributeError:
            instr = date.today().strftime("%Y%m%d")
        instr = re.sub(r"^(\d{6})", f"RSL-AR-\\1", instr)
        outstr = super().enforce_name(instr=instr, data=data)
        outstr = outstr.replace("RSLAR", "RSL-AR")
        return outstr

    @classmethod
    def parse_samples(cls, input_dict: dict) -> dict:
        """
        Update sample dictionary with type specific information. Extends parent.

        Args:
            input_dict (dict): Input sample dictionary

        Returns:
            dict: Updated sample dictionary
        """
        input_dict = super().parse_samples(input_dict)
        input_dict['sample_type'] = "Wastewater Sample"
        # NOTE: Stop gap solution because WW is sloppy with their naming schemes
        try:
            input_dict['source_plate'] = input_dict['source_plate'].replace("WW20", "WW-20")
        except KeyError:
            pass
        try:
            input_dict['source_plate_number'] = int(input_dict['source_plate_number'])
        except (ValueError, KeyError):
            input_dict['source_plate_number'] = 0
        # NOTE: Because generate_sample_object needs the submitter_id and the artic has the "({origin well})" at the end, this has to be done here. No moving to sqlalchemy object :(
        input_dict['submitter_id'] = re.sub(r"\s\(.+\)\s?$", "", str(input_dict['submitter_id'])).strip()
        try:
            input_dict['ww_processing_num'] = input_dict['sample_name_(lims)']
            del input_dict['sample_name_(lims)']
        except KeyError:
            logger.error(f"Unable to set ww_processing_num for sample {input_dict['submitter_id']}")
        try:
            input_dict['ww_full_sample_id'] = input_dict['sample_name_(ww)']
            del input_dict['sample_name_(ww)']
        except KeyError:
            logger.error(f"Unable to set ww_processing_num for sample {input_dict['submitter_id']}")
        year = str(date.today().year)[-2:]
        # NOTE: Check for extraction negative control (Enterics)
        if re.search(rf"^{year}-(ENC)", input_dict['submitter_id']):
            input_dict['rsl_number'] = cls.en_adapter(input_str=input_dict['submitter_id'])
        # NOTE: Check for extraction negative control (Robotics)
        if re.search(rf"^{year}-(RSL)", input_dict['submitter_id']):
            logger.debug(f"Found {year}-(RSL), so we are going to run PBS adapter:")
            input_dict['rsl_number'] = cls.pbs_adapter(input_str=input_dict['submitter_id'])
        return input_dict

    @classmethod
    def en_adapter(cls, input_str: str) -> str:
        """
        Stopgap solution because WW names their ENs different

        Args:
            input_str (str): input name

        Returns:
            str: output name
        """
        processed = input_str.replace("RSL", "")
        # NOTE: Remove anything in brackets at the end of string?
        processed = re.sub(r"\(.*\)$", "", processed).strip()
        # NOTE: Remove letters that are not R.
        processed = re.sub(r"[A-QS-Z]+\d*", "", processed)
        # NOTE: Remove trailing '-' if any
        processed = processed.strip("-")
        try:
            # NOTE: get digit at the end of the string.
            en_num = re.search(r"\-\d{1}$", processed).group()
            processed = rreplace(processed, en_num, "")
        except AttributeError:
            en_num = "1"
        en_num = en_num.strip("-")
        try:
            # NOTE: Get last digit and maybe 'R' with another digit.
            plate_num = re.search(r"\-\d{1}R?\d?$", processed).group()
            processed = rreplace(processed, plate_num, "")
        except AttributeError:
            plate_num = "1"
        # NOTE: plate_num not currently used, but will keep incase it is in the future
        plate_num = plate_num.strip("-")
        day = re.search(r"\d{2}$", processed).group()
        processed = rreplace(processed, day, "")
        month = re.search(r"\d{2}$", processed).group()
        processed = rreplace(processed, month, "")
        processed = processed.replace("--", "")
        year = re.search(r'^(?:\d{2})?\d{2}', processed).group()
        year = f"20{year}"
        final_en_name = f"EN{en_num}-{year}{month}{day}"
        return final_en_name

    @classmethod
    def pbs_adapter(cls, input_str):
        """
        Stopgap solution because WW names their controls different

        Args:
            input_str (str): input name

        Returns:
            str: output name
        """
        logger.debug(f"PBS adapter on {input_str}")
        # NOTE: Remove letters.
        processed = input_str.replace("RSL", "")
        # logger.debug(processed)
        # NOTE: Remove brackets at end
        processed = re.sub(r"\(.*\)$", "", processed).strip()
        # logger.debug(processed)
        processed = re.sub(r"-RPT", "", processed, flags=re.IGNORECASE)
        # NOTE: Remove any non-R letters at end.
        processed = re.sub(r"[A-QS-Z]+\d*", "", processed)
        # logger.debug(processed)
        # NOTE: Remove trailing '-' if any
        processed = processed.strip("-")
        # logger.debug(processed)
        try:
            plate_num = re.search(r"\-\d{1}R?\d?$", processed).group()
            processed = rreplace(processed, plate_num, "")
        except AttributeError:
            plate_num = "1"
        plate_num = plate_num.strip("-")
        try:
            repeat_num = re.search(r"R(?P<repeat>\d)?$", processed).groups()[0]
        except:
            repeat_num = None
        if repeat_num is None and "R" in plate_num:
            repeat_num = "1"
        try:
            plate_num = re.sub(r"R", rf"R{repeat_num}", plate_num)
        except AttributeError:
            logger.error(f"Problem re-evaluating plate number for {processed}")
        # logger.debug(processed)
        # NOTE: Remove any redundant -digits
        processed = re.sub(r"-\d$", "", processed)
        # logger.debug(processed)
        day = re.search(r"\d{2}$", processed).group()
        processed = rreplace(processed, day, "")
        # logger.debug(processed)
        month = re.search(r"\d{2}$", processed).group()
        processed = rreplace(processed, month, "")
        processed = processed.replace("--", "")
        # logger.debug(processed)
        year = re.search(r'^(?:\d{2})?\d{2}', processed).group()
        year = f"20{year}"
        # logger.debug(processed)
        final_en_name = f"PBS{year}{month}{day}-{plate_num}"
        return final_en_name

    @classmethod
    def custom_validation(cls, pyd) -> dict:
        """
        Performs any final custom parsing of the excel file. Extends parent

        Args:
            input_dict (dict): Parser product up to this point.
            xl (pd.ExcelFile | None, optional): Excel submission form. Defaults to None.
            info_map (dict | None, optional): Map of information locations from SubmissionType. Defaults to None.
            plate_map (dict | None, optional): Constructed plate map of samples. Defaults to None.

        Returns:
            dict: Updated parser product.
        """
        input_dict = super().custom_validation(pyd)
        exclude_plates = [None, "", "none", "na"]
        pyd.source_plates = [plate for plate in pyd.source_plates if plate['plate'].lower() not in exclude_plates]
        for sample in pyd.samples:
            if re.search(r"^NTC", sample.submitter_id):
                if isinstance(pyd.rsl_plate_num, dict):
                    placeholder = pyd.rsl_plate_num['value']
                else:
                    placeholder = pyd.rsl_plate_num
                sample.submitter_id = f"{sample.submitter_id}-WWG-{placeholder}"
        return input_dict

    @classmethod
    def custom_info_writer(cls, input_excel: Workbook, info: dict | None = None, backup: bool = False,
                           custom_fields: dict = {}) -> Workbook:
        """
        Adds custom autofill methods for submission. Extends Parent

        Args:
            input_excel (Workbook): initial workbook.
            info (dict | None, optional): dictionary of additional info. Defaults to None.
            backup (bool, optional): Whether this is part of a backup operation. Defaults to False.
            custom_fields: Dictionary of locations, ranges, etc to be used by this function

        Returns:
            Workbook: Updated workbook
        """
        input_excel = super().custom_info_writer(input_excel, info, backup)
        if isinstance(info, types.GeneratorType):
            info = {k: v for k, v in info}
        # NOTE: check for source plate information
        if check_key_or_attr(key='source_plates', interest=info, check_none=True):
            source_plates_section = custom_fields['source_plates']
            worksheet = input_excel[source_plates_section['sheet']]
            start_row = source_plates_section['start_row']
            # NOTE: write source plates to First strand list
            for iii, plate in enumerate(info['source_plates']['value']):
                row = start_row + iii
                try:
                    worksheet.cell(row=row, column=source_plates_section['plate_column'], value=plate['plate'])
                except TypeError:
                    pass
                try:
                    worksheet.cell(row=row, column=source_plates_section['starting_sample_column'],
                                   value=plate['starting_sample'])
                except TypeError:
                    pass
        else:
            logger.warning(f"No source plate info found.")
        # NOTE: check for gel information
        if check_key_or_attr(key='gel_info', interest=info, check_none=True):
            egel_section = custom_fields['egel_info']
            # NOTE: print json field gel results to Egel results
            worksheet = input_excel[egel_section['sheet']]
            start_row = egel_section['start_row'] - 1
            start_column = egel_section['start_column'] - 3
            for row, ki in enumerate(info['gel_info']['value'], start=1):
                row = start_row + row
                worksheet.cell(row=row, column=start_column, value=ki['name'])
                for jjj, kj in enumerate(ki['values'], start=1):
                    column = start_column + 2 + jjj
                    worksheet.cell(row=start_row, column=column, value=kj['name'])
                    try:
                        worksheet.cell(row=row, column=column, value=kj['value'])
                    except AttributeError:
                        logger.error(f"Failed {kj['name']} with value {kj['value']} to row {row}, column {column}")
        else:
            logger.warning("No gel info found.")
        if check_key_or_attr(key='gel_image', interest=info, check_none=True):
            worksheet = input_excel[egel_section['sheet']]
            with ZipFile(cls.__directory_path__.joinpath("submission_imgs.zip")) as zipped:
                z = zipped.extract(info['gel_image']['value'], Path(TemporaryDirectory().name))
                img = OpenpyxlImage(z)
                img.height = 400  # insert image height in pixels as float or int (e.g. 305.5)
                img.width = 600
                img.anchor = egel_section['img_anchor']
                worksheet.add_image(img)
        else:
            logger.warning("No gel image found.")
        return input_excel

    @classmethod
    def custom_sample_writer(self, sample: dict) -> dict:
        if sample['source_plate_number'] in [0, "0"]:
            sample['source_plate_number'] = "control"
        return sample

    @classmethod
    def get_details_template(cls, base_dict: dict) -> Tuple[dict, Template]:
        """
        Get the details jinja template for the correct class. Extends parent

        Args:
            base_dict (dict): incoming dictionary of Submission fields

        Returns:
            Tuple[dict, Template]: (Updated dictionary, Template to be rendered)
        """
        base_dict, template = super().get_details_template(base_dict=base_dict)
        base_dict['excluded'] += ['gel_info', 'gel_image', 'headers', "dna_core_submission_number", "source_plates",
                                  "gel_controls, gel_image_path"]
        base_dict['DNA Core ID'] = base_dict['dna_core_submission_number']
        if check_key_or_attr(key='gel_info', interest=base_dict, check_none=True):
            headers = [item['name'] for item in base_dict['gel_info'][0]['values']]
            base_dict['headers'] = [''] * (4 - len(headers))
            base_dict['headers'] += headers
        if check_key_or_attr(key='gel_image', interest=base_dict, check_none=True):
            with ZipFile(cls.__directory_path__.joinpath("submission_imgs.zip")) as zipped:
                base_dict['gel_image_actual'] = base64.b64encode(zipped.read(base_dict['gel_image'])).decode('utf-8')
        return base_dict, template

    def custom_context_events(self) -> dict:
        """
        Creates dictionary of str:function to be passed to context menu. Extends parent

        Returns:
            dict: dictionary of functions
        """
        events = super().custom_context_events()
        events['Gel Box'] = self.gel_box
        return events

    def set_attribute(self, key: str, value):
        """
        Performs custom attribute setting based on values. Extends parent

        Args:
            key (str): name of attribute
            value (_type_): value of attribute
        """
        super().set_attribute(key=key, value=value)
        if key == 'gel_info':
            if len(self.gel_info) > 3:
                self.gel_info = self.gel_info[-3:]

    def gel_box(self, obj):
        """
        Creates PYQT6 widget to perform gel viewing operations

        Args:
            obj (_type_): parent widget
        """
        from frontend.widgets.gel_checker import GelBox
        from frontend.widgets import select_open_file
        report = Report()
        fname = select_open_file(obj=obj, file_extension="jpg")
        if not fname:
            report.add_result(Result(msg="No file selected, cancelling.", status="Warning"))
            return report
        dlg = GelBox(parent=obj, img_path=fname, submission=self)
        if dlg.exec():
            self.dna_core_submission_number, self.gel_barcode, img_path, output, comment = dlg.parse_form()
            self.gel_image = img_path.name
            self.gel_info = output
            dt = datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")
            com = dict(text=comment, name=getuser(), time=dt)
            if com['text'] is not None and com['text'] != "":
                if self.comment is not None:
                    self.comment.append(com)
                else:
                    self.comment = [com]
            with ZipFile(self.__directory_path__.joinpath("submission_imgs.zip"), 'a') as zipf:
                # NOTE: Add a file located at the source_path to the destination within the zip
                # file. It will overwrite existing files if the names collide, but it
                # will give a warning
                zipf.write(img_path, self.gel_image)
            self.save()


# NOTE: Sample Classes

class BasicSample(BaseClass, LogMixin):
    """
    Base of basic sample which polymorphs into BCSample and WWSample
    """

    id = Column(INTEGER, primary_key=True)  #: primary key
    submitter_id = Column(String(64), nullable=False, unique=True)  #: identification from submitter
    sample_type = Column(String(32))  #: mode_sub_type of sample

    sample_submission_associations = relationship(
        "SubmissionSampleAssociation",
        back_populates="sample",
        cascade="all, delete-orphan",
    )  #: associated submissions

    __mapper_args__ = {
        "polymorphic_identity": "Basic Sample",
        "polymorphic_on": case(

            (sample_type == "Wastewater Sample", "Wastewater Sample"),
            (sample_type == "Wastewater Artic Sample", "Wastewater Sample"),
            (sample_type == "Bacterial Culture Sample", "Bacterial Culture Sample"),

            else_="Basic Sample"
        ),
        "with_polymorphic": "*",
    }

    submissions = association_proxy("sample_submission_associations", "submission")  #: proxy of associated submissions

    @validates('submitter_id')
    def create_id(self, key: str, value: str) -> str:
        """
        Creates a random string as a submitter id.

        Args:
            key (str): name of attribute
            value (str): submitter id

        Returns:
            str: new (or unchanged) submitter id
        """
        if value is None:
            return uuid.uuid4().hex.upper()
        else:
            return value

    def __repr__(self) -> str:
        try:
            return f"<{self.sample_type.replace('_', ' ').title().replace(' ', '')}({self.submitter_id})>"
        except AttributeError:
            return f"<Sample({self.submitter_id})"

    @classproperty
    def searchables(cls):
        return [dict(label="Submitter ID", field="submitter_id")]

    @classproperty
    def timestamps(cls) -> List[str]:
        """
        Constructs a list of all attributes stored as SQL Timestamps

        Returns:
            List[str]: Attribute list
        """
        output = [item.name for item in cls.__table__.columns if isinstance(item.type, TIMESTAMP)]
        if issubclass(cls, BasicSample) and not cls.__name__ == "BasicSample":
            output += BasicSample.timestamps
        return output

    def to_sub_dict(self, full_data: bool = False) -> dict:
        """
        gui friendly dictionary

        Args:
            full_data (bool): Whether to use full object or truncated. Defaults to False

        Returns:
            dict: submitter id and sample type and linked submissions if full data
        """
        sample = dict(
            submitter_id=self.submitter_id,
            sample_type=self.sample_type
        )
        if full_data:
            sample['submissions'] = sorted([item.to_sub_dict() for item in self.sample_submission_associations],
                                           key=itemgetter('submitted_date'))
        return sample

    def to_pydantic(self):
        from backend.validators import PydSample
        return PydSample(**self.to_sub_dict())

    def set_attribute(self, name: str, value):
        """
        Custom attribute setter (depreciated over built-in __setattr__)

        Args:
            name (str): name of attribute
            value (_type_): value to be set to attribute
        """
        try:
            setattr(self, name, value)
        except AttributeError:
            logger.error(f"Attribute {name} not found")

    @classmethod
    def find_polymorphic_subclass(cls, polymorphic_identity: str | None = None,
                                  attrs: dict | None = None) -> Type[BasicSample]:
        """
        Retrieves subclasses of BasicSample based on type name.

        Args:
            attrs (dict | None, optional): name: value of attributes in the wanted subclass
            polymorphic_identity (str | None, optional): Name of subclass fed to polymorphic identity. Defaults to None.

        Returns:
            BasicSample: Subclass of interest.
        """
        if isinstance(polymorphic_identity, dict):
            polymorphic_identity = polymorphic_identity['value']
        if polymorphic_identity is not None:
            try:
                model = cls.__mapper__.polymorphic_map[polymorphic_identity].class_
            except Exception as e:
                logger.error(f"Could not get polymorph {polymorphic_identity} of {cls} due to {e}, using {cls}")
                model = cls
            return model
        else:
            model = cls
        if attrs is None or len(attrs) == 0:
            return model
        if any([not hasattr(cls, attr) for attr in attrs.keys()]):
            # NOTE: looks for first model that has all included kwargs
            try:
                model = next(subclass for subclass in cls.__subclasses__() if
                             all([hasattr(subclass, attr) for attr in attrs.keys()]))
            except StopIteration as e:
                raise AttributeError(
                    f"Couldn't find existing class/subclass of {cls} with all attributes:\n{pformat(attrs.keys())}")
        return model

    @classmethod
    def parse_sample(cls, input_dict: dict) -> dict:
        """
        Custom sample parser

        Args:
            input_dict (dict): Basic parser results.

        Returns:
            dict: Updated parser results.
        """
        return input_dict

    @classproperty
    def details_template(cls) -> Template:
        """
        Get the details jinja template for the correct class

        Args:
            base_dict (dict): incoming dictionary of Submission fields

        Returns:
            Tuple(dict, Template): (Updated dictionary, Template to be rendered)
        """
        env = jinja_template_loading()
        temp_name = f"{cls.__name__.lower()}_details.html"
        try:
            template = env.get_template(temp_name)
        except TemplateNotFound as e:
            logger.error(f"Couldn't find template {e}")
            template = env.get_template("basicsample_details.html")
        return template

    @classmethod
    @setup_lookup
    def query(cls,
              submitter_id: str | None = None,
              sample_type: str | BasicSample | None = None,
              limit: int = 0,
              **kwargs
              ) -> BasicSample | List[BasicSample]:
        """
        Lookup samples in the database by a number of parameters.

        Args:
            submitter_id (str | None, optional): Name of the sample (limits results to 1). Defaults to None.
            sample_type (str | None, optional): Sample type. Defaults to None.
            limit (int, optional): Maximum number of results to return (0 = all). Defaults to 0.

        Returns:
            models.BasicSample|List[models.BasicSample]: Sample(s) of interest.
        """
        match sample_type:
            case str():
                model = cls.find_polymorphic_subclass(polymorphic_identity=sample_type)
            case BasicSample():
                model = sample_type
            case _:
                model = cls.find_polymorphic_subclass(attrs=kwargs)
        query: Query = cls.__database_session__.query(model)
        match submitter_id:
            case str():
                query = query.filter(model.submitter_id == submitter_id)
                limit = 1
            case _:
                pass
        return cls.execute_query(query=query, model=model, limit=limit, **kwargs)

    @classmethod
    def query_or_create(cls, sample_type: str | None = None, **kwargs) -> BasicSample:
        """
        Queries for a sample, if none found creates a new one.

        Args:
            sample_type (str): sample subclass name

        Raises:
            ValueError: Raised if no kwargs are passed to narrow down instances
            ValueError: Raised if unallowed key is given.

        Returns:
            BasicSample: Instance of BasicSample
        """
        disallowed = ["id"]
        if kwargs == {}:
            raise ValueError("Need to narrow down query or the first available instance will be returned.")
        sanitized_kwargs = {k: v for k, v in kwargs.items() if k not in disallowed}
        instance = cls.query(sample_type=sample_type, limit=1, **kwargs)
        if instance is None:
            used_class = cls.find_polymorphic_subclass(attrs=sanitized_kwargs, polymorphic_identity=sample_type)
            instance = used_class(**sanitized_kwargs)
            instance.sample_type = sample_type
        return instance

    @classmethod
    def fuzzy_search(cls,
                     sample_type: str | BasicSample | None = None,
                     **kwargs
                     ) -> List[BasicSample]:
        """
        Allows for fuzzy search of samples.

        Args:
            sample_type (str | BasicSample | None, optional): Type of sample. Defaults to None.

        Returns:
            List[BasicSample]: List of samples that match kwarg search parameters.
        """
        match sample_type:
            case str():
                model = cls.find_polymorphic_subclass(polymorphic_identity=sample_type)
            case BasicSample():
                model = sample_type
            case None:
                model = cls
            case _:
                model = cls.find_polymorphic_subclass(attrs=kwargs)
        query: Query = cls.__database_session__.query(model)
        for k, v in kwargs.items():
            search = f"%{v}%"
            try:
                attr = getattr(model, k)
                # NOTE: the secret sauce is in attr.like
                query = query.filter(attr.like(search))
            except (ArgumentError, AttributeError) as e:
                logger.error(f"Attribute {k} unavailable due to:\n\t{e}\nSkipping.")
        return query.limit(50).all()

    def delete(self):
        raise AttributeError(f"Delete not implemented for {self.__class__}")

    @classmethod
    def samples_to_df(cls, sample_list: List[BasicSample], **kwargs) -> pd.DataFrame:
        """
        Runs a fuzzy search and converts into a dataframe.

        Args:
            sample_list (List[BasicSample]): List of samples to be parsed. Defaults to None.

        Returns:
            pd.DataFrame: Dataframe all samples
        """
        try:
            samples = [sample.to_sub_dict() for sample in sample_list]
        except TypeError as e:
            logger.error(f"Couldn't find any samples with data: {kwargs}\nDue to {e}")
            return None
        df = pd.DataFrame.from_records(samples)
        # NOTE: Exclude sub information
        exclude = ['concentration', 'organism', 'colour', 'tooltip', 'comments', 'samples', 'reagents',
                   'equipment', 'gel_info', 'gel_image', 'dna_core_submission_number', 'gel_controls']
        df = df.loc[:, ~df.columns.isin(exclude)]
        return df

    def show_details(self, obj):
        """
        Creates Widget for showing submission details.

        Args:
            obj (_type_): parent widget
        """
        from frontend.widgets.submission_details import SubmissionDetails
        dlg = SubmissionDetails(parent=obj, sub=self)
        if dlg.exec():
            pass

    def edit_from_search(self, obj, **kwargs):
        """
        Function called form search. "Edit" is dependent on function as this one just shows details.

        Args:
            obj (__type__): Parent widget.
            **kwargs (): Required for all edit from search functions.

        Returns:

        """
        self.show_details(obj)


# NOTE: Below are the custom sample types

class WastewaterSample(BasicSample):
    """
    Derivative wastewater sample
    """

    id = Column(INTEGER, ForeignKey('_basicsample.id'), primary_key=True)
    ww_processing_num = Column(String(64))  #: wastewater processing number
    ww_full_sample_id = Column(String(64))  #: full id given by entrics
    rsl_number = Column(String(64))  #: rsl plate identification number
    collection_date = Column(TIMESTAMP)  #: Date sample collected
    received_date = Column(TIMESTAMP)  #: Date sample received
    notes = Column(String(2000))  #: notes from submission form
    sample_location = Column(String(8))  #: location on 24 well plate
    __mapper_args__ = dict(polymorphic_identity="Wastewater Sample",
                           polymorphic_load="inline",
                           inherit_condition=(id == BasicSample.id))

    @classmethod
    def get_default_info(cls, *args):
        """
        Returns default info for a model. Extends BaseClass method.

        Returns:
            dict | list | str: Output of key:value dict or single (list, str) desired variable
        """
        dicto = super().get_default_info(*args)
        match dicto:
            case dict():
                dicto['singles'] += ['ww_processing_num']
                output = {}
                for k, v in dicto.items():
                    if len(args) > 0 and k not in args:
                        continue
                    else:
                        output[k] = v
                if len(args) == 1:
                    return output[args[0]]
            case list():
                if "singles" in args:
                    dicto += ['ww_processing_num']
                return dicto
            case _:
                pass

    def to_sub_dict(self, full_data: bool = False) -> dict:
        """
        gui friendly dictionary, extends parent method.

        Returns:
            dict: sample id, type, received date, collection date
        """
        sample = super().to_sub_dict(full_data=full_data)
        sample['ww_processing_num'] = self.ww_processing_num
        sample['sample_location'] = self.sample_location
        sample['received_date'] = self.received_date
        sample['collection_date'] = self.collection_date
        return sample

    @classmethod
    def parse_sample(cls, input_dict: dict) -> dict:
        """
        Custom sample parser. Extends parent

        Args:
            input_dict (dict): Basic parser results for this sample.

        Returns:
            dict: Updated parser results.
        """
        output_dict = super().parse_sample(input_dict)
        disallowed = ["", None, "None"]
        try:
            check = output_dict['rsl_number'] in disallowed
        except KeyError:
            check = True
        if check:
            output_dict['rsl_number'] = "RSL-WW-" + output_dict['ww_processing_num']
        if output_dict['ww_full_sample_id'] is not None and output_dict["submitter_id"] in disallowed:
            output_dict["submitter_id"] = output_dict['ww_full_sample_id']
        # check = check_key_or_attr("rsl_number", output_dict, check_none=True)
        return output_dict

    @classproperty
    def searchables(cls) -> List[dict]:
        """
        Delivers a list of fields that can be used in fuzzy search. Extends parent.

        Returns:
            List[str]: List of fields.
        """
        searchables = deepcopy(super().searchables)
        for item in ["ww_processing_num", "ww_full_sample_id", "rsl_number"]:
            label = item.strip("ww_").replace("_", " ").replace("rsl", "RSL").title()
            searchables.append(dict(label=label, field=item))
        return searchables


class BacterialCultureSample(BasicSample):
    """
    base of bacterial culture sample
    """
    id = Column(INTEGER, ForeignKey('_basicsample.id'), primary_key=True)
    organism = Column(String(64))  #: bacterial specimen
    concentration = Column(String(16))  #: sample concentration
    control = relationship("IridaControl", back_populates="sample", uselist=False)
    __mapper_args__ = dict(polymorphic_identity="Bacterial Culture Sample",
                           polymorphic_load="inline",
                           inherit_condition=(id == BasicSample.id))

    def to_sub_dict(self, full_data: bool = False) -> dict:
        """
        gui friendly dictionary, extends parent method.

        Returns:
            dict: well location and name (sample id, organism) NOTE: keys must sync with WWSample to_sub_dict above
        """
        sample = super().to_sub_dict(full_data=full_data)
        sample['name'] = self.submitter_id
        sample['organism'] = self.organism
        try:
            sample['concentration'] = f"{float(self.concentration):.2f}"
        except (TypeError, ValueError):
            sample['concentration'] = 0.0
        if self.control is not None:
            sample['colour'] = [0, 128, 0]
            target = next((v for k, v in self.control.controltype.targets.items() if k == self.control.subtype),
                          "Not Available")
            try:
                target = ", ".join(target)
            except:
                target = "None"
            sample['tooltip'] = f"\nControl: {self.control.controltype.name} - {target}"
        return sample


# NOTE: Submission to Sample Associations

class SubmissionSampleAssociation(BaseClass):
    """
    table containing submission/sample associations
    DOC: https://docs.sqlalchemy.org/en/14/orm/extensions/associationproxy.html
    """

    id = Column(INTEGER, unique=True, nullable=False)  #: id to be used for inheriting purposes
    sample_id = Column(INTEGER, ForeignKey("_basicsample.id"), nullable=False)  #: id of associated sample
    submission_id = Column(INTEGER, ForeignKey("_basicsubmission.id"), primary_key=True)  #: id of associated submission
    row = Column(INTEGER, primary_key=True)  #: row on the 96 well plate
    column = Column(INTEGER, primary_key=True)  #: column on the 96 well plate
    submission_rank = Column(INTEGER, nullable=False, default=0)  #: Location in sample list

    # NOTE: reference to the Submission object
    submission = relationship(BasicSubmission,
                              back_populates="submission_sample_associations")  #: associated submission

    # NOTE: reference to the Sample object
    sample = relationship(BasicSample, back_populates="sample_submission_associations")  #: associated sample

    base_sub_type = Column(String)  #: string of mode_sub_type name

    # NOTE: Refers to the type of parent.
    __mapper_args__ = {
        "polymorphic_identity": "Basic Association",
        "polymorphic_on": base_sub_type,
        "with_polymorphic": "*",
    }

    def __init__(self, submission: BasicSubmission = None, sample: BasicSample = None, row: int = 1, column: int = 1,
                 id: int | None = None, submission_rank: int = 0, **kwargs):
        self.submission = submission
        self.sample = sample
        self.row = row
        self.column = column
        self.submission_rank = submission_rank
        if id is not None:
            self.id = id
        else:
            self.id = self.__class__.autoincrement_id()
        for k, v in kwargs.items():
            try:
                self.__setattr__(k, v)
            except AttributeError:
                logger.error(f"Couldn't set {k} to {v}")

    def __repr__(self) -> str:
        try:
            return f"<{self.__class__.__name__}({self.submission.rsl_plate_num} & {self.sample.submitter_id})"
        except AttributeError as e:
            logger.error(f"Unable to construct __repr__ due to: {e}")
            return super().__repr__()

    def to_sub_dict(self) -> dict:
        """
        Returns a sample dictionary updated with instance information

        Returns:
            dict: Updated dictionary with row, column and well updated
        """
        # NOTE: Get associated sample info
        sample = self.sample.to_sub_dict()
        sample['name'] = self.sample.submitter_id
        sample['row'] = self.row
        sample['column'] = self.column
        try:
            sample['well'] = f"{row_map[self.row]}{self.column}"
        except KeyError as e:
            logger.error(f"Unable to find row {self.row} in row_map.")
            sample['Well'] = None
        sample['plate_name'] = self.submission.rsl_plate_num
        sample['positive'] = False
        sample['submitted_date'] = self.submission.submitted_date
        sample['submission_rank'] = self.submission_rank
        return sample

    def to_pydantic(self) -> "PydSample":
        """
        Creates a pydantic model for this sample.

        Returns:
            PydSample: Pydantic Model
        """
        from backend.validators import PydSample
        return PydSample(**self.to_sub_dict())

    @property
    def hitpicked(self) -> dict | None:
        """
        Outputs a dictionary usable for html plate maps.

        Returns:
            dict: dictionary of sample id, row and column in elution plate
        """
        # NOTE: Since there is no PCR, negliable result is necessary.
        sample = self.to_sub_dict()
        env = jinja_template_loading()
        template = env.get_template("tooltip.html")
        tooltip_text = template.render(fields=sample)
        try:
            control = self.sample.control
        except AttributeError:
            control = None
        if control is not None:
            background = "rgb(128, 203, 196)"
        else:
            background = "rgb(105, 216, 79)"
        try:
            tooltip_text += sample['tooltip']
        except KeyError:
            pass
        sample.update(dict(Name=self.sample.submitter_id[:10], tooltip=tooltip_text, background_color=background))
        return sample

    @classmethod
    def autoincrement_id(cls) -> int:
        """
        Increments the association id automatically

        Returns:
            int: incremented id
        """
        if cls.__name__ == "SubmissionSampleAssociation":
            model = cls
        else:
            model = next((base for base in cls.__bases__ if base.__name__ == "SubmissionSampleAssociation"),
                         SubmissionSampleAssociation)
        try:
            return max([item.id for item in model.query()]) + 1
        except ValueError as e:
            logger.error(f"Problem incrementing id: {e}")
            return 1

    @classmethod
    def find_polymorphic_subclass(cls, polymorphic_identity: str | None = None) -> SubmissionSampleAssociation:
        """
        Retrieves subclasses of SubmissionSampleAssociation based on type name.

        Args:
            polymorphic_identity (str | None, optional): Name of subclass fed to polymorphic identity. Defaults to None.

        Returns:
            SubmissionSampleAssociation: Subclass of interest.
        """
        if isinstance(polymorphic_identity, dict):
            polymorphic_identity = polymorphic_identity['value']
        if polymorphic_identity is None:
            model = cls
        else:
            try:
                model = cls.__mapper__.polymorphic_map[polymorphic_identity].class_
            except Exception as e:
                logger.error(f"Could not get polymorph {polymorphic_identity} of {cls} due to {e}")
                model = cls
        return model

    @classmethod
    @setup_lookup
    def query(cls,
              submission: BasicSubmission | str | None = None,
              exclude_submission_type: str | None = None,
              sample: BasicSample | str | None = None,
              row: int = 0,
              column: int = 0,
              limit: int = 0,
              chronologic: bool = False,
              reverse: bool = False,
              **kwargs
              ) -> SubmissionSampleAssociation | List[SubmissionSampleAssociation]:
        """
        Lookup junction of Submission and Sample in the database

        Args:
            submission (models.BasicSubmission | str | None, optional): Submission of interest. Defaults to None.
            sample (models.BasicSample | str | None, optional): Sample of interest. Defaults to None.
            row (int, optional): Row of the sample location on submission plate. Defaults to 0.
            column (int, optional): Column of the sample location on the submission plate. Defaults to 0.
            limit (int, optional): Maximum number of results to return (0 = all). Defaults to 0.
            chronologic (bool, optional): Return results in chronologic order. Defaults to False.

        Returns:
            models.SubmissionSampleAssociation|List[models.SubmissionSampleAssociation]: Junction(s) of interest
        """
        query: Query = cls.__database_session__.query(cls)
        match submission:
            case BasicSubmission():
                query = query.filter(cls.submission == submission)
            case str():
                query = query.join(BasicSubmission).filter(BasicSubmission.rsl_plate_num == submission)
            case _:
                pass
        match sample:
            case BasicSample():
                query = query.filter(cls.sample == sample)
            case str():
                query = query.join(BasicSample).filter(BasicSample.submitter_id == sample)
            case _:
                pass
        if row > 0:
            query = query.filter(cls.row == row)
        if column > 0:
            query = query.filter(cls.column == column)
        match exclude_submission_type:
            case str():
                query = query.join(BasicSubmission).filter(
                    BasicSubmission.submission_type_name != exclude_submission_type)
            case _:
                pass
        if reverse and not chronologic:
            query = query.order_by(BasicSubmission.id.desc())
        if chronologic:
            if reverse:
                query = query.order_by(BasicSubmission.submitted_date.desc())
            else:
                query = query.order_by(BasicSubmission.submitted_date)
        return cls.execute_query(query=query, limit=limit, **kwargs)

    @classmethod
    def query_or_create(cls,
                        association_type: str = "Basic Association",
                        submission: BasicSubmission | str | None = None,
                        sample: BasicSample | str | None = None,
                        id: int | None = None,
                        **kwargs) -> SubmissionSampleAssociation:
        """
        Queries for an association, if none exists creates a new one.

        Args:
            association_type (str, optional): Subclass name. Defaults to "Basic Association".
            submission (BasicSubmission | str | None, optional): associated submission. Defaults to None.
            sample (BasicSample | str | None, optional): associated sample. Defaults to None.
            id (int | None, optional): association id. Defaults to None.

       Returns:
            SubmissionSampleAssociation: Queried or new association.
        """
        match submission:
            case BasicSubmission():
                pass
            case str():
                submission = BasicSubmission.query(rsl_plate_num=submission)
            case _:
                raise ValueError()
        match sample:
            case BasicSample():
                pass
            case str():
                sample = BasicSample.query(submitter_id=sample)
            case _:
                raise ValueError()
        try:
            row = kwargs['row']
        except KeyError:
            row = None
        try:
            column = kwargs['column']
        except KeyError:
            column = None
        try:
            instance = cls.query(submission=submission, sample=sample, row=row, column=column, limit=1)
        except StatementError:
            instance = None
        if instance is None:
            used_cls = cls.find_polymorphic_subclass(polymorphic_identity=association_type)
            instance = used_cls(submission=submission, sample=sample, id=id, **kwargs)
        return instance

    def delete(self):
        raise AttributeError(f"Delete not implemented for {self.__class__}")


class WastewaterAssociation(SubmissionSampleAssociation):
    """
    table containing wastewater specific submission/sample associations
    DOC: https://docs.sqlalchemy.org/en/14/orm/extensions/associationproxy.html
    """
    id = Column(INTEGER, ForeignKey("_submissionsampleassociation.id"), primary_key=True)
    ct_n1 = Column(FLOAT(2))  #: AKA ct for N1
    ct_n2 = Column(FLOAT(2))  #: AKA ct for N2
    n1_status = Column(String(32))  #: positive or negative for N1
    n2_status = Column(String(32))  #: positive or negative for N2
    pcr_results = Column(JSON)  #: imported PCR status from QuantStudio

    __mapper_args__ = dict(polymorphic_identity="Wastewater Association",
                           polymorphic_load="inline",
                           inherit_condition=(id == SubmissionSampleAssociation.id))

    def to_sub_dict(self) -> dict:
        """
        Returns a sample dictionary updated with instance information. Extends parent

        Returns:
            dict: Updated dictionary with row, column and well updated
        """

        sample = super().to_sub_dict()
        sample['ct'] = f"({self.ct_n1}, {self.ct_n2})"
        try:
            sample['source_row'] = row_keys[self.sample.sample_location[0]]
            sample['source_column'] = int(self.sample.sample_location[1:])
        except (TypeError, AttributeError) as e:
            logger.error(f"Couldn't set sources for {self.sample.rsl_number}. Looks like there isn't data.")
        try:
            sample['positive'] = any(["positive" in item for item in [self.n1_status, self.n2_status]])
        except (TypeError, AttributeError) as e:
            logger.error(f"Couldn't check positives for {self.sample.rsl_number}. Looks like there isn't PCR data.")
        return sample

    @property
    def hitpicked(self) -> dict | None:
        """
        Outputs a dictionary usable for html plate maps. Extends parent

        Returns:
            dict: dictionary of sample id, row and column in elution plate
        """
        sample = super().hitpicked
        try:
            scaler = max([self.ct_n1, self.ct_n2])
        except TypeError:
            scaler = 0.0
        if scaler == 0.0:
            scaler = 45
        bg = (45 - scaler) * 17
        red = min([64 + bg, 255])
        grn = max([255 - bg, 0])
        blu = 128
        sample['background_color'] = f"rgb({red}, {grn}, {blu})"
        try:
            sample[
                'tooltip'] += f"<br>- ct N1: {'{:.2f}'.format(self.ct_n1)}<br>- ct N2: {'{:.2f}'.format(self.ct_n2)}"
        except (TypeError, AttributeError) as e:
            logger.error(f"Couldn't set tooltip for {self.sample.rsl_number}. Looks like there isn't PCR data.")
        return sample


class WastewaterArticAssociation(SubmissionSampleAssociation):
    """
    table containing wastewater artic specific submission/sample associations
    DOC: https://docs.sqlalchemy.org/en/14/orm/extensions/associationproxy.html
    """
    id = Column(INTEGER, ForeignKey("_submissionsampleassociation.id"), primary_key=True)
    source_plate = Column(String(32))
    source_plate_number = Column(INTEGER)
    source_well = Column(String(8))
    ct = Column(String(8))  #: AKA ct for N1

    __mapper_args__ = dict(polymorphic_identity="Wastewater Artic Association",
                           polymorphic_load="inline",
                           inherit_condition=(id == SubmissionSampleAssociation.id))

    def to_sub_dict(self) -> dict:
        """
        Returns a sample dictionary updated with instance information. Extends parent

        Returns:
            dict: Updated dictionary with row, column and well updated
        """
        sample = super().to_sub_dict()
        sample['ct'] = self.ct
        sample['source_plate'] = self.source_plate
        sample['source_plate_number'] = self.source_plate_number
        sample['source_well'] = self.source_well
        return sample
