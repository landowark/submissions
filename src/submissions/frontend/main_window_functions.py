'''
contains operations used by multiple widgets.
'''
from datetime import date
import difflib
from getpass import getuser
import inspect
import pprint
import yaml
import json
from typing import Tuple, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
import pandas as pd
from backend.db.models import *
import logging
from PyQt6.QtWidgets import (
    QMainWindow, QLabel, QWidget, QPushButton, 
    QLineEdit, QComboBox, QDateEdit
)
from .all_window_functions import extract_form_info, select_open_file, select_save_file
from PyQt6.QtCore import QSignalBlocker
from backend.db.functions import (
    construct_submission_info, lookup_reagents, construct_kit_from_yaml, construct_org_from_yaml, get_control_subtypes,
    update_subsampassoc_with_pcr, check_kit_integrity, update_last_used, lookup_organizations, lookup_kit_types, 
    lookup_submissions, lookup_controls, lookup_samples, lookup_submission_sample_association, store_object
)
from backend.excel.parser import SheetParser, PCRParser, SampleParser
from backend.excel.reports import make_report_html, make_report_xlsx, convert_data_list_to_df
from tools import check_not_nan, convert_well_to_row_column
from .custom_widgets.pop_ups import AlertPop, QuestionAsker
from .custom_widgets import ReportDatePicker
from .custom_widgets.misc import ImportReagent, ParsedQLabel
from .visualizations.control_charts import create_charts, construct_html

logger = logging.getLogger(f"submissions.{__name__}")

def import_submission_function(obj:QMainWindow) -> Tuple[QMainWindow, dict|None]:
    """
    Import a new submission to the app window

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict|None]: Collection of new main app window and result dict
    """    
    logger.debug(f"\n\nStarting Import...\n\n")
    result = None
    logger.debug(obj.ctx)
    # initialize samples
    obj.samples = []
    
    obj.missing_info = []
    
    # set file dialog
    fname = select_open_file(obj, file_extension="xlsx")
    logger.debug(f"Attempting to parse file: {fname}")
    if not fname.exists():
        result = dict(message=f"File {fname.__str__()} not found.", status="critical")
        return obj, result
    # create sheetparser using excel sheet and context from gui
    try:
        obj.prsr = SheetParser(ctx=obj.ctx, filepath=fname)
    except PermissionError:
        logger.error(f"Couldn't get permission to access file: {fname}")
        return obj, result
    try:
        logger.debug(f"Submission dictionary:\n{pprint.pformat(obj.prsr.sub)}")
        pyd = obj.prsr.to_pydantic()
        logger.debug(f"Pydantic result: \n\n{pprint.pformat(pyd)}\n\n")
    except Exception as e:
        return obj, dict(message= f"Problem creating pydantic model:\n\n{e}", status="critical")
    # destroy any widgets from previous imports
    for item in obj.table_widget.formlayout.parentWidget().findChildren(QWidget):
        item.setParent(None)
    obj.current_submission_type = pyd.submission_type['value']
    # Get list of fields from pydantic model.
    fields = list(pyd.model_fields.keys()) + list(pyd.model_extra.keys())
    fields.remove('filepath')
    logger.debug(f"pydantic fields: {fields}")
    for field in fields:
        value = getattr(pyd, field)
        logger.debug(f"Checking: {field}: {value}")
        # Get from pydantic model whether field was completed in the form
        if isinstance(value, dict) and field != 'ctx':
            logger.debug(f"The field {field} is a dictionary: {value}")
            if not value['parsed']:
                obj.missing_info.append(field)
        label = ParsedQLabel(value, field)
        match field:
            case 'submitting_lab':
                logger.debug(f"{field}: {value['value']}")
                # create combobox to hold looked up submitting labs
                add_widget = QComboBox()
                # labs = [item.__str__() for item in lookup_all_orgs(ctx=obj.ctx)]
                labs = [item.__str__() for item in lookup_organizations(ctx=obj.ctx)]
                # try to set closest match to top of list
                try:
                    labs = difflib.get_close_matches(value['value'], labs, len(labs), 0)
                except (TypeError, ValueError):
                    pass
                # set combobox values to lookedup values
                add_widget.addItems(labs)
            case 'extraction_kit':
                # if extraction kit not available, all other values fail
                if not check_not_nan(value['value']):
                    msg = AlertPop(message="Make sure to check your extraction kit in the excel sheet!", status="warning")
                    msg.exec()
                # create combobox to hold looked up kits
                add_widget = QComboBox()
                # lookup existing kits by 'submission_type' decided on by sheetparser
                logger.debug(f"Looking up kits used for {pyd.submission_type['value']}")
                # uses = [item.__str__() for item in lookup_kittype_by_use(ctx=obj.ctx, used_for=pyd.submission_type['value'])]
                uses = [item.__str__() for item in lookup_kit_types(ctx=obj.ctx, used_for=pyd.submission_type['value'])]
                logger.debug(f"Kits received for {pyd.submission_type['value']}: {uses}")
                if check_not_nan(value['value']):
                    logger.debug(f"The extraction kit in parser was: {value['value']}")
                    uses.insert(0, uses.pop(uses.index(value['value'])))
                    obj.ext_kit = value['value']
                else:
                    logger.error(f"Couldn't find {obj.prsr.sub['extraction_kit']}")
                    obj.ext_kit = uses[0]
                # Run reagent scraper whenever extraction kit is changed.
                add_widget.currentTextChanged.connect(obj.scrape_reagents)
                # add_widget.addItems(uses)
            case 'submitted_date':
                # uses base calendar
                add_widget = QDateEdit(calendarPopup=True)
                # sets submitted date based on date found in excel sheet
                try:
                    add_widget.setDate(value['value'])
                # if not found, use today
                except:
                    add_widget.setDate(date.today())
            case 'samples':
                # hold samples in 'obj' until form submitted
                logger.debug(f"{field}:\n\t{value}")
                obj.samples = value
                continue
            case "ctx":
                continue
            case 'reagents':
                # NOTE: This is now set to run when the extraction kit is updated.
                continue
            case _:
                # anything else gets added in as a line edit
                add_widget = QLineEdit()
                logger.debug(f"Setting widget text to {str(value['value']).replace('_', ' ')}")
                add_widget.setText(str(value['value']).replace("_", " "))
        try:
            add_widget.setObjectName(field)
            logger.debug(f"Widget name set to: {add_widget.objectName()}")
            obj.table_widget.formlayout.addWidget(label)
            obj.table_widget.formlayout.addWidget(add_widget)
        except AttributeError as e:
            logger.error(e)
    kit_widget = obj.table_widget.formlayout.parentWidget().findChild(QComboBox, 'extraction_kit')
    kit_widget.addItems(uses)
    # compare obj.reagents with expected reagents in kit
    if obj.prsr.sample_result != None:
        msg = AlertPop(message=obj.prsr.sample_result, status="WARNING")
        msg.exec()
    logger.debug(f"Pydantic extra fields: {pyd.model_extra}")
    if "csv" in pyd.model_extra:
        obj.csv = pyd.model_extra['csv']
    logger.debug(f"All attributes of obj:\n{pprint.pformat(obj.__dict__)}")
    return obj, result

def kit_reload_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Reload the fields in the form

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    for item in obj.table_widget.formlayout.parentWidget().findChildren(QWidget):
        if isinstance(item, QLabel):
            if item.text().startswith("Lot"):
                item.setParent(None)
        else:
            logger.debug(f"Type of {item.objectName()} is {type(item)}")
            if item.objectName().startswith("lot_"):
                item.setParent(None)
    obj.kit_integrity_completion_function()
    return obj, result

def kit_integrity_completion_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Compare kit contents to parsed contents

    Args:
        obj (QMainWindow): The original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    logger.debug(inspect.currentframe().f_back.f_code.co_name)
    # find the widget that contains kit info
    kit_widget = obj.table_widget.formlayout.parentWidget().findChild(QComboBox, 'extraction_kit')
    logger.debug(f"Kit selector: {kit_widget}")
    # get current kit being used
    obj.ext_kit = kit_widget.currentText()
    for item in obj.reagents:
        obj.table_widget.formlayout.addWidget(ParsedQLabel({'parsed':True}, item.type, title=False, label_name=f"lot_{item.type}_label"))
        reagent = dict(type=item.type, lot=item.lot, exp=item.exp, name=item.name)
        add_widget = ImportReagent(ctx=obj.ctx, reagent=reagent, extraction_kit=obj.ext_kit)
        obj.table_widget.formlayout.addWidget(add_widget)
    logger.debug(f"Checking integrity of {obj.ext_kit}")
    # see if there are any missing reagents
    if len(obj.missing_reagents) > 0:
        result = dict(message=f"The submission you are importing is missing some reagents expected by the kit.\n\nIt looks like you are missing: {[item.type.upper() for item in obj.missing_reagents]}\n\nAlternatively, you may have set the wrong extraction kit.\n\nThe program will populate lists using existing reagents.\n\nPlease make sure you check the lots carefully!", status="Warning")
    for item in obj.missing_reagents:
        # Add label that has parsed as False to show "MISSING" label.
        obj.table_widget.formlayout.addWidget(ParsedQLabel({'parsed':False}, item.type, title=False, label_name=f"missing_{item.type}_label"))
        # Set default parameters for the empty reagent.
        reagent = dict(type=item.type, lot=None, exp=date.today(), name=None)
        # create and add widget
        # add_widget = ImportReagent(ctx=obj.ctx, reagent=PydReagent(**reagent), extraction_kit=obj.ext_kit)
        add_widget = ImportReagent(ctx=obj.ctx, reagent=reagent, extraction_kit=obj.ext_kit)
        obj.table_widget.formlayout.addWidget(add_widget)
    # Add submit button to the form.
    submit_btn = QPushButton("Submit")
    submit_btn.setObjectName("lot_submit_btn")
    obj.table_widget.formlayout.addWidget(submit_btn)
    submit_btn.clicked.connect(obj.submit_new_sample)
    return obj, result

def submit_new_sample_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Parse forms and add sample to the database.

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    logger.debug(f"\n\nBeginning Submission\n\n")
    result = None
    # extract info from the form widgets
    info = extract_form_info(obj.table_widget.tab1)
    # seperate out reagents
    reagents = {k.replace("lot_", ""):v for k,v in info.items() if k.startswith("lot_")}
    info = {k:v for k,v in info.items() if not k.startswith("lot_")}
    logger.debug(f"Info: {info}")
    logger.debug(f"Reagents: {reagents}")
    parsed_reagents = []
    # compare reagents in form to reagent database
    for reagent in reagents:
        # Lookup any existing reagent of this type with this lot number
        wanted_reagent = lookup_reagents(ctx=obj.ctx, lot_number=reagents[reagent], reagent_type=reagent)
        logger.debug(f"Looked up reagent: {wanted_reagent}")
        # if reagent not found offer to add to database
        if wanted_reagent == None:
            r_lot = reagents[reagent]
            dlg = QuestionAsker(title=f"Add {r_lot}?", message=f"Couldn't find reagent type {reagent.strip('Lot')}: {r_lot} in the database.\n\nWould you like to add it?")
            if dlg.exec():
                logger.debug(f"Looking through {pprint.pformat(obj.reagents)} for reagent {reagent}")
                try:
                    picked_reagent = [item for item in obj.reagents if item.type == reagent][0]
                except IndexError:
                    logger.error(f"Couldn't find {reagent} in obj.reagents. Checking missing reagents {pprint.pformat(obj.missing_reagents)}")
                    picked_reagent = [item for item in obj.missing_reagents if item.type == reagent][0]
                logger.debug(f"checking reagent: {reagent} in obj.reagents. Result: {picked_reagent}")
                expiry_date = picked_reagent.exp
                wanted_reagent = obj.add_reagent(reagent_lot=r_lot, reagent_type=reagent.replace("lot_", ""), expiry=expiry_date, name=picked_reagent.name)
            else:
                # In this case we will have an empty reagent and the submission will fail kit integrity check
                logger.debug("Will not add reagent.")
                return obj, dict(message="Failed integrity check", status="critical")
        parsed_reagents.append(wanted_reagent)
    # move samples into preliminary submission dict
    info['samples'] = obj.samples
    info['uploaded_by'] = getuser()
    # construct submission object
    logger.debug(f"Here is the info_dict: {pprint.pformat(info)}")
    base_submission, result = construct_submission_info(ctx=obj.ctx, info_dict=info)
    # check output message for issues
    match result['code']:
        # code 1: ask for overwrite
        case 1:
            dlg = QuestionAsker(title=f"Review {base_submission.rsl_plate_num}?", message=result['message'])
            if dlg.exec():
                # Do not add duplicate reagents.
                base_submission.reagents = []
            else:
                obj.ctx.database_session.rollback()
                return obj, dict(message="Overwrite cancelled", status="Information")
        # code 2: No RSL plate number given
        case 2:
            return obj, dict(message=result['message'], status='critical')
        case _:
            pass
    # add reagents to submission object
    for reagent in parsed_reagents:
        base_submission.reagents.append(reagent)
        update_last_used(ctx=obj.ctx, reagent=reagent, kit=base_submission.extraction_kit)
    logger.debug(f"Parsed reagents: {pprint.pformat(parsed_reagents)}")
    logger.debug("Checking kit integrity...")
    kit_integrity = check_kit_integrity(base_submission)
    if kit_integrity != None:
        return obj, dict(message=kit_integrity['message'], status="critical")
    logger.debug(f"Sending submission: {base_submission.rsl_plate_num} to database.")
    result = store_object(ctx=obj.ctx, object=base_submission)
    # update summary sheet
    obj.table_widget.sub_wid.setData()
    # reset form
    for item in obj.table_widget.formlayout.parentWidget().findChildren(QWidget):
        item.setParent(None)
    logger.debug(f"All attributes of obj: {pprint.pformat(obj.__dict__)}")
    if len(obj.missing_reagents + obj.missing_info) > 0:
        logger.debug(f"We have blank reagents in the excel sheet.\n\tLet's try to fill them in.") 
        extraction_kit = lookup_kit_types(ctx=obj.ctx, name=obj.ext_kit)
        logger.debug(f"We have the extraction kit: {extraction_kit.name}")
        excel_map = extraction_kit.construct_xl_map_for_use(obj.current_submission_type)
        logger.debug(f"Extraction kit map:\n\n{pprint.pformat(excel_map)}")
        input_reagents = [item.to_reagent_dict(extraction_kit=base_submission.extraction_kit) for item in parsed_reagents]
        logger.debug(f"Parsed reagents going into autofile: {pprint.pformat(input_reagents)}")
        autofill_excel(obj=obj, xl_map=excel_map, reagents=input_reagents, missing_reagents=obj.missing_reagents, info=info, missing_info=obj.missing_info)
    if hasattr(obj, 'csv'):
        dlg = QuestionAsker("Export CSV?", "Would you like to export the csv file?")
        if dlg.exec():
            fname = select_save_file(obj, f"{base_submission.rsl_plate_num}.csv", extension="csv")
            try:
                obj.csv.to_csv(fname.__str__(), index=False)
            except PermissionError:
                logger.debug(f"Could not get permissions to {fname}. Possibly the request was cancelled.")
    try:
        delattr(obj, "csv")
    except AttributeError:
        pass
    return obj, result

def generate_report_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Generate a summary of activities for a time period

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    # ask for date ranges
    dlg = ReportDatePicker()
    if dlg.exec():
        info = extract_form_info(dlg)
        logger.debug(f"Report info: {info}")
        # find submissions based on date range
        # subs = lookup_submissions_by_date_range(ctx=obj.ctx, start_date=info['start_date'], end_date=info['end_date'])
        subs = lookup_submissions(ctx=obj.ctx, start_date=info['start_date'], end_date=info['end_date'])
        # convert each object to dict
        records = [item.report_dict() for item in subs]
        # make dataframe from record dictionaries
        detailed_df, summary_df = make_report_xlsx(records=records)
        html = make_report_html(df=summary_df, start_date=info['start_date'], end_date=info['end_date'])
        # get save location of report
        fname = select_save_file(obj=obj, default_name=f"Submissions_Report_{info['start_date']}-{info['end_date']}.pdf", extension="pdf")
        # logger.debug(f"report output name: {fname}")
        with open(fname, "w+b") as f:
            pisa.CreatePDF(html, dest=f)
        writer = pd.ExcelWriter(fname.with_suffix(".xlsx"), engine='openpyxl')
        summary_df.to_excel(writer, sheet_name="Report")
        detailed_df.to_excel(writer, sheet_name="Details", index=False)
        worksheet = writer.sheets['Report']
        for idx, col in enumerate(summary_df):  # loop through all columns
            series = summary_df[col]
            max_len = max((
                series.astype(str).map(len).max(),  # len of largest item
                len(str(series.name))  # len of column name/header
                )) + 20  # adding a little extra space
            try:
                worksheet.column_dimensions[get_column_letter(idx)].width = max_len 
            except ValueError:
                pass
        for cell in worksheet['D']:
            if cell.row > 1:
                cell.style = 'Currency'
        writer.close()
    return obj, result

def add_kit_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Add a new kit to the database.

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    # setup file dialog to find yaml file
    fname = select_open_file(obj, file_extension="yml")
    assert fname.exists()
    # read yaml file
    try:
        with open(fname.__str__(), "r") as stream:
            try:
                exp = yaml.load(stream, Loader=yaml.Loader)
            except yaml.YAMLError as exc:
                logger.error(f'Error reading yaml file {fname}: {exc}')
                return {}
    except PermissionError:
        return
    # send to kit creator function
    result = construct_kit_from_yaml(ctx=obj.ctx, exp=exp)
    return obj, result

def add_org_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Add a new organization to the database.

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    # setup file dialog to find yaml flie
    fname = select_open_file(obj, extension="yml")
    assert fname.exists()
    # read yaml file
    try:
        with open(fname.__str__(), "r") as stream:
            try:
                org = yaml.load(stream, Loader=yaml.Loader)
            except yaml.YAMLError as exc:
                logger.error(f'Error reading yaml file {fname}: {exc}')
                return obj, dict(message=f"There was a problem reading yaml file {fname.__str__()}", status="critical")
    except PermissionError:
        return obj, result
    # send to kit creator function
    result = construct_org_from_yaml(ctx=obj.ctx, org=org)
    return obj, result

def controls_getter_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Get controls based on start/end dates

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    # subtype defaults to disabled  
    try:
        obj.table_widget.sub_typer.disconnect()
    except TypeError:
        pass
    # correct start date being more recent than end date and rerun
    if obj.table_widget.datepicker.start_date.date() > obj.table_widget.datepicker.end_date.date():
        logger.warning("Start date after end date is not allowed!")
        threemonthsago = obj.table_widget.datepicker.end_date.date().addDays(-60)
        # block signal that will rerun controls getter and set start date
        # Without triggering this function again
        with QSignalBlocker(obj.table_widget.datepicker.start_date) as blocker:
            obj.table_widget.datepicker.start_date.setDate(threemonthsago)
        obj._controls_getter()
        return obj, result
    # convert to python useable date objects
    obj.start_date = obj.table_widget.datepicker.start_date.date().toPyDate()
    obj.end_date = obj.table_widget.datepicker.end_date.date().toPyDate()
    obj.con_type = obj.table_widget.control_typer.currentText()
    obj.mode = obj.table_widget.mode_typer.currentText()
    obj.table_widget.sub_typer.clear()
    # lookup subtypes
    sub_types = get_control_subtypes(ctx=obj.ctx, type=obj.con_type, mode=obj.mode)
    # sub_types = lookup_controls(ctx=obj.ctx, control_type=obj.con_type)
    if sub_types != []:
        # block signal that will rerun controls getter and update sub_typer
        with QSignalBlocker(obj.table_widget.sub_typer) as blocker: 
            obj.table_widget.sub_typer.addItems(sub_types)
        obj.table_widget.sub_typer.setEnabled(True)
        obj.table_widget.sub_typer.currentTextChanged.connect(obj._chart_maker)
    else:
        obj.table_widget.sub_typer.clear()
        obj.table_widget.sub_typer.setEnabled(False)
    obj._chart_maker()
    return obj, result

def chart_maker_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Create html chart for controls reporting

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    logger.debug(f"Control getter context: \n\tControl type: {obj.con_type}\n\tMode: {obj.mode}\n\tStart Date: {obj.start_date}\n\tEnd Date: {obj.end_date}")
    # set the subtype for kraken
    if obj.table_widget.sub_typer.currentText() == "":
        obj.subtype = None
    else:
        obj.subtype = obj.table_widget.sub_typer.currentText()
    logger.debug(f"Subtype: {obj.subtype}")
    # query all controls using the type/start and end dates from the gui
    # controls = get_all_controls_by_type(ctx=obj.ctx, con_type=obj.con_type, start_date=obj.start_date, end_date=obj.end_date)
    controls = lookup_controls(ctx=obj.ctx, control_type=obj.con_type, start_date=obj.start_date, end_date=obj.end_date)
    # if no data found from query set fig to none for reporting in webview
    if controls == None:
        fig = None
    else:
        # change each control to list of dictionaries
        data = [control.convert_by_mode(mode=obj.mode) for control in controls]
        # flatten data to one dimensional list
        data = [item for sublist in data for item in sublist]
        logger.debug(f"Control objects going into df conversion: {type(data)}")
        if data == []:
            return obj, dict(status="Critical", message="No data found for controls in given date range.")
        # send to dataframe creator
        df = convert_data_list_to_df(ctx=obj.ctx, input=data, subtype=obj.subtype)
        if obj.subtype == None:
            title = obj.mode
        else:
            title = f"{obj.mode} - {obj.subtype}"
        # send dataframe to chart maker
        fig = create_charts(ctx=obj.ctx, df=df, ytitle=title)
    logger.debug(f"Updating figure...")
    # construct html for webview
    html = construct_html(figure=fig)
    logger.debug(f"The length of html code is: {len(html)}")
    obj.table_widget.webengineview.setHtml(html)
    obj.table_widget.webengineview.update()
    logger.debug("Figure updated... I hope.")
    return obj, result

def link_controls_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Link scraped controls to imported submissions.

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    # all_bcs = lookup_all_submissions_by_type(obj.ctx, "Bacterial Culture")
    all_bcs = lookup_submissions(ctx=obj.ctx, submission_type="Bacterial Culture")
    logger.debug(all_bcs)
    # all_controls = get_all_controls(obj.ctx)
    all_controls = lookup_controls(ctx=obj.ctx)
    ac_list = [control.name for control in all_controls]
    count = 0
    for bcs in all_bcs:
        logger.debug(f"Running for {bcs.rsl_plate_num}")
        logger.debug(f"Here is the current control: {[control.name for control in bcs.controls]}")
        samples = [sample.submitter_id for sample in bcs.samples]
        logger.debug(bcs.controls)
        for sample in samples:
            # replace below is a stopgap method because some dingus decided to add spaces in some of the ATCC49... so it looks like "ATCC 49"...
            if " " in sample:
                logger.warning(f"There is not supposed to be a space in the sample name!!!")
                sample = sample.replace(" ", "")
            if not any([ac.startswith(sample) for ac in ac_list]):
                continue
            else:
                for control in all_controls:
                    diff = difflib.SequenceMatcher(a=sample, b=control.name).ratio()
                    if control.name.startswith(sample):
                        logger.debug(f"Checking {sample} against {control.name}... {diff}")
                        logger.debug(f"Found match:\n\tSample: {sample}\n\tControl: {control.name}\n\tDifference: {diff}")
                        if control in bcs.controls:
                            logger.debug(f"{control.name} already in {bcs.rsl_plate_num}, skipping")
                            continue
                        else:
                            logger.debug(f"Adding {control.name} to {bcs.rsl_plate_num} as control")
                            bcs.controls.append(control)
                            control.submission = bcs
                            control.submission_id = bcs.id
                            obj.ctx.database_session.add(control)
                            count += 1
        obj.ctx.database_session.add(bcs)
        logger.debug(f"Here is the new control: {[control.name for control in bcs.controls]}")
    result = dict(message=f"We added {count} controls to bacterial cultures.", status="information")
    logger.debug(result)
    obj.ctx.database_session.commit()
    return obj, result

def link_extractions_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Link extractions from runlogs to imported submissions

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    fname = select_open_file(obj, file_extension="csv")
    with open(fname.__str__(), 'r') as f:
        # split csv on commas
        runs = [col.strip().split(",") for col in f.readlines()]
    count = 0
    for run in runs:
        new_run = dict(
                start_time=run[0].strip(), 
                rsl_plate_num=run[1].strip(), 
                sample_count=run[2].strip(), 
                status=run[3].strip(),
                experiment_name=run[4].strip(),
                end_time=run[5].strip()
            )
        # elution columns are item 6 in the comma split list to the end
        for ii in range(6, len(run)):
            new_run[f"column{str(ii-5)}_vol"] = run[ii]
        # Lookup imported submissions
        # sub = lookup_submission_by_rsl_num(ctx=obj.ctx, rsl_num=new_run['rsl_plate_num'])
        sub = lookup_submissions(ctx=obj.ctx, rsl_number=new_run['rsl_plate_num'])
        # If no such submission exists, move onto the next run
        try:
            logger.debug(f"Found submission: {sub.rsl_plate_num}")
            count += 1
        except AttributeError:
            continue
        if sub.extraction_info != None:
            existing = json.loads(sub.extraction_info)
        else:
            existing = None
        # Check if the new info already exists in the imported submission
        try:
            if json.dumps(new_run) in sub.extraction_info:
                logger.debug(f"Looks like we already have that info.")
                continue
        except TypeError:
            pass
        # Update or create the extraction info
        if existing != None:
            try:
                logger.debug(f"Updating {type(existing)}: {existing} with {type(new_run)}: {new_run}")
                existing.append(new_run)
                logger.debug(f"Setting: {existing}")
                sub.extraction_info = json.dumps(existing)
            except TypeError:
                logger.error(f"Error updating!")
                sub.extraction_info = json.dumps([new_run])
            logger.debug(f"Final ext info for {sub.rsl_plate_num}: {sub.extraction_info}")
        else:
            sub.extraction_info = json.dumps([new_run])        
        obj.ctx.database_session.add(sub)
        obj.ctx.database_session.commit()
    result = dict(message=f"We added {count} logs to the database.", status='information') 
    return obj, result

def link_pcr_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Link PCR data from run logs to an imported submission

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    fname = select_open_file(obj, file_extension="csv")
    with open(fname.__str__(), 'r') as f:
        # split csv rows on comma
        runs = [col.strip().split(",") for col in f.readlines()]
    count = 0
    for run in runs:
        new_run = dict(
                start_time=run[0].strip(), 
                rsl_plate_num=run[1].strip(), 
                biomek_status=run[2].strip(), 
                quant_status=run[3].strip(),
                experiment_name=run[4].strip(),
                end_time=run[5].strip()
            )
        # lookup imported submission
        # sub = lookup_submission_by_rsl_num(ctx=obj.ctx, rsl_num=new_run['rsl_plate_num'])
        sub = lookup_submissions(ctx=obj.ctx, rsl_number=new_run['rsl_plate_num'])
        # if imported submission doesn't exist move on to next run
        try:
            logger.debug(f"Found submission: {sub.rsl_plate_num}")
        except AttributeError:
            continue
        # check if pcr_info already exists
        if hasattr(sub, 'pcr_info') and sub.pcr_info != None:
            existing = json.loads(sub.pcr_info)
        else:
            existing = None
        # check if this entry already exists in imported submission
        try:
            if json.dumps(new_run) in sub.pcr_info:
                logger.debug(f"Looks like we already have that info.")
                continue
            else:
                count += 1
        except TypeError:
            logger.error(f"No json to dump")
        if existing != None:
            try:
                logger.debug(f"Updating {type(existing)}: {existing} with {type(new_run)}: {new_run}")
                existing.append(new_run)
                logger.debug(f"Setting: {existing}")
                sub.pcr_info = json.dumps(existing)
            except TypeError:
                logger.error(f"Error updating!")
                sub.pcr_info = json.dumps([new_run])
            logger.debug(f"Final ext info for {sub.rsl_plate_num}: {sub.pcr_info}")
        else:
            sub.pcr_info = json.dumps([new_run])        
        obj.ctx.database_session.add(sub)
        obj.ctx.database_session.commit()
    result = dict(message=f"We added {count} logs to the database.", status='information')
    return obj, result

def import_pcr_results_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Import Quant-studio PCR data to an imported submission

    Args:
        obj (QMainWindow): original app window

    Returns:
        Tuple[QMainWindow, dict]: Collection of new main app window and result dict
    """    
    result = None
    fname = select_open_file(obj, file_extension="xlsx")
    parser = PCRParser(ctx=obj.ctx, filepath=fname)
    logger.debug(f"Attempting lookup for {parser.plate_num}")
    # sub = lookup_submission_by_rsl_num(ctx=obj.ctx, rsl_num=parser.plate_num)
    sub = lookup_submissions(ctx=obj.ctx, rsl_number=parser.plate_num)
    try:
        logger.debug(f"Found submission: {sub.rsl_plate_num}")
    except AttributeError:
        # If no plate is found, may be because this is a repeat. Lop off the '-1' or '-2' and repeat
        logger.error(f"Submission of number {parser.plate_num} not found. Attempting rescue of plate repeat.")
        parser.plate_num = "-".join(parser.plate_num.split("-")[:-1])
        # sub = lookup_submission_by_rsl_num(ctx=obj.ctx, rsl_num=parser.plate_num)
        sub = lookup_submissions(ctx=obj.ctx, rsl_number=parser.plate_num)
        try:
            logger.debug(f"Found submission: {sub.rsl_plate_num}")
        except AttributeError:
            logger.error(f"Rescue of {parser.plate_num} failed.")
            return obj, dict(message="Couldn't find a submission with that RSL number.", status="warning")
    # Check if PCR info already exists
    if hasattr(sub, 'pcr_info') and sub.pcr_info != None:
        existing = json.loads(sub.pcr_info)
    else:
        existing = None
    if existing != None:
        # update pcr_info
        try:
            logger.debug(f"Updating {type(existing)}: {existing} with {type(parser.pcr)}: {parser.pcr}")
            if json.dumps(parser.pcr) not in sub.pcr_info:
                existing.append(parser.pcr)
            logger.debug(f"Setting: {existing}")
            sub.pcr_info = json.dumps(existing)
        except TypeError:
            logger.error(f"Error updating!")
            sub.pcr_info = json.dumps([parser.pcr])
        logger.debug(f"Final pcr info for {sub.rsl_plate_num}: {sub.pcr_info}")
    else:
        sub.pcr_info = json.dumps([parser.pcr])
    obj.ctx.database_session.add(sub)
    logger.debug(f"Existing {type(sub.pcr_info)}: {sub.pcr_info}")
    logger.debug(f"Inserting {type(json.dumps(parser.pcr))}: {json.dumps(parser.pcr)}")
    obj.ctx.database_session.commit()
    logger.debug(f"Got {len(parser.samples)} samples to update!")
    logger.debug(f"Parser samples: {parser.samples}")
    for sample in sub.samples:
        logger.debug(f"Running update on: {sample}")
        try:
            sample_dict = [item for item in parser.samples if item['sample']==sample.rsl_number][0]
        except IndexError:
            continue
        update_subsampassoc_with_pcr(ctx=obj.ctx, submission=sub, sample=sample, input_dict=sample_dict)

    result = dict(message=f"We added PCR info to {sub.rsl_plate_num}.", status='information')
    return obj, result

def autofill_excel(obj:QMainWindow, xl_map:dict, reagents:List[dict], missing_reagents:List[str], info:dict, missing_info:List[str]):
    """
    Automatically fills in excel cells with submission info.

    Args:
        obj (QMainWindow): Original main app window
        xl_map (dict): Map of where each item goes in the excel workbook.
        reagents (List[dict]): All reagents placed in the submission form.
        missing_reagents (List[str]): Reagents that are required for the kit that were not present.
        info (dict): Dictionary of misc info from submission
        missing_info (List[str]): Plate info missing from the excel sheet.
    """    
    # logger.debug(reagents)

    logger.debug(f"Here is the info dict coming in:\n{pprint.pformat(info)}")
    logger.debug(f"Here are the missing reagents:\n{missing_reagents}")
    logger.debug(f"Here are the missing info:\n{missing_info}")
    logger.debug(f"Here is the xl map: {pprint.pformat(xl_map)}")
    # pare down the xl map to only the missing data.
    relevant_reagent_map = {k:v for k,v in xl_map.items() if k in [reagent.type for reagent in missing_reagents]}
    # pare down reagents to only what's missing
    logger.debug(f"Checking {[item['type'] for item in reagents]} against {[reagent.type for reagent in missing_reagents]}")
    relevant_reagents = [item for item in reagents if item['type'] in [reagent.type for reagent in missing_reagents]]
    logger.debug(f"Here are the relevant reagents: {pprint.pformat(relevant_reagents)}")
    # hacky manipulation of submission type so it looks better.
    # pare down info to just what's missing
    relevant_info_map = {k:v for k,v in xl_map['info'].items() if k in missing_info and k != 'samples'}
    relevant_info = {k:v for k,v in info.items() if k in missing_info}
    logger.debug(f"Here is the relevant info: {pprint.pformat(relevant_info)}")
    # construct new objects to put into excel sheets:
    new_reagents = []
    logger.debug(f"Parsing from relevant reagent map: {pprint.pformat(relevant_reagent_map)}")
    for reagent in relevant_reagents:
        new_reagent = {}
        new_reagent['type'] = reagent['type']
        new_reagent['lot'] = relevant_reagent_map[new_reagent['type']]['lot']
        new_reagent['lot']['value'] = reagent['lot']
        new_reagent['expiry'] = relevant_reagent_map[new_reagent['type']]['expiry']
        new_reagent['expiry']['value'] = reagent['expiry']
        new_reagent['sheet'] = relevant_reagent_map[new_reagent['type']]['sheet']
        # name is only present for Bacterial Culture
        try:
            new_reagent['name'] = relevant_reagent_map[new_reagent['type']]['name']
            new_reagent['name']['value'] = reagent['name']
        except Exception as e:
            logger.error(f"Couldn't get name due to {e}")
        new_reagents.append(new_reagent)
    # construct new info objects to put into excel sheets
    new_info = []
    logger.debug(f"Parsing from relevant info map: {pprint.pformat(relevant_info_map)}")
    for item in relevant_info:
        new_item = {}
        new_item['type'] = item
        new_item['location'] = relevant_info_map[item]
        new_item['value'] = relevant_info[item]
        new_info.append(new_item)
    logger.debug(f"New reagents: {new_reagents}")
    logger.debug(f"New info: {new_info}")
    # open a new workbook using openpyxl
    workbook = load_workbook(obj.prsr.xl.io)
    # get list of sheet names
    sheets = workbook.sheetnames
    # logger.debug(workbook.sheetnames)
    for sheet in sheets:
        # open sheet
        worksheet=workbook[sheet]
        # Get relevant reagents for that sheet
        sheet_reagents = [item for item in new_reagents if sheet in item['sheet']]
        for reagent in sheet_reagents:
            logger.debug(f"Attempting to write lot {reagent['lot']['value']} in: row {reagent['lot']['row']}, column {reagent['lot']['column']}")
            worksheet.cell(row=reagent['lot']['row'], column=reagent['lot']['column'], value=reagent['lot']['value'])
            logger.debug(f"Attempting to write expiry {reagent['expiry']['value']} in: row {reagent['expiry']['row']}, column {reagent['expiry']['column']}")
            worksheet.cell(row=reagent['expiry']['row'], column=reagent['expiry']['column'], value=reagent['expiry']['value'])
            try:
                logger.debug(f"Attempting to write name {reagent['name']['value']} in: row {reagent['name']['row']}, column {reagent['name']['column']}")
                worksheet.cell(row=reagent['name']['row'], column=reagent['name']['column'], value=reagent['name']['value'])
            except Exception as e:
                logger.error(f"Could not write name {reagent['name']['value']} due to {e}")
        # Get relevant info for that sheet
        sheet_info = [item for item in new_info if sheet in item['location']['sheets']]
        for item in sheet_info:
            logger.debug(f"Attempting: {item['type']}")
            worksheet.cell(row=item['location']['row'], column=item['location']['column'], value=item['value'])
        # Hacky way to pop in 'signed by'
        if info['submission_type'] == "Bacterial Culture":
            workbook["Sample List"].cell(row=14, column=2, value=getuser()[0:2].upper())
    fname = select_save_file(obj=obj, default_name=info['rsl_plate_num'], extension="xlsx")
    workbook.save(filename=fname.__str__())

def construct_first_strand_function(obj:QMainWindow) -> Tuple[QMainWindow, dict]:
    """
    Generates a csv file from client submitted xlsx file.

    Args:
        obj (QMainWindow): Main application

    Returns:
        Tuple[QMainWindow, dict]: Updated main application and result
    """    
    def get_plates(input_sample_number:str, plates:list) -> Tuple[int, str]:
        logger.debug(f"Looking up {input_sample_number} in {plates}")
        # samp = lookup_ww_sample_by_processing_number(ctx=obj.ctx, processing_number=input_sample_number)
        samp = lookup_samples(ctx=obj.ctx, ww_processing_num=input_sample_number)
        if samp ==  None:
            # samp = lookup_sample_by_submitter_id(ctx=obj.ctx, submitter_id=input_sample_number)
            samp = lookup_samples(ctx=obj.ctx, submitter_id=input_sample_number)
        logger.debug(f"Got sample: {samp}")
        # new_plates = [(iii+1, lookup_sub_samp_association_by_plate_sample(ctx=obj.ctx, rsl_sample_num=samp, rsl_plate_num=lookup_submission_by_rsl_num(ctx=obj.ctx, rsl_num=plate))) for iii, plate in enumerate(plates)]
        new_plates = [(iii+1, lookup_submission_sample_association(ctx=obj.ctx, sample=samp, submission=plate)) for iii, plate in enumerate(plates)]
        logger.debug(f"Associations: {pprint.pformat(new_plates)}")
        try:
            plate_num, plate = next(assoc for assoc in new_plates if assoc[1] is not None)
        except StopIteration:
            plate_num, plate = None, None
        logger.debug(f"Plate number {plate_num} is {plate}")
        return plate_num, plate
    fname = select_open_file(obj=obj, file_extension="xlsx")
    xl = pd.ExcelFile(fname)
    sprsr = SampleParser(ctx=obj.ctx, xl=xl, submission_type="First Strand")
    _, samples = sprsr.parse_samples(generate=False)
    plates = sprsr.grab_plates()
    output_samples = []
    logger.debug(f"Samples: {pprint.pformat(samples)}")
    old_plate_number = 1
    for item in samples:
        new_dict = {}
        new_dict['sample'] = item['submitter_id']
        if item['submitter_id'] == "NTC1":
            new_dict['destination_row'] = 8
            new_dict['destination_column'] = 2
            new_dict['plate_number'] = 'control'
        elif item['submitter_id'] == "NTC2":
            new_dict['destination_row'] = 8
            new_dict['destination_column'] = 5
            new_dict['plate_number'] = 'control'
        else:
            new_dict['destination_row'] = item['row']
            new_dict['destination_column'] = item['column']
        plate_num, plate = get_plates(input_sample_number=new_dict['sample'], plates=plates)
        if plate_num == None:
            plate_num = str(old_plate_number) + "*"
        else:
            old_plate_number = plate_num
        logger.debug(f"Got plate number: {plate_num}, plate: {plate}")
        if plate == None:
            try:
                new_dict['source_row'], new_dict['source_column'] = convert_well_to_row_column(item['well'])
                new_dict['plate_number'] = plate_num
            except KeyError:
                pass
        else:
            new_dict['plate_number'] = plate_num
            new_dict['plate'] = plate.submission.rsl_plate_num
            new_dict['source_row'] = plate.row
            new_dict['source_column'] = plate.column
        output_samples.append(new_dict)
    df = pd.DataFrame.from_records(output_samples)
    df.sort_values(by=['destination_column', 'destination_row'], ascending=True, inplace=True)
    columnsTitles = ['sample', 'destination_column', 'destination_row', 'plate_number', 'plate', "source_column", 'source_row']
    df = df.reindex(columns=columnsTitles)
    ofname = select_save_file(obj=obj, default_name=f"First Strand {date.today()}", extension="csv")
    df.to_csv(ofname, index=False)
    return obj, None

def scrape_reagents(obj:QMainWindow, extraction_kit:str) -> Tuple[QMainWindow, dict]:
    """
    Extracted scrape reagents function that will run when 
    form 'extraction_kit' widget is updated.

    Args:
        obj (QMainWindow): updated main application
        extraction_kit (str): name of extraction kit (in 'extraction_kit' widget)

    Returns:
        Tuple[QMainWindow, dict]: Updated application and result
    """    
    logger.debug("\n\nHello from reagent scraper!!\n\n")
    logger.debug(f"Extraction kit: {extraction_kit}")
    obj.reagents = []
    obj.missing_reagents = []
    [item.setParent(None) for item in obj.table_widget.formlayout.parentWidget().findChildren(QWidget) if item.objectName().startswith("lot_") or item.objectName().startswith("missing_")]
    reagents = obj.prsr.parse_reagents(extraction_kit=extraction_kit)
    logger.debug(f"Got reagents: {reagents}")
    for reagent in obj.prsr.sub['reagents']:
        # create label
        if reagent['parsed']:
            obj.reagents.append(reagent['value'])
        else:
            obj.missing_reagents.append(reagent['value'])
    logger.debug(f"Imported reagents: {obj.reagents}")
    logger.debug(f"Missing reagents: {obj.missing_reagents}")
    return obj, None
    


